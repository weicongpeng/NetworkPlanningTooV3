#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
合并版网络规划工具 - LTE/NR PCI规划、参数更新和邻区规划
包含GUI界面和命令行功能
"""

# First check license
try:
    import sys
    import json
    import os
    import hashlib
    import base64
    import hmac
    from datetime import datetime
    from cryptography.fernet import Fernet

    class LicenseManager:
        def __init__(self, license_path="license.dat", secret_key=None):
            # Use a hardcoded key (in production, this should be embedded securely in the compiled code)
            self.secret_key = secret_key or b'networkplanningtool_default_secret_key_for_encryption'
            self.license_path = license_path

        def check_license(self):
            """Check if license is valid and hasn't expired"""
            if not os.path.exists(self.license_path):
                return False, "License file not found. 请联系作者：weicongpeng1@163.com或15220958556"

            try:
                # Read the encrypted license file
                with open(self.license_path, 'rb') as f:
                    encrypted_content = f.read()

                # Parse the JSON
                try:
                    license_package = json.loads(encrypted_content.decode('utf-8'))
                except UnicodeDecodeError:
                    return False, "License file has been tampered with. 请联系作者：weicongpeng1@163.com或15220958556"

                # Extract encrypted data and signature
                try:
                    encrypted_data = base64.b64decode(license_package['encrypted_data'])
                    stored_signature = license_package['signature']
                except (KeyError, ValueError):
                    return False, "License file has been tampered with. 请联系作者：weicongpeng1@163.com或15220958556"

                # Verify the signature
                calculated_signature = hmac.new(self.secret_key, encrypted_data, hashlib.sha256).hexdigest()

                if not hmac.compare_digest(calculated_signature, stored_signature):
                    return False, "License file has been tampered with. 请联系作者：weicongpeng1@163.com或15220958556"

                # Decrypt the data
                derived_key = base64.urlsafe_b64encode(hashlib.sha256(self.secret_key).digest())
                fernet = Fernet(derived_key)
                try:
                    decrypted_data = fernet.decrypt(encrypted_data)
                except Exception:
                    return False, "License file has been tampered with. 请联系作者：weicongpeng1@163.com或15220958556"

                # Parse the decrypted data
                try:
                    license_data = json.loads(decrypted_data.decode())
                except json.JSONDecodeError:
                    return False, "License file has been tampered with. 请联系作者：weicongpeng1@163.com或15220958556"

                # Check expiration
                try:
                    expire_date = datetime.strptime(license_data['expire_date'], '%Y-%m-%d')
                except ValueError:
                    return False, "License file has been tampered with. 请联系作者：weicongpeng1@163.com或15220958556"

                current_date = datetime.now()

                if current_date > expire_date:
                    return False, f"License expired on {license_data['expire_date']}. 请联系作者：weicongpeng1@163.com或15220958556"

                return True, "License valid"

            except Exception as e:
                return False, f"Invalid license file: {str(e)}"

    license_mgr = LicenseManager()
    is_valid, message = license_mgr.check_license()

    if not is_valid:
        import tkinter as tk
        from tkinter import messagebox
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        root.attributes('-topmost', True)  # Bring to front
        full_message = "License验证失败，请联系作者获取授权。\n\n联系作者：weicongpeng1@163.com或15220958556"
        messagebox.showerror("License Error", full_message)
        print(message)  # Also print to console for debugging
        root.destroy()
        exit(1)
except ImportError:
    import tkinter as tk
    from tkinter import messagebox
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    root.attributes('-topmost', True)  # Bring to front
    full_message_import = "缺少License管理模块，请联系作者获取支持。\n\n联系作者：weicongpeng1@163.com或15220958556"
    messagebox.showerror("Import Error", full_message_import)
    root.destroy()
    exit(1)

import pandas as pd
import numpy as np
import math
from typing import Dict, List, Tuple, Set, Optional, Union
from datetime import datetime
import time
import warnings
import zipfile
import os
import glob
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import queue
import subprocess
import sys

warnings.filterwarnings('ignore')

class NetworkParameterUpdater:
    """
    现网工参更新工具类
    用于从现网工参压缩包中提取信息并更新全量工参文件
    支持模糊列名匹配和特殊处理逻辑
    """

    def __init__(self):
        self.full_param_dir = "全量工参"
        self.baseline_zip_pattern = "BaselineLab_*.zip"
        self.lte_online_params = None
        self.nr_online_params = None
        self.extracted_files = []

    def _get_latest_parameter_file(self, file_list):
        """从文件列表中选择时间戳最新的文件"""
        if not file_list:
            return None

        import re
        from datetime import datetime

        # 提取文件名中的时间戳并排序
        file_timestamps = []
        for file_path in file_list:
            file_name = os.path.basename(file_path)
            # 查找文件名中的14位时间戳 (YYYYMMDDHHMMSS)
            timestamp_match = re.search(r'(\d{14})', file_name)
            if timestamp_match:
                timestamp_str = timestamp_match.group(1)
                try:
                    # 尝试解析时间戳
                    timestamp = datetime.strptime(timestamp_str, '%Y%m%d%H%M%S')
                    file_timestamps.append((timestamp, file_path, f"14位时间戳: {timestamp_str}"))
                except ValueError:
                    # 如果不是14位时间戳，尝试查找8位日期
                    date_match = re.search(r'(\d{8})', file_name)
                    if date_match:
                        date_str = date_match.group(1)
                        try:
                            date = datetime.strptime(date_str, '%Y%m%d')
                            file_timestamps.append((date, file_path, f"8位日期: {date_str}"))
                        except ValueError:
                            # 如果都不是有效时间格式，使用文件修改时间
                            file_timestamps.append((datetime.fromtimestamp(os.path.getmtime(file_path)), file_path, f"文件修改时间"))
            else:
                # 如果没有找到时间戳，使用文件修改时间
                file_timestamps.append((datetime.fromtimestamp(os.path.getmtime(file_path)), file_path, f"文件修改时间"))

        # 按时间戳排序，选择最新的文件
        if file_timestamps:
            file_timestamps.sort(key=lambda x: x[0], reverse=True)
            latest_file = file_timestamps[0][1]
            timestamp_info = file_timestamps[0][2]
            print(f"选择最新文件: {os.path.basename(latest_file)} ({timestamp_info})")

            # 显示所有文件的时间戳信息（用于调试）
            print(f"找到 {len(file_timestamps)} 个候选文件:")
            for i, (timestamp, file_path, info) in enumerate(file_timestamps[:5]):  # 只显示前5个
                file_name = os.path.basename(file_path)
                print(f"  {i+1}. {file_name} - {info} ({timestamp.strftime('%Y-%m-%d %H:%M:%S')})")

            return latest_file

        return file_list[0] if file_list else None

    def update_network_parameters(self):
        """更新现网工参到全量工参"""
        try:
            # 查找全量工参文件
            full_param_files = glob.glob(f"{self.full_param_dir}/ProjectParameter_mongoose*.xlsx")
            if not full_param_files:
                print("❌ 未找到全量工参文件")
                return False

            # 选择最新的工参文件
            full_param_file = self._get_latest_parameter_file(full_param_files)
            if not full_param_file:
                print("❌ 无法确定最新的全量工参文件")
                return False

            print(f"找到最新的全量工参文件: {full_param_file}")

            # 查找现网工参压缩包
            baseline_zips = glob.glob(f"{self.full_param_dir}/{self.baseline_zip_pattern}")
            if not baseline_zips:
                print("❌ 未找到现网工参压缩包")
                return False

            baseline_zip = baseline_zips[0]
            print(f"正在提取现网工参压缩包: {baseline_zip}")

            # 提取压缩包中的工参文件
            lte_sdr_files, lte_itbbu_files, nr_files = self._extract_parameter_files(baseline_zip)

            if not any([lte_sdr_files, lte_itbbu_files, nr_files]):
                print("❌ 未找到有效的工参文件")
                return False

            # 加载现网工参数据
            lte_online_df = self._load_lte_online_data(lte_sdr_files, lte_itbbu_files)
            nr_online_df = self._load_nr_online_data(nr_files)

            if lte_online_df is None and nr_online_df is None:
                print("❌ 未能加载任何工参数据")
                return False

            # 加载全量工参数据
            print(f"开始更新全量工参文件: {full_param_file}")
            lte_full_df, nr_full_df = self._load_full_parameters(full_param_file)

            if lte_full_df is not None:
                print(f"全量LTE工参数: {len(lte_full_df)}")
            if nr_full_df is not None:
                print(f"全量NR工参数: {len(nr_full_df)}")

            # 更新工参数据
            updated = False
            if lte_online_df is not None and lte_full_df is not None:
                lte_full_df = self._update_lte_parameters(lte_full_df, lte_online_df)
                updated = True

            if nr_online_df is not None and nr_full_df is not None:
                nr_full_df = self._update_nr_parameters(nr_full_df, nr_online_df)
                updated = True

            if not updated:
                print("❌ 没有需要更新的工参数据")
                return False

            # 保存更新后的工参文件
            self._save_updated_parameters(full_param_file, lte_full_df, nr_full_df)

            print("✅ 现网工参更新完成")
            return True

        except Exception as e:
            print(f"❌ 更新全量工参失败: {e}")
            return False

    def _extract_parameter_files(self, zip_path):
        """从压缩包中提取工参文件"""
        lte_sdr_files = []
        lte_itbbu_files = []
        nr_files = []

        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                # 创建临时目录存放解压文件
                extract_dir = "temp_extract"
                os.makedirs(extract_dir, exist_ok=True)

                # 只提取需要的文件
                for file_info in zip_ref.infolist():
                    filename = file_info.filename
                    # 处理路径分隔符问题，确保正确提取
                    clean_filename = filename.replace('\\', '/').split('/')[-1]

                    if 'LTE_SDR_CellInfo' in clean_filename and clean_filename.endswith('.csv'):
                        zip_ref.extract(file_info, extract_dir)
                        extracted_path = os.path.join(extract_dir, filename.replace('\\', '/'))
                        lte_sdr_files.append(extracted_path)
                    elif 'LTE_ITBBU_CellInfo' in clean_filename and clean_filename.endswith('.csv'):
                        zip_ref.extract(file_info, extract_dir)
                        extracted_path = os.path.join(extract_dir, filename.replace('\\', '/'))
                        lte_itbbu_files.append(extracted_path)
                    elif 'NR_CellInfo' in clean_filename and clean_filename.endswith('.csv') and not clean_filename.endswith('.csvX'):
                        zip_ref.extract(file_info, extract_dir)
                        extracted_path = os.path.join(extract_dir, filename.replace('\\', '/'))
                        nr_files.append(extracted_path)

                print(f"找到LTE_SDR文件: {len(lte_sdr_files)} 个")
                print(f"找到LTE_ITBBU文件: {len(lte_itbbu_files)} 个")
                print(f"找到NR文件: {len(nr_files)} 个")

                # 如果NR文件在根目录但路径处理有问题，重新检查
                if not nr_files:
                    for file_info in zip_ref.infolist():
                        filename = file_info.filename
                        clean_filename = filename.replace('\\', '/').split('/')[-1]
                        if 'NR_CellInfo' in clean_filename and clean_filename.endswith('.csv') and not clean_filename.endswith('.csvX'):
                            # 确保文件正确提取到临时目录
                            zip_ref.extract(file_info, extract_dir)
                            extracted_path = os.path.join(extract_dir, filename.replace('\\', '/'))
                            if os.path.exists(extracted_path):
                                nr_files.append(extracted_path)
                                print(f"重新提取NR文件: {clean_filename}")

        except Exception as e:
            print(f"提取压缩包文件失败: {e}")

        return lte_sdr_files, lte_itbbu_files, nr_files

    def _load_lte_online_data(self, lte_sdr_files, lte_itbbu_files):
        """加载LTE现网工参数据"""
        lte_dfs = []

        # 加载LTE_SDR文件
        for file in lte_sdr_files:
            try:
                df = pd.read_csv(file, encoding='utf-8-sig')
                print(f"加载LTE_SDR文件: {os.path.basename(file)}, 记录数: {len(df)}")
                lte_dfs.append(df)
                # 清理临时文件
                os.remove(file)
            except Exception as e:
                print(f"加载LTE_SDR文件失败 {file}: {e}")

        # 加载LTE_ITBBU文件
        for file in lte_itbbu_files:
            try:
                df = pd.read_csv(file, encoding='utf-8-sig')
                # 处理LTE_ITBBU文件的列名大小写问题 - 只标准化cellName相关列
                for col in df.columns:
                    if 'cellname' in col.lower():
                        # 将cellName相关列统一为'cellName'
                        df = df.rename(columns={col: 'cellName'})
                print(f"加载LTE_ITBBU文件: {os.path.basename(file)}, 记录数: {len(df)}")
                lte_dfs.append(df)
                # 清理临时文件
                os.remove(file)
            except Exception as e:
                print(f"加载LTE_ITBBU文件失败 {file}: {e}")

        # 清理临时目录
        temp_dir = "temp_extract"
        if os.path.exists(temp_dir) and os.path.isdir(temp_dir):
            import shutil
            shutil.rmtree(temp_dir, ignore_errors=True)

        if not lte_dfs:
            return None

        # 合并所有LTE数据
        lte_online_df = pd.concat(lte_dfs, ignore_index=True)
        print(f"合并LTE现网工参总数: {len(lte_online_df)}")
        return lte_online_df

    def _load_nr_online_data(self, nr_files):
        """加载NR现网工参数据"""
        nr_dfs = []

        for file in nr_files:
            try:
                # 检查文件是否存在
                if not os.path.exists(file):
                    print(f"NR文件不存在: {file}")
                    # 尝试直接提取文件
                    # 动态查找现网工参压缩包
                    baseline_zips = glob.glob(f"{self.full_param_dir}/{self.baseline_zip_pattern}")
                    if baseline_zips:
                        zip_path = baseline_zips[0]
                        filename = os.path.basename(file)
                        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                            zip_ref.extract(filename, "temp_extract")
                            extracted_file = os.path.join("temp_extract", filename)
                            if os.path.exists(extracted_file):
                                file = extracted_file
                                print(f"重新提取NR文件: {filename}")

                if os.path.exists(file):
                    df = pd.read_csv(file, encoding='utf-8-sig')
                    print(f"加载NR文件: {os.path.basename(file)}, 记录数: {len(df)}")
                    nr_dfs.append(df)

                    # 显示NR文件的列名用于调试
                    #print(f"NR文件列名: {list(df.columns)}")
                else:
                    print(f"NR文件仍然不存在: {file}")

            except Exception as e:
                print(f"加载NR文件失败 {file}: {e}")
                # 尝试其他编码
                try:
                    df = pd.read_csv(file, encoding='gbk')
                    print(f"使用GBK编码加载NR文件成功: {os.path.basename(file)}, 记录数: {len(df)}")
                    nr_dfs.append(df)
                    #print(f"NR文件列名: {list(df.columns)}")
                except Exception as e2:
                    print(f"使用GBK编码加载NR文件也失败 {file}: {e2}")

        # 清理临时文件
        for file in nr_files:
            try:
                if os.path.exists(file):
                    os.remove(file)
            except:
                pass

        # 清理临时目录
        temp_dir = "temp_extract"
        if os.path.exists(temp_dir) and os.path.isdir(temp_dir):
            import shutil
            shutil.rmtree(temp_dir, ignore_errors=True)

        if not nr_dfs:
            return None

        # 合并所有NR数据
        nr_online_df = pd.concat(nr_dfs, ignore_index=True)
        print(f"合并NR现网工参总数: {len(nr_online_df)}")
        return nr_online_df

    def _load_full_parameters(self, file_path):
        """加载全量工参数据"""
        try:
            lte_full_df = None
            nr_full_df = None

            # 读取LTE工参表
            try:
                lte_full_df = pd.read_excel(file_path, sheet_name='LTE Project Parameters')
            except:
                print("未找到LTE工参表")

            # 读取NR工参表
            try:
                nr_full_df = pd.read_excel(file_path, sheet_name='NR Project Parameters')
            except:
                print("未找到NR工参表")

            return lte_full_df, nr_full_df

        except Exception as e:
            print(f"加载全量工参失败: {e}")
            return None, None

    def _apply_cell_formatting(self, file_path: str) -> None:
        """
        应用单元格格式设置 - 已移除颜色填充功能

        Args:
            file_path: Excel文件路径
        """
        try:
            # 加载工作簿
            wb = load_workbook(file_path)

            # 处理LTE工作表
            if 'LTE Project Parameters' in wb.sheetnames:
                lte_ws = wb['LTE Project Parameters']

                try:
                    # 安全地获取工作表的行列数
                    max_row = lte_ws.max_row
                    max_col = lte_ws.max_column

                    # 检查工作表是否为空且有足够的行数
                    if max_row >= 4 and max_col > 0:
                        # 获取列名映射
                        lte_columns = {cell.value: cell.column for cell in lte_ws[1]}

                        # 不再应用颜色填充
                        pass
                except Exception as e:
                    print(f"处理LTE工作表时出错: {e}")

            # 处理NR工作表
            if 'NR Project Parameters' in wb.sheetnames:
                nr_ws = wb['NR Project Parameters']

                try:
                    # 安全地获取工作表的行列数
                    max_row = nr_ws.max_row
                    max_col = nr_ws.max_column

                    # 检查工作表是否为空且有足够的行数
                    if max_row >= 4 and max_col > 0:
                        # 获取列名映射
                        nr_columns = {cell.value: cell.column for cell in nr_ws[1]}

                        # 不再应用颜色填充
                        pass
                except Exception as e:
                    print(f"处理NR工作表时出错: {e}")

            # 保存格式设置
            wb.save(file_path)
            print("✅ 单元格格式设置完成")

        except Exception as e:
            print(f"单元格格式设置失败: {e}")

    def _find_column_by_prefix(self, df: pd.DataFrame, prefix: str) -> str:
        """
        通过前缀模糊查找列名

        Args:
            df: 数据框
            prefix: 列名前缀

        Returns:
            匹配的列名，如果找不到返回None
        """
        for col in df.columns:
            if col.startswith(prefix):
                return col
        return None

    def _find_column_by_fuzzy_match(self, df: pd.DataFrame, keyword: str) -> str:
        """
        通过关键词模糊查找列名，支持部分匹配
        例如：列名为"子网ID SubNetwork ID long[0...128]"，只用"子网ID"即可匹配

        Args:
            df: 数据框
            keyword: 关键词

        Returns:
            匹配的列名，如果找不到返回None
        """
        # 首先尝试精确匹配
        if keyword in df.columns:
            return keyword

        # 然后尝试部分匹配 - 检查关键词是否包含在列名中
        for col in df.columns:
            if keyword in col:
                return col

        # 最后尝试前缀匹配作为备选
        return self._find_column_by_prefix(df, keyword)

    def _update_lte_parameters(self, lte_full_df: pd.DataFrame, lte_online_df: pd.DataFrame) -> pd.DataFrame:
        """
        更新LTE工参表

        Args:
            lte_full_df: 全量LTE工参
            lte_online_df: 现网LTE工参

        Returns:
            更新后的LTE工参
        """
        print("开始更新LTE工参表...")

        # 创建合并键
        # 使用模糊匹配查找列名
        enodeb_id_col = self._find_column_by_fuzzy_match(lte_full_df, 'eNodeB标识')
        cell_id_col = self._find_column_by_fuzzy_match(lte_full_df, '小区标识')

        # 确保列名有效再进行下一步操作
        if enodeb_id_col is None:
            print("错误: 无法在LTE全量工参中找到eNodeB标识列")
            return lte_full_df  # 返回原数据框，不进行更新
        if cell_id_col is None:
            print("错误: 无法在LTE全量工参中找到小区标识列")
            return lte_full_df  # 返回原数据框，不进行更新

        lte_full_df['merge_key'] = lte_full_df[enodeb_id_col].astype(str) + "_" + lte_full_df[cell_id_col].astype(str)
        lte_online_df['merge_key'] = lte_online_df['eNBId'].astype(str) + "_" + lte_online_df['cellLocalId'].astype(str)

        # 字段映射 - 支持LTE_SDR和LTE_ITBBU文件的不同列名
        field_mapping = {
            'SubNetwork': '子网ID',
            'ManagedElement': '管理网元ID',
            'eNBName': '基站名称',
            'eNBId': 'eNodeB标识',
            'cellName': '小区名称',  # LTE_SDR文件中的列名
            'CellName': '小区名称',  # LTE_ITBBU文件中的列名
            'cellLocalId': '小区标识',
            'tac': '跟踪区码',
            'pci': '物理小区识别码',
            'frequency': '下行链路的中心载频'
        }

        # 使用迭代方式更新，避免merge的性能问题
        updated_count = 0
        new_cells_count = 0

        # 创建全量工参的键值映射，跳过表头行（第1行和第2行，索引0和1）
        # 从第3行开始创建键值映射（索引从2开始）
        full_keys_set = set(lte_full_df['merge_key'][2:]) if len(lte_full_df) > 2 else set()

        for online_idx, online_row in lte_online_df.iterrows():
            merge_key = online_row['merge_key']

            if merge_key in full_keys_set:
                # 更新现有记录，跳过表头行（第1行和第2行，索引0和1）
                # 只从第3行开始查找匹配项（索引从2开始）
                matching_indices = lte_full_df[(lte_full_df['merge_key'] == merge_key) & (lte_full_df.index >= 2)].index

                for online_field, full_field in field_mapping.items():
                    if online_field in online_row:
                        # 使用模糊匹配查找全量工参中的目标列
                        target_col = self._find_column_by_fuzzy_match(lte_full_df, full_field)
                        if target_col and pd.notna(online_row[online_field]):
                            # 只更新数据行，跳过表头行
                            lte_full_df.loc[matching_indices, target_col] = online_row[online_field]
                            updated_count += len(matching_indices)
                        elif target_col is None:
                            # 当在全量工参中找不到对应列时，创建新列并填充"未找到"
                            # 只对数据行（从第3行开始，索引2）填充"未找到"
                            lte_full_df.loc[matching_indices, full_field] = "未找到"
                            updated_count += len(matching_indices)
            else:
                # 添加新小区
                new_record = {}
                for online_field, full_field in field_mapping.items():
                    if online_field in online_row:
                        # 使用模糊匹配查找全量工参中的目标列
                        target_col = self._find_column_by_fuzzy_match(lte_full_df, full_field)
                        if target_col:
                            new_record[target_col] = online_row[online_field]

                # 只添加有有效数据的记录
                if new_record:
                    lte_full_df = pd.concat([lte_full_df, pd.DataFrame([new_record])], ignore_index=True)
                    new_cells_count += 1

        print(f"LTE工参更新: 更新了 {updated_count} 个字段值")
        print(f"LTE工参新增: 添加了 {new_cells_count} 个新小区")

        # 移除临时列
        lte_full_df = lte_full_df.drop('merge_key', axis=1)

        # 填充默认值
        lte_full_df = self._fill_default_values(lte_full_df, 'LTE')

        print(f"LTE工参更新完成，总记录数: {len(lte_full_df)}")
        return lte_full_df

    def _update_nr_parameters(self, nr_full_df: pd.DataFrame, nr_online_df: pd.DataFrame) -> pd.DataFrame:
        """
        更新NR工参表

        Args:
            nr_full_df: 全量NR工参
            nr_online_df: 现网NR工参

        Returns:
            更新后的NR工参
        """
        print("开始更新NR工参表...")

        # 创建合并键 - 调整为全局匹配方式
        # 全量工参中的"移动国家码"&"移动网络码"&"gNodeB标识"&"小区标识"合并值
        mcc_col = self._find_column_by_fuzzy_match(nr_full_df, '移动国家码')
        mnc_col = self._find_column_by_fuzzy_match(nr_full_df, '移动网络码')
        gnodb_id_col = self._find_column_by_fuzzy_match(nr_full_df, 'gNodeB标识')
        cell_id_col = self._find_column_by_fuzzy_match(nr_full_df, '小区标识')

        # 确保列名有效再进行下一步操作
        if mcc_col is None:
            print("错误: 无法在NR全量工参中找到移动国家码列")
            return nr_full_df  # 返回原数据框，不进行更新
        if mnc_col is None:
            print("错误: 无法在NR全量工参中找到移动网络码列")
            return nr_full_df  # 返回原数据框，不进行更新
        if gnodb_id_col is None:
            print("错误: 无法在NR全量工参中找到gNodeB标识列")
            return nr_full_df  # 返回原数据框，不进行更新
        if cell_id_col is None:
            print("错误: 无法在NR全量工参中找到小区标识列")
            return nr_full_df  # 返回原数据框，不进行更新

        # 处理plmn字段：删除"-"，并将值为"1"的移动网络码改为"01"
        def format_plmn(plmn):
            """
            处理plmn字段：删除"-"，并将值为"1"的移动网络码改为"01"
            """
            if pd.isna(plmn):
                return None  # 不设置默认值，保持为空

            plmn_str = str(plmn).strip()

            # 如果包含"-"
            if "-" in plmn_str:
                parts = plmn_str.split("-")
                if len(parts) >= 2:
                    left_part = parts[0]  # 左边字符（MCC）
                    right_part = parts[1]  # 右边字符（MNC）

                    # 右边字符如果是"1"则改为"01"
                    if right_part == "1":
                        right_part = "01"
                    elif len(right_part) == 1 and right_part.isdigit():
                        # 其他单数字也补齐为2位
                        right_part = "0" + right_part

                    return left_part + right_part

            # 如果不包含"-"，假设是46011这种格式
            if len(plmn_str) == 5 and plmn_str.startswith("460"):
                # 检查是否需要将"1"改为"01"
                if plmn_str.endswith("1") and not plmn_str.endswith("01") and not plmn_str.endswith("11") and not plmn_str.endswith("21") and not plmn_str.endswith("31") and not plmn_str.endswith("41") and not plmn_str.endswith("51") and not plmn_str.endswith("61") and not plmn_str.endswith("71") and not plmn_str.endswith("81") and not plmn_str.endswith("91"):
                    return plmn_str[:-1] + "01"
                return plmn_str

            # 保持原值
            return None

        # 确保移动网络码为2位字符串，不足用0补齐
        # 修复：处理"460-XX"格式，只保留"XX"部分
        def format_mnc(mnc):
            if pd.notna(mnc) and str(mnc).strip():
                mnc_str = str(mnc).strip()
                # 如果是"460-XX"格式，只保留"XX"部分
                if mnc_str.startswith("460-"):
                    mnc_part = mnc_str[4:]  # 去掉"460-"前缀
                    return mnc_part.zfill(2) if mnc_part.isdigit() else mnc_part
                # 其他情况保持原有逻辑
                return mnc_str.zfill(2) if mnc_str.isdigit() else mnc_str
            return None  # 不设置默认值，保持为空

        # 处理全量工参的移动网络码字段，应用format_mnc函数
        nr_full_df[mnc_col] = nr_full_df[mnc_col].apply(format_mnc)

        # 创建全量工参合并键
        nr_full_df[mcc_col] = nr_full_df[mcc_col].fillna('460')  # 填充缺失的移动国家码
        # 不再强制格式化移动网络码，保持原有值，只在匹配时更新
        nr_full_df['merge_key'] = (nr_full_df[mcc_col].astype(str) +
                                  nr_full_df[mnc_col].astype(str) +
                                  nr_full_df[gnodb_id_col].astype(str) +
                                  nr_full_df[cell_id_col].astype(str))

        # 现网工参中的"plmn"&"gNBId"&"cellLocalId"合并值（使用新的处理方法）
        nr_online_df['processed_plmn'] = nr_online_df['plmn'].apply(format_plmn)
        nr_online_df['merge_key'] = (nr_online_df['processed_plmn'].astype(str) +
                                    nr_online_df['gNBId'].astype(str) +
                                    nr_online_df['cellLocalId'].astype(str))

        #print(f"NR全量工参列名: {list(nr_full_df.columns)}")
        #print(f"NR现网工参列名: {list(nr_online_df.columns)}")
        #print(f"模糊匹配结果 - 移动国家码: {mcc_col}, 移动网络码: {mnc_col}, gNodeB标识: {gnodb_id_col}, 小区标识: {cell_id_col}")

        # 字段映射 - 从现网工参字段 -> 全量工参字段
        field_mapping = {
            'SubNetwork': '子网ID',
            'plmn': '移动网络码',
            'gNBName': '基站名称',
            'gNBId': 'gNodeB标识',
            'CellName': '小区名称',
            'cellLocalId': '小区标识',
            'ssbFrequency': '填写SSB频点',
            'pci': '物理小区识别码',
            'gNodeBLength': 'gNodeBLength'
        }

        # 使用迭代方式更新，避免merge的性能问题
        updated_count = 0
        new_cells_count = 0

        # 创建全量工参的键值映射，跳过表头行（第1行和第2行，索引0和1）
        # 从第3行开始创建键值映射（索引从2开始）
        full_keys_set = set(nr_full_df['merge_key'][2:]) if len(nr_full_df) > 2 else set()

        # 更新现有记录
        updated_count = 0
        for online_idx, online_row in nr_online_df.iterrows():
            merge_key = online_row['merge_key']

            if merge_key in full_keys_set:
                # 更新现有记录，跳过表头行（第1行和第2行，索引0和1）
                # 只从第3行开始查找匹配项（索引从2开始）
                matching_indices = nr_full_df[(nr_full_df['merge_key'] == merge_key) & (nr_full_df.index >= 2)].index

                # 只更新第一条匹配的记录，避免重复更新
                if len(matching_indices) > 0:
                    first_matching_index = matching_indices[0]

                    for online_field, full_field in field_mapping.items():
                        if online_field in online_row:
                            # 使用模糊匹配查找全量工参中的目标列
                            target_col = self._find_column_by_fuzzy_match(nr_full_df, full_field)
                            if target_col and pd.notna(online_row[online_field]):
                                # 特殊处理plmn字段，应用format_mnc函数提取和格式化MNC
                                if online_field == 'plmn':
                                    nr_full_df.loc[first_matching_index, target_col] = format_mnc(online_row[online_field])
                                else:
                                    # 只更新数据行，跳过表头行
                                    nr_full_df.loc[first_matching_index, target_col] = online_row[online_field]
                                updated_count += 1

                    # 更新full_keys_set中的键值，确保后续匹配正确
                    # 重新计算该记录的合并键
                    new_merge_key = (str(nr_full_df.loc[first_matching_index, mcc_col]) +
                                   str(nr_full_df.loc[first_matching_index, mnc_col]) +
                                   str(nr_full_df.loc[first_matching_index, gnodb_id_col]) +
                                   str(nr_full_df.loc[first_matching_index, cell_id_col]))
                    # 从集合中移除旧键值并添加新键值
                    full_keys_set.discard(merge_key)
                    full_keys_set.add(new_merge_key)
            else:
                # 添加新小区
                new_record = {}
                for online_field, full_field in field_mapping.items():
                    if online_field in online_row:
                        # 使用模糊匹配查找全量工参中的目标列
                        target_col = self._find_column_by_fuzzy_match(nr_full_df, full_field)
                        if target_col:
                            # 特殊处理plmn字段，应用format_mnc函数提取和格式化MNC
                            if online_field == 'plmn':
                                new_record[target_col] = format_mnc(online_row[online_field])
                            else:
                                new_record[target_col] = online_row[online_field]

                # 只添加有有效数据的记录，并确保不重复
                if new_record:
                    # 创建新记录的合并键用于去重检查
                    new_gnodeb_id = new_record.get(gnodb_id_col, '')
                    new_cell_id = new_record.get(cell_id_col, '')
                    new_mcc = new_record.get(mcc_col, '460')
                    new_mnc = new_record.get(mnc_col, '')

                    # 生成新记录的合并键
                    new_merge_key = str(new_mcc) + str(new_mnc) + str(new_gnodeb_id) + str(new_cell_id)

                    # 检查是否已存在相同的记录（跳过表头行）
                    if new_merge_key not in full_keys_set:
                        nr_full_df = pd.concat([nr_full_df, pd.DataFrame([new_record])], ignore_index=True)
                        new_cells_count += 1
                        # 更新键值集合
                        full_keys_set.add(new_merge_key)
                    # 如果已存在，则跳过添加

        print(f"NR工参更新: 更新了 {updated_count} 个字段值")
        print(f"NR工参新增: 添加了 {new_cells_count} 个新小区")

        # 对NR工参进行去重处理，基于移动国家码、移动网络码、gNodeB标识和小区标识的组合
        print(f"去重前记录数: {len(nr_full_df)}")
        # 跳过表头行（第1行和第2行，索引0和1）
        header_df = nr_full_df.iloc[:2] if len(nr_full_df) >= 2 else nr_full_df
        data_df = nr_full_df.iloc[2:] if len(nr_full_df) > 2 else pd.DataFrame()

        if not data_df.empty:
            # 基于移动国家码、移动网络码、gNodeB标识和小区标识的组合进行去重，保留第一条记录
            data_df = data_df.drop_duplicates(subset=[mcc_col, mnc_col, gnodb_id_col, cell_id_col], keep='first')
            # 重新合并表头和数据
            nr_full_df = pd.concat([header_df, data_df], ignore_index=True)

        print(f"去重后记录数: {len(nr_full_df)}")

        # 移除临时列
        nr_full_df = nr_full_df.drop('merge_key', axis=1)
        if 'processed_plmn' in nr_online_df.columns:
            nr_online_df = nr_online_df.drop('processed_plmn', axis=1)

        # 填充默认值
        nr_full_df = self._fill_default_values(nr_full_df, 'NR')

        print(f"NR工参更新完成，总记录数: {len(nr_full_df)}")
        return nr_full_df

    def _fill_default_values(self, df: pd.DataFrame, network_type: str) -> pd.DataFrame:
        """
        填充默认值

        Args:
            df: 工参数据
            network_type: 网络类型 ('LTE' 或 'NR')

        Returns:
            填充默认值后的数据
        """
        if network_type == 'LTE':
            # 填充LTE默认值 - 强制填充
            df = self._fill_lte_defaults(df)
        elif network_type == 'NR':
            # 填充NR默认值 - 强制填充
            df = self._fill_nr_defaults(df)

        return df

    def _fill_lte_defaults(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        强制填充LTE默认值，跳过表头行（第1行和第2行）
        """
        # 模糊查找列名并强制填充，只对数据行（从第3行开始，索引2）进行填充

        for col in df.columns:
            # 第一分组 - 只对新增小区强制填入"电信中兴"，其他情况不操作
            if '第一分组' in col:
                # 只对值为空的单元格填入默认值，跳过表头行
                if len(df) > 2:  # 确保有足够的数据行
                    df.iloc[2:, df.columns.get_loc(col)] = df.iloc[2:, df.columns.get_loc(col)].fillna('电信中兴')
            # 系统制式 - 强制填入"1"，跳过表头行
            elif '系统制式' in col:
                if len(df) > 2:  # 确保有足够的数据行
                    df.iloc[2:, df.columns.get_loc(col)] = '1'
            # 移动国家码 - 强制填入"460"，跳过表头行
            elif '移动国家码' in col:
                if len(df) > 2:  # 确保有足够的数据行
                    df.iloc[2:, df.columns.get_loc(col)] = '460'
            # 移动网络码 - 默认填入"11"，跳过表头行
            elif '移动网络码' in col:
                if len(df) > 2:  # 确保有足够的数据行
                    df.iloc[2:, df.columns.get_loc(col)] = df.iloc[2:, df.columns.get_loc(col)].fillna('11')
            # 基站名称 - 保持为空时不处理
            elif '基站名称' in col:
                df[col] = df[col]  # 保持原值
            # 子网ID - 保持为空时不处理
            elif '子网ID' in col:
                df[col] = df[col]  # 保持原值
            # 管理网元ID - 保持为空时不处理
            elif '管理网元ID' in col:
                df[col] = df[col]  # 保持原值

        return df

    def _fill_nr_defaults(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        强制填充NR默认值，跳过表头行（第1行和第2行）
        """
        # 模糊查找列名并强制填充，只对数据行（从第3行开始，索引2）进行填充
        for col in df.columns:
            # 第一分组 - 只对新增小区强制填入"河源电信"，其他情况不操作
            if '第一分组' in col:
                # 只对值为空的单元格填入默认值，跳过表头行
                if len(df) > 2:  # 确保有足够的数据行
                    df.iloc[2:, df.columns.get_loc(col)] = df.iloc[2:, df.columns.get_loc(col)].fillna('河源电信')
            # 移动国家码 - 强制填入"460"，跳过表头行
            elif '移动国家码' in col:
                if len(df) > 2:  # 确保有足够的数据行
                    df.iloc[2:, df.columns.get_loc(col)] = '460'
            # gNodeBLength - 根据SSB频点判断
            elif 'gNodeBLength' in col:
                df = self._fill_gnodeb_length(df, col)
            # 移动网络码 - 从plmn字段提取最后两位
            elif '移动网络码' in col:
                df = self._fill_mnc_from_plmn(df, col)
            # 子网ID - 保持为空时不处理
            elif '子网ID' in col:
                df[col] = df[col]  # 保持原值

        return df

    def _fill_gnodeb_length(self, df: pd.DataFrame, length_col: str) -> pd.DataFrame:
        """
        根据SSB频点判断gNodeBLength值，跳过表头行（第1行和第2行）
        """
        # 查找SSB频点列
        ssb_col = None
        for col in df.columns:
            if 'SSB频点' in col or 'ssbFrequency' in col:
                ssb_col = col
                break

        if ssb_col is None:
            print("警告: 未找到SSB频点列，无法判断gNodeBLength")
            # 只对数据行（从第3行开始，索引2）填充默认值
            if len(df) > 2:  # 确保有足够的数据行
                df.iloc[2:, df.columns.get_loc(length_col)] = df.iloc[2:, df.columns.get_loc(length_col)].fillna(5)
            return df

        # 根据SSB频点范围判断gNodeBLength
        def determine_gnodeb_length(ssb_freq):
            if pd.isna(ssb_freq):
                return "未找到"  # 当SSB频点为空时填充"未找到"

            try:
                ssb_freq = float(ssb_freq)
                if 870 <= ssb_freq <= 881:
                    return 5
                elif 2113 <= ssb_freq <= 2145:
                    return 1
                elif 3407 <= ssb_freq <= 3510:
                    return 78
                else:
                    return "未找到"  # 当SSB频点不在匹配范围时填充"未找到"
            except (ValueError, TypeError):
                return "未找到"  # 当SSB频点无法转换为数值时填充"未找到"

        # 只对数据行（从第3行开始，索引2）应用判断逻辑
        if len(df) > 2:  # 确保有足够的数据行
            df.iloc[2:, df.columns.get_loc(length_col)] = df.iloc[2:, df.columns.get_loc(ssb_col)].apply(determine_gnodeb_length)
        return df

    def _fill_mnc_from_plmn(self, df: pd.DataFrame, mnc_col: str) -> pd.DataFrame:
        """
        从plmn字段提取移动网络码，跳过表头行（第1行和第2行）
        """
        # 查找plmn列
        plmn_col = None
        for col in df.columns:
            if 'plmn' in col.lower() or 'PLMN' in col:
                plmn_col = col
                break

        if plmn_col is None:
            print("警告: 未找到plmn列，无法提取移动网络码")
            # 不再填充默认值，保持原有值
            return df

        # 从plmn值提取移动网络码
        def extract_mnc_from_plmn(plmn_value):
            if pd.isna(plmn_value):
                return None  # 不设置默认值，保持为空

            plmn_str = str(plmn_value).strip()

            # 如果包含"-"
            if "-" in plmn_str:
                parts = plmn_str.split("-")
                if len(parts) >= 2:
                    right_part = parts[1]  # 右边字符（MNC）

                    # 右边字符如果是"1"则改为"01"
                    if right_part == "1":
                        right_part = "01"
                    elif len(right_part) == 1 and right_part.isdigit():
                        # 其他单数字也补齐为2位
                        right_part = "0" + right_part

                    return right_part

            # 如果不包含"-"，假设是46011这种格式
            if len(plmn_str) == 5 and plmn_str.startswith("460"):
                # 提取后两位作为MNC
                return plmn_str[3:]

            # 保持原值
            return None

        # 只对数据行（从第3行开始，索引2）应用提取逻辑，且只更新空值
        mask = (df.index >= 2) & (df[mnc_col].isna())
        df.loc[mask, mnc_col] = df.loc[mask, plmn_col].apply(extract_mnc_from_plmn)
        return df

    def _save_updated_parameters(self, file_path, lte_full_df, nr_full_df):
        """保存更新后的工参文件"""
        try:
            # 生成新的文件名，保持原有结构，只更新日期部分
            import os
            import re
            from datetime import datetime

            # 获取当前时间戳 (YYYYMMDDHHMMSS格式) 确保唯一性
            current_timestamp = datetime.now().strftime('%Y%m%d%H%M%S')

            # 提取原文件名中的信息
            base_name = os.path.basename(file_path)
            dir_path = os.path.dirname(file_path)

            # 检查文件名格式，确定时间戳的位置和长度
            # 模式1: ProjectParameter_mongoose河源电联20250928183756.xlsx (14位时间戳在末尾)
            # 模式2: ProjectParameter_mongoose20250923河源电联.xlsx (8位日期在中间)

            name_parts = os.path.splitext(base_name)
            file_name_without_ext = name_parts[0]
            file_ext = name_parts[1]

            # 尝试匹配末尾的14位时间戳 (YYYYMMDDHHMMSS)
            timestamp_match = re.search(r'(\d{14})$', file_name_without_ext)
            if timestamp_match:
                # 替换末尾的14位时间戳
                original_timestamp = timestamp_match.group(1)
                new_filename = file_name_without_ext[:-14] + current_timestamp + file_ext
                print(f"检测到末尾时间戳 {original_timestamp}，替换为 {current_timestamp}")
            else:
                # 尝试匹配末尾的8位日期 (YYYYMMDD)
                date_match = re.search(r'(\d{8})$', file_name_without_ext)
                if date_match:
                    # 替换末尾的8位日期为当前日期部分
                    original_date = date_match.group(1)
                    current_date = current_timestamp[:8]  # 只取前8位日期部分
                    new_filename = file_name_without_ext[:-8] + current_date + file_ext
                    print(f"检测到末尾日期 {original_date}，替换为 {current_date}")
                else:
                    # 如果没有找到时间戳模式，查找文件名中的8位或14位数字
                    all_numbers = re.findall(r'\d+', file_name_without_ext)
                    if all_numbers:
                        # 找到最大的数字序列并替换
                        max_number = max(all_numbers, key=len)
                        if len(max_number) == 14:
                            # 替换14位时间戳
                            new_filename = file_name_without_ext.replace(max_number, current_timestamp)
                        elif len(max_number) == 8:
                            # 替换8位日期
                            current_date = current_timestamp[:8]
                            new_filename = file_name_without_ext.replace(max_number, current_date)
                        else:
                            # 其他长度，在末尾添加时间戳
                            new_filename = file_name_without_ext + current_timestamp + file_ext
                    else:
                        # 如果没有数字，在末尾添加时间戳
                        new_filename = file_name_without_ext + current_timestamp + file_ext

            # 构建新的完整文件路径
            new_file_path = os.path.join(dir_path, new_filename)

            print(f"正在保存更新后的工参文件: {new_file_path}")
            print(f"原文件名: {base_name} -> 新文件名: {new_filename}")

            with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                if lte_full_df is not None:
                    lte_full_df.to_excel(writer, sheet_name='LTE Project Parameters', index=False)
                if nr_full_df is not None:
                    nr_full_df.to_excel(writer, sheet_name='NR Project Parameters', index=False)
            print("✅ 工参文件保存成功")
        except Exception as e:
            print(f"保存工参文件失败: {e}")
            # 如果追加模式失败，尝试创建新文件
            try:
                with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
                    if lte_full_df is not None:
                        lte_full_df.to_excel(writer, sheet_name='LTE Project Parameters', index=False)
                    if nr_full_df is not None:
                        nr_full_df.to_excel(writer, sheet_name='NR Project Parameters', index=False)
                print("✅ 工参文件保存成功（新文件）")
            except Exception as e2:
                print(f"创建新工参文件失败: {e2}")


class NeighborPlanningTool:
    """
    邻区规划工具类
    支持NR到NR、LTE到LTE、NR到LTE的邻区关系规划
    """

    def __init__(self, max_neighbors: int = 32,
                 coverage_distance_factor: float = 5/9, coverage_radius_factor: float = 5/9):
        """
        初始化邻区规划工具（仅使用覆盖圆算法）

        Args:
            max_neighbors: 每个小区的最大邻区数量
            coverage_distance_factor: 覆盖圆距离系数k（站点到覆盖圆心距离 = k * Co）
            coverage_radius_factor: 覆盖圆半径系数m（覆盖半径 = m * Co）
        """
        self.max_neighbors = max_neighbors
        # 始终使用覆盖圆算法
        self.coverage_distance_factor = coverage_distance_factor  # k系数
        self.coverage_radius_factor = coverage_radius_factor      # m系数
        self.lte_cells = None
        self.nr_cells = None
        self.distance_cache = {}
        self.params_file = None  # 全量工参文件路径

        print(f"初始化邻区规划工具")
        print(f"使用覆盖圆算法")
        print(f"覆盖圆距离系数k: {coverage_distance_factor}")
        print(f"覆盖圆半径系数m: {coverage_radius_factor}")
        print(f"最大邻区数量: {max_neighbors}")

    def generate_timestamp_suffix(self) -> str:
        """生成时间戳后缀"""
        return datetime.now().strftime("%Y%m%d_%H%M%S")

    def calculate_distances_vectorized(self, target_lat: float, target_lon: float,
                                     source_lats: np.ndarray, source_lons: np.ndarray) -> np.ndarray:
        """
        向量化计算距离 - 优化版本
        批量计算一个目标点到多个源点的距离

        Args:
            target_lat: 目标纬度
            target_lon: 目标经度
            source_lats: 源点纬度数组
            source_lons: 源点经度数组

        Returns:
            np.ndarray: 距离数组
        """
        # 转换为弧度
        target_lat_rad = math.radians(target_lat)
        target_lon_rad = math.radians(target_lon)
        source_lats_rad = np.radians(source_lats)
        source_lons_rad = np.radians(source_lons)

        # Haversine公式
        dlat = source_lats_rad - target_lat_rad
        dlon = source_lons_rad - target_lon_rad
        a = np.sin(dlat/2)**2 + math.cos(target_lat_rad) * np.cos(source_lats_rad) * np.sin(dlon/2)**2
        c = 2 * np.arcsin(np.sqrt(a))

        # 地球半径（公里）
        return c * 6371.0

    def calculate_distance(self, lat1: float, lon1: float, lat2: float, lon2: float) -> float:
        """
        计算两点间的球面距离（公里）
        使用Haversine公式
        """
        # 检查缓存
        cache_key = (lat1, lon1, lat2, lon2)
        if cache_key in self.distance_cache:
            return self.distance_cache[cache_key]

        # 将角度转换为弧度
        lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])

        # Haversine公式
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
        c = 2 * math.asin(math.sqrt(a))

        # 地球半径（公里）
        earth_radius_km = 6371.0
        distance = earth_radius_km * c

        # 缓存结果
        self.distance_cache[cache_key] = distance
        return distance

    def calculate_coverage_circle_center(self, lat: float, lon: float, azimuth: float, distance_factor: float, coverage_distance: float) -> Tuple[float, float]:
        """
        计算覆盖圆心坐标

        Args:
            lat: 站点纬度
            lon: 站点经度
            azimuth: 天线方位角（度）
            distance_factor: 站点到覆盖圆心距离系数k
            coverage_distance: 正向覆盖距离Co（米转公里）

        Returns:
            tuple: (圆心纬度, 圆心经度)
        """
        # 转换方位角为弧度
        azimuth_rad = math.radians(azimuth)

        # 计算站点到覆盖圆心的距离（公里）
        d_s = distance_factor * coverage_distance / 1000  # 转换为公里

        # 使用简单的平面近似计算（适用于较短距离）
        # 计算纬度变化（每度约111公里）
        lat_delta = (d_s * math.cos(azimuth_rad)) / 111.0
        # 计算经度变化（每度约111公里，但需要考虑纬度影响）
        lon_delta = (d_s * math.sin(azimuth_rad)) / (111.0 * math.cos(math.radians(lat)))

        center_lat = lat + lat_delta
        center_lon = lon + lon_delta

        return center_lat, center_lon

    def calculate_coverage_radius(self, coverage_distance: float, radius_factor: float) -> float:
        """
        计算覆盖半径

        Args:
            coverage_distance: 正向覆盖距离Co（米）
            radius_factor: 覆盖半径系数m

        Returns:
            float: 覆盖半径（公里）
        """
        return radius_factor * coverage_distance / 1000  # 转换为公里

    def are_coverage_circles_intersecting(self, center1_lat: float, center1_lon: float, radius1: float,
                                          center2_lat: float, center2_lon: float, radius2: float) -> bool:
        """
        判断两个覆盖圆是否相交

        Args:
            center1_lat, center1_lon: 第一个圆心位置
            radius1: 第一个圆的半径
            center2_lat, center2_lon: 第二个圆心位置
            radius2: 第二个圆的半径

        Returns:
            bool: 如果相交返回True，否则返回False
        """
        # 计算两个圆心之间的距离
        center_distance = self.calculate_distance(center1_lat, center1_lon, center2_lat, center2_lon)

        # 两圆相交的条件：圆心距离 <= 半径之和
        return center_distance <= (radius1 + radius2)

    def preprocess_parameter_cells(self, raw_df: pd.DataFrame, network_type: str) -> pd.DataFrame:
        """
        预处理全量工参中的小区数据，标准化列名
        """
        if raw_df.empty:
            return pd.DataFrame()

        processed = pd.DataFrame()

        if network_type == "LTE":
            # LTE列映射
            processed['enodeb_id'] = raw_df['eNodeB标识\neNodeB ID\nlong:[0..1048575]']
            processed['cell_id'] = raw_df['小区标识\ncellLocalId\ninteger:[0~2147483647]']
            processed['cell_name'] = raw_df['小区名称\nuserLabel\nstring[0..128]']
            processed['pci'] = raw_df['物理小区识别码\nPCI\nlong:[0..503]']
            processed['lat'] = raw_df['小区纬度\neNodeB Latitude\ndouble:[-90..90]']
            processed['lon'] = raw_df['小区经度\neNodeB Longitude double:[-180..180]']
            processed['earfcn_dl'] = raw_df['下行链路的中心载频\nearfcnDl\ndouble Step：0.1 \nUnite：MHz']
            # 使用模糊匹配查找天线方向角列
            antenna_azimuth_col = self._find_column_by_fuzzy_match(raw_df, '天线方向角') or '天线方向角'
            processed['antenna_azimuth'] = raw_df.get(antenna_azimuth_col, None)
        else:  # NR
            processed['enodeb_id'] = raw_df['gNodeB标识\ngNodeB ID\nLong:[0..1048575]']
            processed['cell_id'] = raw_df['小区标识\ncellLocalId\nInteger:[0~2147483647]']
            processed['cell_name'] = raw_df['小区名称\nCELL NAME\nString[0..128]']
            processed['pci'] = raw_df['物理小区识别码\nPCI\nLong:[0..1007]']
            processed['lat'] = raw_df['小区纬度\nCell  Latitude\nDouble:[-90..90]']
            processed['lon'] = raw_df['小区经度\nCell  Longitude Double:[-180..180]']
            processed['earfcn_dl'] = raw_df['填写SSB频点\nSSB Frequency\nDouble Step：0.01 \nUnite：MHz']
            # 使用模糊匹配查找天线方向角列
            antenna_azimuth_col = self._find_column_by_fuzzy_match(raw_df, '天线方向角') or '天线方向角'
            processed['antenna_azimuth'] = raw_df.get(antenna_azimuth_col, None)

        processed['cell_type'] = network_type

        # 转换所有数值列（cell_name保持字符串格式）
        numeric_columns = ['enodeb_id', 'cell_id', 'pci', 'lat', 'lon', 'earfcn_dl']
        for col in numeric_columns:
            if col in processed.columns:
                processed[col] = pd.to_numeric(processed[col], errors='coerce')

        return processed

    def load_full_parameter_data(self, params_file: str) -> bool:
        """
        从全量工参文件中加载小区数据

        Args:
            params_file: 全量工参文件路径

        Returns:
            bool: 加载成功返回True
        """
        try:
            self.params_file = params_file
            print(f"正在加载全量工参文件: {os.path.basename(params_file)}")

            # 读取LTE数据
            try:
                lte_raw = pd.read_excel(params_file, sheet_name='LTE Project Parameters')
                # 跳过前两行（可能是表头说明）
                if len(lte_raw) > 2:
                    lte_raw = lte_raw.iloc[2:].reset_index(drop=True)
                self.lte_cells = self.preprocess_parameter_cells(lte_raw, 'LTE')
                print(f"加载LTE小区数据: {len(self.lte_cells)} 个小区")
            except Exception as e:
                print(f"加载LTE数据失败: {e}")
                self.lte_cells = None

            # 读取NR数据
            try:
                nr_raw = pd.read_excel(params_file, sheet_name='NR Project Parameters')
                # 跳过前两行（可能是表头说明）
                if len(nr_raw) > 2:
                    nr_raw = nr_raw.iloc[2:].reset_index(drop=True)
                self.nr_cells = self.preprocess_parameter_cells(nr_raw, 'NR')
                print(f"加载NR小区数据: {len(self.nr_cells)} 个小区")
            except Exception as e:
                print(f"加载NR数据失败: {e}")
                self.nr_cells = None

            return self.lte_cells is not None or self.nr_cells is not None

        except Exception as e:
            print(f"加载全量工参文件失败: {e}")
            return False

    def load_cell_data(self, cells_file: str) -> bool:
        """
        加载小区数据

        Args:
            cells_file: 小区数据文件路径

        Returns:
            bool: 加载成功返回True
        """
        try:
            # 读取Excel文件
            xl_file = pd.ExcelFile(cells_file)
            sheet_names = xl_file.sheet_names

            self.lte_cells = None
            self.nr_cells = None


            # 查找LTE工作表
            lte_sheets = [name for name in sheet_names if 'LTE' in name.upper()]
            if lte_sheets:
                self.lte_cells = pd.read_excel(cells_file, sheet_name=lte_sheets[0])
                print(f"加载LTE小区数据: {len(self.lte_cells)} 个小区")

            # 查找NR工作表
            nr_sheets = [name for name in sheet_names if 'NR' in name.upper()]
            if nr_sheets:
                self.nr_cells = pd.read_excel(cells_file, sheet_name=nr_sheets[0])
                print(f"加载NR小区数据: {len(self.nr_cells)} 个小区")

            return self.lte_cells is not None or self.nr_cells is not None

        except Exception as e:
            print(f"加载小区数据失败: {e}")
            return False

    def get_cell_key(self, row: pd.Series, network_type: str) -> str:
        """
        获取小区的唯一标识键

        Args:
            row: 小区数据行
            network_type: 网络类型 ('LTE' 或 'NR')

        Returns:
            str: 小区唯一标识
        """
        if network_type == 'LTE':
            # LTE使用eNodeB ID + Cell ID
            enodeb_id = str(row.get('enodeb_id', '')).strip()
            cell_id = str(row.get('cell_id', '')).strip()
            return f"{enodeb_id}_{cell_id}"
        else:  # NR
            # NR使用gNodeB ID + Cell ID
            gnodeb_id = str(row.get('enodeb_id', '')).strip()
            cell_id = str(row.get('cell_id', '')).strip()
            return f"{gnodeb_id}_{cell_id}"

    def get_cell_key_no_prefix(self, row: pd.Series, network_type: str) -> str:
        """
        获取小区的唯一标识键（无网络类型前缀）

        Args:
            row: 小区数据行
            network_type: 网络类型 ('LTE' 或 'NR')

        Returns:
            str: 小区唯一标识（无LTE/NR前缀）
        """
        if network_type == 'LTE':
            # LTE使用eNodeB ID + Cell ID
            enodeb_id = str(row.get('enodeb_id', '')).strip()
            cell_id = str(row.get('cell_id', '')).strip()
            return f"{enodeb_id}_{cell_id}"
        else:  # NR
            # NR使用MCC + MNC + gNodeB ID + Cell ID
            mcc = str(row.get('mcc', '')).strip()
            mnc = str(row.get('mnc', '')).strip()
            gnodeb_id = str(row.get('enodeb_id', '')).strip()
            cell_id = str(row.get('cell_id', '')).strip()
            return f"{mcc}_{mnc}_{gnodeb_id}_{cell_id}"

    def get_neighbor_planning_cell_key(self, row: pd.Series, network_type: str) -> str:
        """
        获取用于邻区规划的小区唯一标识键（不考虑移动网络码）
        NR网络仅使用gNodeB ID + Cell ID，LTE网络使用eNodeB ID + Cell ID

        Args:
            row: 小区数据行
            network_type: 网络类型 ('LTE' 或 'NR')

        Returns:
            str: 用于邻区规划的小区唯一标识
        """
        if network_type == 'LTE':
            # LTE使用eNodeB ID + Cell ID
            enodeb_id = str(row.get('enodeb_id', '')).strip()
            cell_id = str(row.get('cell_id', '')).strip()
            return f"{enodeb_id}_{cell_id}"
        else:  # NR
            # NR使用gNodeB ID + Cell ID (邻区规划时不考虑移动网络码)
            gnodeb_id = str(row.get('enodeb_id', '')).strip()
            cell_id = str(row.get('cell_id', '')).strip()
            return f"{gnodeb_id}_{cell_id}"

    def get_cell_location(self, row: pd.Series, network_type: str) -> tuple:
        """
        获取小区的经纬度

        Args:
            row: 小区数据行
            network_type: 网络类型 ('LTE' 或 'NR')

        Returns:
            tuple: (纬度, 经度)
        """
        lat = row.get('lat', None)
        lon = row.get('lon', None)

        if pd.isna(lat) or pd.isna(lon):
            return None, None

        try:
            return float(lat), float(lon)
        except (ValueError, TypeError):
            return None, None

    def calculate_azimuth_angle(self, lat1: float, lon1: float, lat2: float, lon2: float) -> float:
        """
        计算从点1到点2的方位角（度）
        北为0度，顺时针方向
        """
        import math
        # 将经纬度从度转换为弧度
        lat1 = math.radians(lat1)
        lon1 = math.radians(lon1)
        lat2 = math.radians(lat2)
        lon2 = math.radians(lon2)

        # 计算方位角
        dlon = lon2 - lon1
        y = math.sin(dlon) * math.cos(lat2)
        x = math.cos(lat1) * math.sin(lat2) - math.sin(lat1) * math.cos(lat2) * math.cos(dlon)
        bearing = math.atan2(y, x)

        # 将弧度转换为度
        bearing = math.degrees(bearing)

        # 将结果标准化到0-360度范围
        bearing = (bearing + 360) % 360

        return bearing

    def calculate_angle_difference(self, azimuth1: float, azimuth2: float) -> float:
        """
        计算两个方向角之间的差值（度）
        返回0-180度之间的最小角度差
        """
        import math
        diff = abs(azimuth1 - azimuth2)
        # 返回最小角度差（考虑到角度是循环的）
        return min(diff, 360 - diff)

    def estimate_distance_by_enodeb_id(self, enodeb1: int, enodeb2: int) -> float:
        """
        根据基站ID估算距离（当没有经纬度时）
        基站ID相近的认为距离较近
        """
        if enodeb1 is None or enodeb2 is None:
            return float('inf')

        # 计算基站ID的差值，转换为估算距离
        id_diff = abs(enodeb1 - enodeb2)

        # 基站ID每相差1000，估算为1公里距离
        estimated_distance = id_diff / 1000.0

        # 设置最大估算距离为10公里
        return min(estimated_distance, 10.0)

    def get_cell_info(self, row: pd.Series, network_type: str) -> dict:
        """
        获取小区的基本信息

        Args:
            row: 小区数据行
            network_type: 网络类型 ('LTE' 或 'NR')

        Returns:
            dict: 小区基本信息
        """
        cell_name = str(row.get('cell_name', '')).strip()
        if not cell_name or cell_name == '':
            enodeb_id = str(row.get('enodeb_id', 'unknown'))
            cell_id = str(row.get('cell_id', 'unknown'))
            cell_name = f"{network_type}_{enodeb_id}_{cell_id}"

        earfcn = row.get('earfcn_dl', '')
        pci = row.get('pci', '')
        antenna_azimuth = row.get('antenna_azimuth', '')  # 获取天线方向角

        return {
            'cell_name': cell_name,
            'earfcn': earfcn,
            'pci': pci,
            'antenna_azimuth': antenna_azimuth
        }

    def get_enodeb_id(self, row: pd.Series, network_type: str):
        """
        获取基站ID - 使用标准化列名
        """
        try:
            # 使用预处理后的标准化列名
            enodeb_id = row.get('enodeb_id', None)
            return int(enodeb_id) if enodeb_id is not None and not pd.isna(enodeb_id) else None
        except (ValueError, TypeError):
            return None

    def _find_column_by_fuzzy_match(self, df: pd.DataFrame, keyword: str) -> str:
        """
        通过关键词模糊查找列名，支持部分匹配
        例如：列名为"子网ID SubNetwork ID long[0...128]"，只用"子网ID"即可匹配

        Args:
            df: 数据框
            keyword: 关键词

        Returns:
            匹配的列名，如果找不到返回None
        """
        # 首先尝试精确匹配
        if keyword in df.columns:
            return keyword

        # 然后尝试部分匹配 - 检查关键词是否包含在列名中
        for col in df.columns:
            if keyword in col:
                return col

        # 最后尝试前缀匹配作为备选
        return self._find_column_by_prefix(df, keyword)

    def _find_column_by_prefix(self, df: pd.DataFrame, prefix: str) -> str:
        """
        通过前缀模糊查找列名

        Args:
            df: 数据框
            prefix: 列名前缀

        Returns:
            匹配的列名，如果找不到返回None
        """
        for col in df.columns:
            if col.startswith(prefix):
                return col
        return None

    def plan_neighbors_for_network(self, source_network: str, target_network: str) -> pd.DataFrame:
        """
        为指定网络类型规划邻区关系（使用覆盖圆算法）

        Args:
            source_network: 源网络类型 ('LTE' 或 'NR')
            target_network: 目标网络类型 ('LTE' 或 'NR')

        Returns:
            pd.DataFrame: 邻区关系数据
        """
        source_cells = self.lte_cells if source_network == 'LTE' else self.nr_cells
        target_cells = self.lte_cells if target_network == 'LTE' else self.nr_cells

        if source_cells is None or target_cells is None:
            print(f"缺少{source_network}或{target_network}小区数据")
            return pd.DataFrame()

        print(f"\n开始规划 {source_network} 到 {target_network} 的邻区关系...")

        neighbor_relations = []

        # 遍历源小区
        for idx1, source_row in source_cells.iterrows():
            source_key = self.get_cell_key(source_row, source_network)
            source_lat, source_lon = self.get_cell_location(source_row, source_network)
            source_info = self.get_cell_info(source_row, source_network)

            # 获取源小区基站ID
            source_enodeb_id = self.get_enodeb_id(source_row, source_network)

            # 获取源小区的方向角和覆盖距离
            source_azimuth = source_info.get('antenna_azimuth', None)
            try:
                source_azimuth = float(source_azimuth) if source_azimuth is not None and str(source_azimuth).strip() != '' else None
            except (ValueError, TypeError):
                source_azimuth = None

            # 假设默认覆盖距离为1000米
            source_coverage_distance = 1000.0

            # 存储候选邻区
            candidate_neighbors = []

            # 遍历目标小区
            for idx2, target_row in target_cells.iterrows():
                target_key = self.get_cell_key(target_row, target_network)
                target_lat, target_lon = self.get_cell_location(target_row, target_network)
                target_info = self.get_cell_info(target_row, target_network)

                # 获取目标小区基站ID
                target_enodeb_id = self.get_enodeb_id(target_row, target_network)

                # 获取目标小区的方向角和覆盖距离
                target_azimuth = target_info.get('antenna_azimuth', None)
                try:
                    target_azimuth = float(target_azimuth) if target_azimuth is not None and str(target_azimuth).strip() != '' else None
                except (ValueError, TypeError):
                    target_azimuth = None

                # 假设默认覆盖距离为1000米
                target_coverage_distance = 1000.0

                # 跳过相同小区
                if source_key == target_key:
                    continue

                # 使用覆盖圆算法判断是否为邻区
                is_neighbor = False
                distance = self.calculate_distance(source_lat, source_lon, target_lat, target_lon)  # 计算源小区到目标小区的距离

                try:
                    # 计算源小区的覆盖圆心和半径
                    if source_azimuth is not None:
                        source_center_lat, source_center_lon = self.calculate_coverage_circle_center(
                            source_lat, source_lon, source_azimuth,
                            self.coverage_distance_factor, source_coverage_distance)
                        source_radius = self.calculate_coverage_radius(
                            source_coverage_distance, self.coverage_radius_factor)
                    else:
                        # 如果没有方位角，则以当前位置为中心
                        source_center_lat, source_center_lon = source_lat, source_lon
                        source_radius = self.calculate_coverage_radius(
                            source_coverage_distance, self.coverage_radius_factor)

                    # 计算目标小区的覆盖圆心和半径
                    if target_azimuth is not None:
                        target_center_lat, target_center_lon = self.calculate_coverage_circle_center(
                            target_lat, target_lon, target_azimuth,
                            self.coverage_distance_factor, target_coverage_distance)
                        target_radius = self.calculate_coverage_radius(
                            target_coverage_distance, self.coverage_radius_factor)
                    else:
                        # 如果没有方位角，则以当前位置为中心
                        target_center_lat, target_center_lon = target_lat, target_lon
                        target_radius = self.calculate_coverage_radius(
                            target_coverage_distance, self.coverage_radius_factor)

                    # 判断覆盖圆是否相交
                    is_neighbor = self.are_coverage_circles_intersecting(
                        source_center_lat, source_center_lon, source_radius,
                        target_center_lat, target_center_lon, target_radius)
                except Exception as e:
                    print(f"覆盖圆计算错误: {e}")
                    # 如果覆盖圆计算失败，不添加为邻区
                    is_neighbor = False

                # 如果满足邻区条件，则添加到候选列表
                if is_neighbor:
                    # 计算从源小区到目标小区的方位角
                    source_to_target_azimuth = self.calculate_azimuth_angle(source_lat, source_lon, target_lat, target_lon)

                    # 计算角度差 - 用于确定方向优先级
                    angle_diff = None
                    if source_azimuth is not None and target_azimuth is not None:
                        # 理想情况下，两个小区应该方向对打（相差约180度）
                        facing_diff = self.calculate_angle_difference(source_azimuth, (target_azimuth + 180) % 360)
                        # 或者目标小区方向应该朝向源小区
                        pointing_diff = self.calculate_angle_difference(target_azimuth, source_to_target_azimuth)
                        # 采用较小的角度差作为参考
                        angle_diff = min(facing_diff, pointing_diff)

                    candidate_neighbors.append({
                        'source_key': source_key,
                        'target_key': target_key,
                        'distance': distance,
                        'angle_diff': angle_diff,  # 方向角度差
                        'source_cell_name': source_info['cell_name'],
                        'target_cell_name': target_info['cell_name'],
                        'source_earfcn': source_info['earfcn'],
                        'target_earfcn': target_info['earfcn'],
                        'source_pci': source_info['pci'],
                        'target_pci': target_info['pci'],
                        'source_lat': source_lat if source_lat is not None else 0,
                        'source_lon': source_lon if source_lon is not None else 0,
                        'target_lat': target_lat if target_lat is not None else 0,
                        'target_lon': target_lon if target_lon is not None else 0,
                        'source_enodeb_id': source_enodeb_id,
                        'target_enodeb_id': target_enodeb_id
                    })

            # 使用智能选择算法按扇区分部选择邻区
            if candidate_neighbors:
                # 按评分排序，选择评分最高的max_neighbors个邻区
                for neighbor in candidate_neighbors:
                    score = self._calculate_neighbor_score(neighbor)
                    neighbor['score'] = score

                # 按评分排序（评分越高优先级越高）
                candidate_neighbors.sort(key=lambda x: x['score'], reverse=True)
                selected_neighbors = candidate_neighbors[:self.max_neighbors]

                # 添加到结果中
                neighbor_relations.extend(selected_neighbors)

                if len(selected_neighbors) > 0:
                    print(f"  {source_info['cell_name']}: 规划 {len(selected_neighbors)} 个邻区")

        # 转换为DataFrame
        if neighbor_relations:
            result_df = pd.DataFrame(neighbor_relations)
            print(f"\n{source_network} 到 {target_network} 邻区关系规划完成，共 {len(result_df)} 条关系")
            return result_df
        else:
            print(f"\n{source_network} 到 {target_network} 没有找到符合条件的邻区关系")
            return pd.DataFrame()

    def save_neighbor_results(self, nr_to_nr_df: pd.DataFrame, lte_to_lte_df: pd.DataFrame,
                            nr_to_lte_df: pd.DataFrame) -> bool:
        """
        保存邻区规划结果到文件

        Args:
            nr_to_nr_df: NR到NR邻区关系
            lte_to_lte_df: LTE到LTE邻区关系
            nr_to_lte_df: NR到LTE邻区关系

        Returns:
            bool: 保存成功返回True
        """
        try:
            timestamp = self.generate_timestamp_suffix()
            output_dir = "输出文件"
            os.makedirs(output_dir, exist_ok=True)

            saved_files = []

            # 保存NR到NR邻区关系
            if not nr_to_nr_df.empty:
                filename = f"{output_dir}/邻区关系_NR到NR_{timestamp}.xlsx"
                nr_to_nr_df.to_excel(filename, index=False, sheet_name='NR到NR邻区关系')
                saved_files.append(filename)
                print(f"✅ NR到NR邻区关系已保存: {filename}")

            # 保存LTE到LTE邻区关系
            if not lte_to_lte_df.empty:
                filename = f"{output_dir}/邻区关系_LTE到LTE_{timestamp}.xlsx"
                lte_to_lte_df.to_excel(filename, index=False, sheet_name='LTE到LTE邻区关系')
                saved_files.append(filename)
                print(f"✅ LTE到LTE邻区关系已保存: {filename}")

            # 保存NR到LTE邻区关系
            if not nr_to_lte_df.empty:
                filename = f"{output_dir}/邻区关系_NR到LTE_{timestamp}.xlsx"
                nr_to_lte_df.to_excel(filename, index=False, sheet_name='NR到LTE邻区关系')
                saved_files.append(filename)
                print(f"✅ NR到LTE邻区关系已保存: {filename}")

            if saved_files:
                print(f"\n邻区规划结果保存完成，共生成 {len(saved_files)} 个文件")
                return True
            else:
                print("没有生成任何邻区关系文件")
                return False

        except Exception as e:
            print(f"保存邻区规划结果失败: {e}")
            return False

    def load_cells_to_plan(self, cells_file: str) -> tuple:
        """
        加载待规划小区清单

        Args:
            cells_file: 待规划小区文件路径

        Returns:
            tuple: (lte_cells_df, nr_cells_df)
        """
        try:
            xl_file = pd.ExcelFile(cells_file)
            sheet_names = xl_file.sheet_names

            lte_cells = None
            nr_cells = None

            # 查找LTE工作表
            lte_sheets = [name for name in sheet_names if 'LTE' in name.upper()]
            if lte_sheets:
                lte_cells = pd.read_excel(cells_file, sheet_name=lte_sheets[0])
                print(f"待规划LTE小区: {len(lte_cells)} 个")

            # 查找NR工作表
            nr_sheets = [name for name in sheet_names if 'NR' in name.upper()]
            if nr_sheets:
                nr_cells = pd.read_excel(cells_file, sheet_name=nr_sheets[0])
                print(f"待规划NR小区: {len(nr_cells)} 个")

            return lte_cells, nr_cells

        except Exception as e:
            print(f"加载待规划小区文件失败: {e}")
            return None, None

    def get_cell_details_from_params(self, cells_to_plan: pd.DataFrame, network_type: str) -> pd.DataFrame:
        """
        从全量工参文件中获取待规划小区的详细信息

        Args:
            cells_to_plan: 待规划小区清单
            network_type: 网络类型 ('LTE' 或 'NR')

        Returns:
            pd.DataFrame: 包含详细信息的待规划小区数据
        """
        if cells_to_plan is None or len(cells_to_plan) == 0:
            return pd.DataFrame()

        source_cells = self.lte_cells if network_type == 'LTE' else self.nr_cells
        if source_cells is None or len(source_cells) == 0:
            return pd.DataFrame()

        print(f"正在从全量工参中获取{network_type}待规划小区的详细信息...")

        detailed_cells = []

        for _, plan_row in cells_to_plan.iterrows():
            if network_type == 'LTE':
                # 从待规划清单中获取关键信息
                plan_enodeb_id = str(plan_row.get('eNodeBID', '')).strip()
                plan_cell_id = str(plan_row.get('CellID', '')).strip()

                # 在全量工参中查找匹配的小区 - 使用标准化列名
                for _, param_row in source_cells.iterrows():
                    param_enodeb_id = str(param_row.get('enodeb_id', '')).strip()
                    param_cell_id = str(param_row.get('cell_id', '')).strip()

                    if plan_enodeb_id == param_enodeb_id and plan_cell_id == param_cell_id:
                        detailed_cells.append(param_row)
                        break
            else:  # NR
                # 从待规划清单中获取关键信息
                plan_gnodeb_id = str(plan_row.get('gNodeBID', '')).strip()
                plan_cell_id = str(plan_row.get('CellID', '')).strip()

                # 在全量工参中查找匹配的小区 - 使用标准化列名
                for _, param_row in source_cells.iterrows():
                    param_gnodeb_id = str(param_row.get('enodeb_id', '')).strip()
                    param_cell_id = str(param_row.get('cell_id', '')).strip()

                    if plan_gnodeb_id == param_gnodeb_id and plan_cell_id == param_cell_id:
                        detailed_cells.append(param_row)
                        break

        if detailed_cells:
            result_df = pd.DataFrame(detailed_cells)
            print(f"成功匹配 {len(result_df)} 个{network_type}待规划小区")
            return result_df
        else:
            print(f"未找到匹配的{network_type}待规划小区")
            return pd.DataFrame()

    def plan_neighbors_for_target_cells(self, target_cells: pd.DataFrame, target_network: str,
                                      source_network: str, all_param_cells: pd.DataFrame) -> pd.DataFrame:
        """
        为指定的待规划小区寻找邻区关系（智能优化版本）

        Args:
            target_cells: 待规划小区（从待规划清单中读取的）
            target_network: 待规划小区的网络类型
            source_network: 寻找邻区的源网络类型
            all_param_cells: 全量工参中的所有小区

        Returns:
            pd.DataFrame: 邻区关系数据
        """
        if target_cells is None or len(target_cells) == 0:
            return pd.DataFrame()

        if all_param_cells is None or len(all_param_cells) == 0:
            return pd.DataFrame()

        relation_name = f"{target_network}到{source_network}"
        print(f"\n开始规划 {relation_name} 的邻区关系...")

        neighbor_relations = []

        # 预计算所有源小区的位置信息，避免重复计算
        print("正在预处理源小区位置信息...")
        source_cell_info = []
        for idx, source_row in all_param_cells.iterrows():
            source_key = self.get_neighbor_planning_cell_key(source_row, source_network)
            source_lat, source_lon = self.get_cell_location(source_row, source_network)
            source_info = self.get_cell_info(source_row, source_network)
            source_enodeb_id = self.get_enodeb_id(source_row, source_network)

            # 获取方向角和覆盖距离（如果有的话）
            source_azimuth = source_info.get('antenna_azimuth', None)
            try:
                source_azimuth = float(source_azimuth) if source_azimuth is not None and str(source_azimuth).strip() != '' else None
            except (ValueError, TypeError):
                source_azimuth = None

            # 尝试从数据中获取覆盖距离（如果存在相关列）
            coverage_distance = 1000.0  # 默认1000米，如果无法获取则使用默认值
            # 这里可以根据实际数据列进行调整

            source_cell_info.append({
                'key': source_key,
                'lat': source_lat,
                'lon': source_lon,
                'info': source_info,
                'enodeb_id': source_enodeb_id,
                'azimuth': source_azimuth,
                'coverage_distance': coverage_distance
            })

        print(f"预处理完成，共 {len(source_cell_info)} 个源小区")

        # 遍历待规划小区
        for idx1, target_row in target_cells.iterrows():
            target_key = self.get_neighbor_planning_cell_key(target_row, target_network)
            target_lat, target_lon = self.get_cell_location(target_row, target_network)
            target_info = self.get_cell_info(target_row, target_network)

            if target_lat is None or target_lon is None:
                continue

            # 获取目标小区基站ID
            target_enodeb_id = self.get_enodeb_id(target_row, target_network)

            # 获取目标小区的方向角和覆盖距离
            target_azimuth = target_info.get('antenna_azimuth', None)
            try:
                target_azimuth = float(target_azimuth) if target_azimuth is not None and str(target_azimuth).strip() != '' else None
            except (ValueError, TypeError):
                target_azimuth = None

            # 默认覆盖距离，如果数据中有具体值可以进一步获取
            target_coverage_distance = 1000.0  # 默认1000米

            # 存储候选邻区
            candidate_neighbors = []

            # 遍历预处理的源小区信息
            for source_data in source_cell_info:
                source_key = source_data['key']
                source_lat = source_data['lat']
                source_lon = source_data['lon']
                source_info = source_data['info']
                source_enodeb_id = source_data['enodeb_id']
                source_azimuth = source_data['azimuth']
                source_coverage_distance = source_data['coverage_distance']

                # 跳过相同小区
                if target_key == source_key:
                    continue

                # 跳过无效位置
                if source_lat is None or source_lon is None:
                    continue

                # 使用覆盖圆算法判断
                is_neighbor = False
                distance = self.calculate_distance(target_lat, target_lon, source_lat, source_lon)

                try:
                    # 计算目标小区的覆盖圆心和半径
                    if target_azimuth is not None:
                        target_center_lat, target_center_lon = self.calculate_coverage_circle_center(
                            target_lat, target_lon, target_azimuth,
                            self.coverage_distance_factor, target_coverage_distance)
                        target_radius = self.calculate_coverage_radius(
                            target_coverage_distance, self.coverage_radius_factor)
                    else:
                        # 如果没有方位角，则以当前位置为中心
                        target_center_lat, target_center_lon = target_lat, target_lon
                        target_radius = self.calculate_coverage_radius(
                            target_coverage_distance, self.coverage_radius_factor)

                    # 计算源小区的覆盖圆心和半径
                    if source_azimuth is not None:
                        source_center_lat, source_center_lon = self.calculate_coverage_circle_center(
                            source_lat, source_lon, source_azimuth,
                            self.coverage_distance_factor, source_coverage_distance)
                        source_radius = self.calculate_coverage_radius(
                            source_coverage_distance, self.coverage_radius_factor)
                    else:
                        # 如果没有方位角，则以当前位置为中心
                        source_center_lat, source_center_lon = source_lat, source_lon
                        source_radius = self.calculate_coverage_radius(
                            source_coverage_distance, self.coverage_radius_factor)

                    # 判断覆盖圆是否相交
                    is_neighbor = self.are_coverage_circles_intersecting(
                        target_center_lat, target_center_lon, target_radius,
                        source_center_lat, source_center_lon, source_radius)
                except Exception as e:
                    print(f"覆盖圆计算错误: {e}")
                    # 如果覆盖圆计算失败，不添加为邻区（覆盖圆算法失效时返回False）
                    is_neighbor = False

                # 如果满足邻区条件，则计算方位角和角度差等信息
                if is_neighbor:
                    # 计算从目标小区到源小区的方位角
                    target_to_source_azimuth = self.calculate_azimuth_angle(target_lat, target_lon, source_lat, source_lon)

                    # 计算角度差 - 用于确定方向优先级
                    angle_diff = None
                    if target_azimuth is not None and source_azimuth is not None:
                        # 理想情况下，两个小区应该方向对打（相差约180度）
                        facing_diff = self.calculate_angle_difference(target_azimuth, (source_azimuth + 180) % 360)
                        # 或者源小区方向应该朝向目标小区
                        pointing_diff = self.calculate_angle_difference(source_azimuth, target_to_source_azimuth)
                        # 采用较小的角度差作为参考
                        angle_diff = min(facing_diff, pointing_diff)

                    candidate_neighbors.append({
                        'source_key': target_key,  # 待规划小区作为源
                        'target_key': source_key,   # 全量工参中的小区作为目标
                        'distance': distance,
                        'target_to_source_azimuth': target_to_source_azimuth,  # 从目标到源的方位角
                        'angle_diff': angle_diff,  # 方向角度差
                        'source_cell_name': target_info['cell_name'],
                        'target_cell_name': source_info['cell_name'],
                        'source_earfcn': target_info['earfcn'],
                        'target_earfcn': source_info['earfcn'],
                        'source_pci': target_info['pci'],
                        'target_pci': source_info['pci'],
                        'source_lat': target_lat,
                        'source_lon': target_lon,
                        'target_lat': source_lat,
                        'target_lon': source_lon,
                        'source_enodeb_id': target_enodeb_id,
                        'target_enodeb_id': source_enodeb_id
                    })

            # 对候选邻区进行去重：基于source_key和target_key确保邻区关系唯一
            if candidate_neighbors:
                # 转换为DataFrame以便去重
                candidate_df = pd.DataFrame(candidate_neighbors)
                original_count = len(candidate_df)

                if 'source_key' in candidate_df.columns and 'target_key' in candidate_df.columns:
                    candidate_df = candidate_df.drop_duplicates(subset=['source_key', 'target_key'], keep='first')
                    deduplicated_count = len(candidate_df)
                    if original_count > deduplicated_count:
                        print(f"  {target_info['cell_name']}: 候选邻区去重 - 从 {original_count} 个候选中移除 {original_count - deduplicated_count} 个重复项")

                # 转换回列表
                candidate_neighbors = candidate_df.to_dict('records')

                # 实现智能邻区选择算法：按扇区分部选择
                selected_neighbors = self._select_neighbors_intelligently(
                    candidate_neighbors, self.max_neighbors, target_info['cell_name'])

                # 添加到结果中
                neighbor_relations.extend(selected_neighbors)

                if len(selected_neighbors) > 0:
                    print(f"  {target_info['cell_name']}: 规划 {len(selected_neighbors)} 个邻区")

        # 转换为DataFrame
        if neighbor_relations:
            result_df = pd.DataFrame(neighbor_relations)
            print(f"\n{relation_name} 邻区关系规划完成，共 {len(result_df)} 条唯一关系")
            return result_df
        else:
            print(f"\n{relation_name} 没有找到符合条件的邻区关系")
            return pd.DataFrame()

    def _select_neighbors_intelligently(self, candidate_neighbors, max_neighbors, cell_name):
        """
        智能选择邻区：考虑方向、距离和扇区分布
        """
        if len(candidate_neighbors) <= max_neighbors:
            # 如果候选邻区数量不超过最大值，直接返回所有邻区
            return candidate_neighbors

        # 按方向和距离计算综合评分
        for neighbor in candidate_neighbors:
            score = self._calculate_neighbor_score(neighbor)
            neighbor['score'] = score

        # 按评分排序（评分越高优先级越高）
        candidate_neighbors.sort(key=lambda x: x['score'], reverse=True)

        # 选择评分最高的邻区，但考虑扇区分布
        selected = []

        # 将360度分成若干扇区（例如6个扇区，每扇区60度）
        sector_size = 60  # 扇区大小（度）
        sectors = {}

        # 将候选邻区按扇区分组
        for neighbor in candidate_neighbors:
            if 'target_to_source_azimuth' in neighbor and neighbor['target_to_source_azimuth'] is not None:
                azimuth = neighbor['target_to_source_azimuth']
                sector_idx = int(azimuth // sector_size)
                # 处理边界情况：将360度归入第0扇区
                sector_idx = sector_idx % (360 // sector_size)
                if sector_idx not in sectors:
                    sectors[sector_idx] = []
                sectors[sector_idx].append(neighbor)

        # 动态计算每个扇区应分配的邻区数量，考虑各扇区的候选数量
        total_candidates = len(candidate_neighbors)
        sector_allocation = {}

        for sector_idx in sectors:
            sector_candidate_count = len(sectors[sector_idx])
            # 根据扇区内候选邻区数量按比例分配，但保证最少有一个邻区
            allocated_count = max(1, int(max_neighbors * (sector_candidate_count / total_candidates)))
            sector_allocation[sector_idx] = allocated_count

        # 确保分配总数不超过最大值
        total_allocated = sum(sector_allocation.values())
        if total_allocated > max_neighbors:
            # 按比例缩减
            scale_factor = max_neighbors / total_allocated
            for sector_idx in sector_allocation:
                sector_allocation[sector_idx] = max(1, int(sector_allocation[sector_idx] * scale_factor))

        # 从每个扇区选择评分最高的邻区
        remaining_candidates = [n for n in candidate_neighbors]
        for sector_idx, sector_neighbors in sectors.items():
            sector_neighbors.sort(key=lambda x: x['score'], reverse=True)
            # 从每个扇区选择相应数量的邻区
            allocated_count = sector_allocation.get(sector_idx, 0)
            selected_from_sector = sector_neighbors[:allocated_count]
            selected.extend(selected_from_sector)

            # 从剩余候选中移除已选的
            for n in selected_from_sector:
                if n in remaining_candidates:
                    remaining_candidates.remove(n)

        # 如果还没有达到最大邻区数，从剩余的邻区中选择评分最高的
        remaining_needed = max_neighbors - len(selected)
        if remaining_needed > 0 and remaining_candidates:
            remaining_candidates.sort(key=lambda x: x['score'], reverse=True)
            selected.extend(remaining_candidates[:remaining_needed])

        # 确保不超过最大邻区数
        return selected[:max_neighbors]

    def _calculate_neighbor_score(self, neighbor):
        """
        根据距离、方向角度等因素计算邻区优先级评分
        """
        import math
        distance = neighbor['distance']
        angle_diff = neighbor.get('angle_diff', float('inf'))  # 角度差，越小越好

        # 检查是否为同站小区（经纬度相同或非常接近）
        # tolerance设为0.0001度（约10米精度）来判断是否为同站
        tolerance = 0.0001
        is_same_site = (abs(neighbor['source_lat'] - neighbor['target_lat']) < tolerance and
                        abs(neighbor['source_lon'] - neighbor['target_lon']) < tolerance)

        # 如果是同站小区，给予最高优先级
        if is_same_site:
            # 给同站小区一个非常高基础分，确保它们总是优先选择
            same_site_bonus = 1000  # 大于任何其他评分的值
            # 同站小区仍然会根据距离和方向进行微调
            distance_score = 100 * math.exp(-distance / 5.0)  # 调整参数5.0以控制距离衰减速度
            if angle_diff is None or angle_diff == float('inf'):
                direction_score = 0
            else:
                # 角度差越小得分越高，0度差得满分，180度差得0分
                direction_score = max(0, (180 - angle_diff) / 180 * 50)  # 最多50分
            total_score = same_site_bonus + distance_score + direction_score
        else:
            # 距离评分：距离越近得分越高
            # 使用负指数函数，使近距离的得分显著高于远距离
            distance_score = 100 * math.exp(-distance / 5.0)  # 调整参数5.0以控制距离衰减速度

            # 方向评分：方向对打或朝向目标小区的得分更高
            # angle_diff = 0 表示完美对打或朝向，angle_diff = 180 表示完全背向
            if angle_diff is None or angle_diff == float('inf'):
                # 如果无法计算方向角，则只考虑距离
                direction_score = 0
            else:
                # 角度差越小得分越高，0度差得满分，180度差得0分
                direction_score = max(0, (180 - angle_diff) / 180 * 50)  # 最多50分

            # 综合评分 = 距离评分 + 方向评分
            total_score = distance_score + direction_score

        return total_score


    def run_neighbor_planning(self, cells_file: str, params_file: str, planning_type: str) -> bool:
        """
        运行指定类型的邻区规划流程


        Args:
            cells_file: 待规划小区文件路径
            params_file: 全量工参文件路径
            planning_type: 规划类型 ('NR到NR', 'LTE到LTE', 'NR到LTE')

        Returns:
            bool: 规划成功返回True
        """
        print("\n" + "="*60)
        print(f"开始{planning_type}邻区关系规划")
        print("="*60)

        # 加载待规划小区清单
        cells_to_plan_lte, cells_to_plan_nr = self.load_cells_to_plan(cells_file)

        # 加载全量工参数据
        if not self.load_full_parameter_data(params_file):
            return False

        result_df = pd.DataFrame()

        if planning_type == 'NR到NR':
            if cells_to_plan_nr is not None and len(cells_to_plan_nr) > 0:
                # 获取待规划NR小区的详细信息
                detailed_nr_cells = self.get_cell_details_from_params(cells_to_plan_nr, 'NR')
                if len(detailed_nr_cells) > 0:
                    result_df = self.plan_neighbors_for_target_cells(detailed_nr_cells, 'NR', 'NR', self.nr_cells)
                else:
                    print("未能获取待规划NR小区的详细信息")
            else:
                print("没有找到待规划的NR小区")

        elif planning_type == 'LTE到LTE':
            if cells_to_plan_lte is not None and len(cells_to_plan_lte) > 0:
                # 获取待规划LTE小区的详细信息
                detailed_lte_cells = self.get_cell_details_from_params(cells_to_plan_lte, 'LTE')
                if len(detailed_lte_cells) > 0:
                    result_df = self.plan_neighbors_for_target_cells(detailed_lte_cells, 'LTE', 'LTE', self.lte_cells)
                else:
                    print("未能获取待规划LTE小区的详细信息")
            else:
                print("没有找到待规划的LTE小区")

        elif planning_type == 'NR到LTE':
            if cells_to_plan_nr is not None and len(cells_to_plan_nr) > 0:
                # 获取待规划NR小区的详细信息
                detailed_nr_cells = self.get_cell_details_from_params(cells_to_plan_nr, 'NR')
                if len(detailed_nr_cells) > 0:
                    result_df = self.plan_neighbors_for_target_cells(detailed_nr_cells, 'NR', 'LTE', self.lte_cells)
                else:
                    print("未能获取待规划NR小区的详细信息")
            else:
                print("没有找到待规划的NR小区")

        # 保存结果
        if planning_type == 'NR到NR':
            return self.save_neighbor_results(result_df, pd.DataFrame(), pd.DataFrame())
        elif planning_type == 'LTE到LTE':
            return self.save_neighbor_results(pd.DataFrame(), result_df, pd.DataFrame())
        elif planning_type == 'NR到LTE':
            return self.save_neighbor_results(pd.DataFrame(), pd.DataFrame(), result_df)

        return False


class LTENRPCIPlanner:
    def __init__(self, reuse_distance_km: float = 3.0, lte_inherit_mod3: bool = False,
                 nr_inherit_mod30: bool = False, network_type: str = "LTE", params_file: str = None,
                 pci_range: List[int] = None):
        """
        初始化LTE/NR分离式PCI规划工具

        Args:
            reuse_distance_km: 最小PCI复用距离（公里）
            lte_inherit_mod3: LTE小区是否继承原PCI的模3值
            nr_inherit_mod30: NR小区是否继承原PCI的模30值
            network_type: 当前处理的网络类型 ("LTE" 或 "NR")
            params_file: 使用的参数文件路径
            pci_range: 用户指定的PCI范围列表，如果为None则使用默认范围
        """
        self.reuse_distance_km = reuse_distance_km
        self.lte_inherit_mod3 = lte_inherit_mod3
        self.nr_inherit_mod30 = nr_inherit_mod30
        self.network_type = network_type
        self.params_file = params_file  # 存储参数文件路径

        # PCI范围根据网络类型设定，或使用用户指定的范围
        if pci_range is not None:
            # 验证用户指定的PCI范围是否符合网络类型标准
            if network_type == "LTE":
                # 验证LTE范围：0-503
                valid_range = list(range(0, 504))
                invalid_pci = [pci for pci in pci_range if pci not in valid_range]
                if invalid_pci:
                    print(f"警告: 用户指定的LTE PCI范围包含无效值 {invalid_pci}，将被过滤")
                    pci_range = [pci for pci in pci_range if pci in valid_range]
                self.pci_range = sorted(list(set(pci_range)))  # 去重并排序
            else:  # NR
                # 验证NR范围：0-1007
                valid_range = list(range(0, 1008))
                invalid_pci = [pci for pci in pci_range if pci not in valid_range]
                if invalid_pci:
                    print(f"警告: 用户指定的NR PCI范围包含无效值 {invalid_pci}，将被过滤")
                    pci_range = [pci for pci in pci_range if pci in valid_range]
                self.pci_range = sorted(list(set(pci_range)))  # 去重并排序
        else:
            # 使用默认范围
            if network_type == "LTE":
                self.pci_range = list(range(0, 504))  # LTE PCI范围 0-503
            else:  # NR
                self.pci_range = list(range(0, 1008))  # NR PCI范围 0-1007

        if network_type == "LTE":
            self.inherit_mod = lte_inherit_mod3
            self.mod_value = 3  # LTE使用mod3
            self.dual_mod_requirement = False  # LTE只需要模3约束
        else:  # NR
            self.inherit_mod = nr_inherit_mod30
            self.mod_value = 30  # NR使用mod30
            self.dual_mod_requirement = True   # NR需要同时满足模3和模30约束

        # 数据存储
        self.cells_to_plan = None
        self.target_cells = None  # 目标网络类型的小区数据
        self.all_cells_combined = None  # 合并的所有小区数据（用于干扰分析）

        # 性能优化缓存
        self.distance_cache = {}  # 距离计算结果缓存
        self.pci_validity_cache = {}  # PCI验证结果缓存
        self.same_site_cache = {}  # 同站点信息缓存

        # 失败原因统计
        self.failure_reasons = {
            'cell_not_found': [],
            'no_location': [],
            'reuse_distance_violation': [],
            'no_compliant_pci': [],
            'fallback_assignments': [],
            'same_site_mod_conflicts': []  # 同站点模值冲突统计
        }

        print(f"初始化{network_type}网络PCI规划工具")
        print(f"同频PCI最小复用距离: {reuse_distance_km}km (优先级最高)")
        print(f"PCI分配策略 (按优先级):")
        print(f"  1. 最小复用距离 (最高优先级) - 确保同频同PCI小区间距离≥{reuse_distance_km}km")
        print(f"  2. 同站点模{self.mod_value}冲突避免 (第二优先级) - 避免同基站小区模值相同")
        print(f"  3. PCI分布均衡性 (第三优先级) - 优选复用距离接近阈值的PCI")

        if network_type == "LTE":
            if lte_inherit_mod3:
                print(f"模3继承: 是 - 严格按照原PCI的mod3值分配")
            else:
                print(f"模3继承: 否 - 自由规划，不考虑mod3值")
        else:
            if nr_inherit_mod30:
                print(f"模30继承: 是 - 严格按照原PCI的mod30值分配")
            else:
                print(f"模30继承: 否 - 自由规划，不考虑mod30值")

        print(f"PCI范围: 0-{max(self.pci_range)}")
        print(f"核心规则: 只有同频(earfcnDl相同)且PCI相同的小区才需要遵循最小复用距离")

    def generate_timestamp_suffix(self) -> str:
        """
        生成时间后缀，格式：YYYYMMDD_HHMMSS
        """
        now = datetime.now()
        return now.strftime("%Y%m%d_%H%M%S")

    def convert_to_numeric(self, df: pd.DataFrame, columns: List[str]) -> pd.DataFrame:
        """
        转换指定列为数值格式，处理文本格式数据
        """
        df_copy = df.copy()

        for col in columns:
            if col in df_copy.columns:
                # 转换为字符串，处理各种数据类型
                df_copy[col] = df_copy[col].astype(str)

                # 移除空格和特殊字符
                df_copy[col] = df_copy[col].str.replace(' ', '').str.replace(',', '')

                # 处理 'nan', 'None', 空字符串等
                df_copy[col] = df_copy[col].replace(['nan', 'None', '', 'null'], np.nan)

                # 转换为数值
                df_copy[col] = pd.to_numeric(df_copy[col], errors='coerce')

        return df_copy

    def load_data(self, cells_file: str, params_file: str):
        """
        加载数据文件 - 支持分离式LTE/NR工作表
        """
        print(f"\n加载数据文件...")
        print(f"规划小区文件: {cells_file}")
        print(f"参数文件: {params_file}")
        print(f"使用的全量工参文件: {os.path.basename(params_file)}")  # 明确显示使用的参数文件


        try:
            # 加载规划小区数据 - 从对应的工作表读取
            worksheet_name = self.network_type  # "LTE" 或 "NR"
            print(f"正在从'{worksheet_name}'工作表加载待规划小区...")

            self.cells_to_plan = pd.read_excel(cells_file, sheet_name=worksheet_name)
            print(f"规划小区数据加载成功: {len(self.cells_to_plan)} 行")

            # 列名标准化检查
            if self.network_type == "NR" and 'gNodeBID' in self.cells_to_plan.columns:
                # NR网络应使用gNodeBID列，不需要映射
                print("NR网络使用gNodeBID列作为基站标识")
            elif self.network_type == "LTE" and 'eNodeBID' in self.cells_to_plan.columns:
                print("LTE网络使用eNodeBID列作为基站标识")

            # 加载参数文件，根据网络类型选择工作表
            if self.network_type == "LTE":
                params_worksheet = "LTE Project Parameters"
            else:
                params_worksheet = "NR Project Parameters"

            print(f"正在加载{params_worksheet}工作表...")

            # 读取目标网络类型的数据
            target_raw = pd.read_excel(params_file, sheet_name=params_worksheet)
            self.target_cells = self.preprocess_target_cells(target_raw)

            # 只使用当前网络类型的数据进行干扰分析（不跨网络类型）
            self.all_cells_combined = self.target_cells.copy()
            print(f"加载{self.network_type}小区数据: {len(self.all_cells_combined)} 个")

        except Exception as e:
            print(f"错误: 无法加载{self.network_type}工作表: {e}")
            raise

    def preprocess_target_cells(self, raw_df: pd.DataFrame) -> pd.DataFrame:
        """
        预处理目标网络类型的小区数据
        """
        if raw_df.empty:
            return pd.DataFrame()

        processed = pd.DataFrame()

        if self.network_type == "LTE":
            # LTE列映射
            processed['enodeb_id'] = raw_df['eNodeB标识\neNodeB ID\nlong:[0..1048575]']
            processed['cell_id'] = raw_df['小区标识\ncellLocalId\ninteger:[0~2147483647]']
            processed['cell_name'] = raw_df['小区名称\nuserLabel\nstring[0..128]']
            processed['pci'] = raw_df['物理小区识别码\nPCI\nlong:[0..503]']
            processed['lat'] = raw_df['小区纬度\neNodeB Latitude\ndouble:[-90..90]']
            processed['lon'] = raw_df['小区经度\neNodeB Longitude double:[-180..180]']
            processed['earfcn_dl'] = raw_df['下行链路的中心载频\nearfcnDl\ndouble Step：0.1 \nUnite：MHz']
        else:
            # NR列映射
            processed['enodeb_id'] = raw_df['gNodeB标识\ngNodeB ID\nLong:[0..1048575]']
            processed['cell_id'] = raw_df['小区标识\ncellLocalId\nInteger:[0~2147483647]']
            processed['cell_name'] = raw_df['小区名称\nCELL NAME\nString[0..128]']
            processed['pci'] = raw_df['物理小区识别码\nPCI\nLong:[0..1007]']
            processed['lat'] = raw_df['小区纬度\nCell  Latitude\nDouble:[-90..90]']
            processed['lon'] = raw_df['小区经度\nCell  Longitude Double:[-180..180]']
            processed['earfcn_dl'] = raw_df['填写SSB频点\nSSB Frequency\nDouble Step：0.01 \nUnite：MHz']

        processed['cell_type'] = self.network_type

        # 转换所有数值列（cell_name保持字符串格式）
        print(f"转换{self.network_type}小区数值格式...")
        numeric_cols = ['enodeb_id', 'cell_id', 'pci', 'lat', 'lon', 'earfcn_dl']
        processed = self.convert_to_numeric(processed, numeric_cols)

        # 过滤掉缺少关键信息的小区
        before_count = len(processed)

        # 记录被移除的小区信息
        invalid_cells = processed[processed['enodeb_id'].isna() | processed['cell_id'].isna()]

        processed = processed.dropna(subset=['enodeb_id', 'cell_id'])
        after_count = len(processed)

        if before_count != after_count:
            print(f"{self.network_type}数据清理: 移除了 {before_count - after_count} 个缺少关键信息的记录")
            if not invalid_cells.empty:
                print("被移除的小区信息:")
                for idx, cell in invalid_cells.iterrows():
                    enodeb_id = cell.get('enodeb_id', '缺失')
                    cell_id = cell.get('cell_id', '缺失')
                    cell_name = cell.get('cell_name', '未知')
                    print(f"  - 基站ID: {enodeb_id}, 小区ID: {cell_id}, 小区名称: {cell_name}")

        print(f"{self.network_type}小区预处理完成，有效小区数量: {len(processed)}")
        return processed


    def calculate_distance_vectorized(self, lat1, lon1, lat2_array, lon2_array):
        """
        向量化距离计算（带缓存优化）
        """
        # 生成缓存键
        cache_key = (round(lat1, 6), round(lon1, 6),
                    tuple(np.round(lat2_array, 6)), tuple(np.round(lon2_array, 6)))

        if cache_key in self.distance_cache:
            return self.distance_cache[cache_key]

        # 转换为弧度
        lat1_rad = math.radians(lat1)
        lon1_rad = math.radians(lon1)
        lat2_rad = np.radians(lat2_array)
        lon2_rad = np.radians(lon2_array)

        # Haversine公式
        dlat = lat2_rad - lat1_rad
        dlon = lon2_rad - lon1_rad
        a = np.sin(dlat/2)**2 + math.cos(lat1_rad) * np.cos(lat2_rad) * np.sin(dlon/2)**2
        c = 2 * np.arcsin(np.sqrt(a))

        result = c * 6371  # 地球半径（公里）
        self.distance_cache[cache_key] = result
        return result

    def calculate_distance(self, lat1: float, lon1: float, lat2: float, lon2: float) -> float:
        """
        计算两点之间的距离（公里）
        """
        # 转换为弧度
        lat1_rad = math.radians(lat1)
        lon1_rad = math.radians(lon1)
        lat2_rad = math.radians(lat2)
        lon2_rad = math.radians(lon2)

        # Haversine公式
        dlat = lat2_rad - lat1_rad
        dlon = lon2_rad - lon1_rad
        a = math.sin(dlat/2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon/2)**2
        c = 2 * math.asin(math.sqrt(a))

        result = c * 6371  # 地球半径（公里）
        return result

    def get_cell_info(self, enodeb_id: int, cell_id: int) -> Tuple[Optional[pd.Series], str]:
        """
        获取小区详细信息

        Returns:
            (cell_info, status)
        """
        # 在所有小区中查找（包含完整的位置信息）
        mask = (self.all_cells_combined['enodeb_id'] == enodeb_id) & (self.all_cells_combined['cell_id'] == cell_id)
        matches = self.all_cells_combined[mask]

        if matches.empty:
            return None, 'cell_not_found'

        cell_info = matches.iloc[0]

        # 检查位置信息
        if pd.isna(cell_info['lat']) or pd.isna(cell_info['lon']):
            return cell_info, 'no_location'

        return cell_info, 'success'

    def validate_pci_reuse_distance(self, candidate_pci: int, target_lat: float, target_lon: float,
                                   target_earfcn: float, exclude_enodeb: int = None, exclude_cell: int = None) -> Tuple[bool, float]:
        """
        验证候选PCI是否满足最小复用距离要求（带缓存优化）
        """
        if self.all_cells_combined is None or self.all_cells_combined.empty:
            return True, float('inf')

        # 生成缓存键 - 关键修复：包含复用距离参数
        cache_key = (candidate_pci, round(target_lat, 6), round(target_lon, 6),
                    round(target_earfcn, 2), exclude_enodeb, exclude_cell,
                    round(self.reuse_distance_km, 2))

        if cache_key in self.pci_validity_cache:
            return self.pci_validity_cache[cache_key]

        # 只查找同频同PCI的小区
        # 修复：应该从 all_cells_combined 中查找，因为包含所有已分配PCI的小区
        # 移除cell_type过滤条件，因为在实际使用中cell_type值可能不一致
        same_freq_same_pci_cells = self.all_cells_combined[
            (self.all_cells_combined['pci'] == candidate_pci) &
            (self.all_cells_combined['pci'].notna()) &
            (self.all_cells_combined['earfcn_dl'] == target_earfcn) &  # 同频条件
            (self.all_cells_combined['earfcn_dl'].notna())
        ].copy()

        # 排除当前小区自己
        if exclude_enodeb is not None and exclude_cell is not None:
            same_freq_same_pci_cells = same_freq_same_pci_cells[
                ~((same_freq_same_pci_cells['enodeb_id'] == exclude_enodeb) &
                  (same_freq_same_pci_cells['cell_id'] == exclude_cell))
            ]

        # 过滤有效位置的小区
        same_freq_same_pci_cells = same_freq_same_pci_cells.dropna(subset=['lat', 'lon'])

        if same_freq_same_pci_cells.empty:
            result = (True, float('inf'))
            self.pci_validity_cache[cache_key] = result
            return result

        # 计算到所有同频同PCI小区的距离
        distances = self.calculate_distance_vectorized(
            target_lat, target_lon,
            same_freq_same_pci_cells['lat'].values,
            same_freq_same_pci_cells['lon'].values
        )

        min_distance = np.min(distances)

        # 检查是否有同站点PCI冲突（最高优先级）
        # 同站点小区不能使用相同的PCI，即使距离满足要求
        same_site_cells = self.get_same_site_cells(target_lat, target_lon, exclude_enodeb, exclude_cell)
        for same_site_cell in same_site_cells:
            same_site_pci = same_site_cell.get('pci')
            if pd.notna(same_site_pci) and same_site_pci == candidate_pci:
                # 发现同站点PCI冲突，直接返回无效
                result = (False, 0.0)  # 距离设为0表示同站冲突
                self.pci_validity_cache[cache_key] = result
                return result

        is_valid = min_distance >= self.reuse_distance_km

        result = (is_valid, min_distance)
        self.pci_validity_cache[cache_key] = result
        return result

    def validate_multiple_pci_reuse_distance(self, candidate_pcis: List[int], target_lat: float, target_lon: float,
                                           target_earfcn: float, exclude_enodeb: int = None,
                                           exclude_cell: int = None) -> List[Tuple[int, bool, float]]:
        """
        批量验证多个PCI的复用距离要求 - 向量化优化版本
        大幅提升性能，避免逐个PCI验证的开销

        Args:
            candidate_pcis: 候选PCI列表
            target_lat: 目标纬度
            target_lon: 目标经度
            target_earfcn: 下行频点
            exclude_enodeb: 要排除的基站ID
            exclude_cell: 要排除的小区ID

        Returns:
            List of (pci, is_valid, min_distance) tuples
        """
        if self.all_cells_combined is None or self.all_cells_combined.empty:
            return [(pci, True, float('inf')) for pci in candidate_pcis]

        results = []

        # 按PCI值分组，避免重复计算相同PCI的距离
        unique_pcis = list(set(candidate_pcis))
        pci_validity_map = {}

        for pci in unique_pcis:
            # 检查缓存
            cache_key = (pci, round(target_lat, 6), round(target_lon, 6),
                        round(target_earfcn, 2), exclude_enodeb, exclude_cell,
                        round(self.reuse_distance_km, 2))

            if cache_key in self.pci_validity_cache:
                pci_validity_map[pci] = self.pci_validity_cache[cache_key]
                continue

            # 只查找同频同PCI的小区
            same_freq_same_pci_cells = self.all_cells_combined[
                (self.all_cells_combined['pci'] == pci) &
                (self.all_cells_combined['pci'].notna()) &
                (self.all_cells_combined['earfcn_dl'] == target_earfcn) &
                (self.all_cells_combined['earfcn_dl'].notna())
            ].copy()

            # 排除当前小区自己
            if exclude_enodeb is not None and exclude_cell is not None:
                same_freq_same_pci_cells = same_freq_same_pci_cells[
                    ~((same_freq_same_pci_cells['enodeb_id'] == exclude_enodeb) &
                      (same_freq_same_pci_cells['cell_id'] == exclude_cell))
                ]

            # 过滤有效位置的小区
            same_freq_same_pci_cells = same_freq_same_pci_cells.dropna(subset=['lat', 'lon'])

            if same_freq_same_pci_cells.empty:
                validity_result = (True, float('inf'))
                pci_validity_map[pci] = validity_result
                self.pci_validity_cache[cache_key] = validity_result
                continue

            # 向量化计算到所有同频同PCI小区的距离
            distances = self.calculate_distance_vectorized(
                target_lat, target_lon,
                same_freq_same_pci_cells['lat'].values,
                same_freq_same_pci_cells['lon'].values
            )

            min_distance = np.min(distances)

            # 检查是否有同站点PCI冲突
            same_site_cells = self.get_same_site_cells(target_lat, target_lon, exclude_enodeb, exclude_cell)
            for same_site_cell in same_site_cells:
                same_site_pci = same_site_cell.get('pci')
                if pd.notna(same_site_pci) and same_site_pci == pci:
                    validity_result = (False, 0.0)
                    pci_validity_map[pci] = validity_result
                    self.pci_validity_cache[cache_key] = validity_result
                    break
            else:
                # 没有发现同站点PCI冲突
                is_valid = min_distance >= self.reuse_distance_km
                validity_result = (is_valid, min_distance)
                pci_validity_map[pci] = validity_result
                self.pci_validity_cache[cache_key] = validity_result

        # 为每个候选PCI返回结果
        return [(pci,) + pci_validity_map[pci] for pci in candidate_pcis]

    def get_same_site_cells(self, target_lat: float, target_lon: float, exclude_enodeb_id: int = None, exclude_cell_id: int = None) -> List[Dict]:
        """
        获取同站点的其他小区信息
        真正基于经纬度判断同站点，不依赖eNodeBID
        """
        if pd.isna(target_lat) or pd.isna(target_lon):
            return []

        # 使用经纬度查找相同位置的小区
        # 容差值设为0.0001度（约10米精度），确保同一物理位置的小区能够正确匹配
        tolerance = 0.0001  # 约10米精度
        location_filter = (
            (abs(self.all_cells_combined['lat'] - target_lat) < tolerance) &
            (abs(self.all_cells_combined['lon'] - target_lon) < tolerance)
        )

        # 从所有小区数据中查找相同位置的小区（包括未分配PCI的小区）
        same_site_cells = self.all_cells_combined[location_filter].copy()

        # 排除指定的基站ID和小区ID
        if exclude_enodeb_id is not None and exclude_cell_id is not None:
            same_site_cells = same_site_cells[
                ~((same_site_cells['enodeb_id'] == exclude_enodeb_id) &
                  (same_site_cells['cell_id'] == exclude_cell_id))
            ]

        # 关键修复：移除cell_type过滤条件，因为cell_type值可能不一致
        # 在测试和实际数据中，目标小区的cell_type可能是"target"，而现有小区可能是"NR"或其他值
        # 为了正确识别同站小区，应该基于位置和基站ID判断，而不是cell_type
        # 保留所有同位置的小区，无论cell_type如何

        # 确保获取到最新的PCI分配信息
        result = same_site_cells.to_dict('records')

        # 调试信息：显示同站点小区的PCI分配情况
        if result and False:  # 关闭调试信息
            print(f"      [DEBUG] 位置({target_lat:.6f}, {target_lon:.6f})同站点小区PCI分配情况:")
            for cell in result:
                enodeb_id = cell.get('enodeb_id', 'N/A')
                cell_id = cell.get('cell_id', 'N/A')
                pci = cell.get('pci', 'N/A')
                mod_val = int(pci) % self.mod_value if pd.notna(pci) and pci != -1 else 'N/A'
                print(f"        基站{enodeb_id}-小区{cell_id}: PCI={pci}, mod{self.mod_value}={mod_val}")

        return result

    def get_cells_at_same_location(self, target_lat: float, target_lon: float,
                                  exclude_enodeb_id: int = None, exclude_cell_id: int = None) -> List[Dict]:
        """
        获取相同经纬度位置的其他小区信息

        Args:
            target_lat: 目标纬度
            target_lon: 目标经度
            exclude_enodeb_id: 要排除的基站ID
            exclude_cell_id: 要排除的小区ID

        Returns:
            相同位置的小区列表
        """
        # 使用经纬度查找相同位置的小区
        # 增大容差值到0.0001度（约10米精度），确保同一基站的小区能够正确匹配
        tolerance = 0.0001  # 约10米精度
        location_filter = (
            (abs(self.all_cells_combined['lat'] - target_lat) < tolerance) &
            (abs(self.all_cells_combined['lon'] - target_lon) < tolerance)
        )

        # 从所有小区数据中查找相同位置的小区（包括未分配PCI的小区）
        # 关键修改：统一使用all_cells_combined作为数据源，确保同站点判断一致性
        same_location_cells = self.all_cells_combined[location_filter]

        # 排除指定的基站ID
        if exclude_enodeb_id is not None:
            same_location_cells = same_location_cells[
                same_location_cells['enodeb_id'] != exclude_enodeb_id
            ]

        # 排除指定的小区ID
        if exclude_cell_id is not None:
            same_location_cells = same_location_cells[
                same_location_cells['cell_id'] != exclude_cell_id
            ]

        # 移除cell_type过滤条件，因为cell_type值可能不一致（如"target" vs "NR"）
        # 保留所有同位置的小区，无论cell_type如何

        return same_location_cells.to_dict('records')

    def check_same_site_mod_conflict(self, candidate_pci: int, target_lat: float, target_lon: float,
                                   exclude_enodeb_id: int = None, exclude_cell_id: int = None) -> bool:
        """
        检查候选PCI是否与同站点小区的模值冲突
        真正基于经纬度判断同站点
        支持NR网络的双模约束（同时检查模3和模30冲突）

        Args:
            candidate_pci: 候选PCI值
            target_lat: 目标小区纬度
            target_lon: 目标小区经度
            exclude_enodeb_id: 要排除的基站ID
            exclude_cell_id: 要排除的小区ID

        Returns:
            True表示无冲突，False表示有冲突
        """
        candidate_mod = candidate_pci % self.mod_value

        # 检查同站点小区模值冲突（基于经纬度判断）
        same_site_cells = self.get_same_site_cells(target_lat, target_lon, exclude_enodeb_id, exclude_cell_id)

        # 统计同站点已有的模值
        existing_mods = []
        conflict_cells = []

        # 检查是否与同站点其他小区的模值冲突
        for cell in same_site_cells:
            existing_pci = cell['pci']
            if pd.notna(existing_pci) and existing_pci != -1:
                existing_mod = int(existing_pci) % self.mod_value
                existing_mods.append(existing_mod)
                if existing_mod == candidate_mod:
                    enodeb_id = cell.get('enodeb_id', 'N/A')
                    cell_id = cell.get('cell_id', 'N/A')
                    conflict_cells.append(f"基站{enodeb_id}-小区{cell_id}(PCI{existing_pci})")

        # 关键修复：无论是否继承模值，都必须避免同站同模冲突
        # 对于LTE网络，同站3个小区必须保证模3值不同
        if self.network_type == "LTE":
            # 检查同站点是否已有3个不同模值
            unique_mods = set(existing_mods)
            if len(unique_mods) >= 3 and candidate_mod in unique_mods:
                # 同站点已经有3个不同模值，且候选PCI的模值已经存在，禁止分配
                print(f"      [DEBUG] 发现LTE同站点模{self.mod_value}冲突: 候选PCI={candidate_pci} (mod{self.mod_value}={candidate_mod})")
                print(f"      [DEBUG] 同站点已有模{self.mod_value}值: {sorted(existing_mods)} (已达3种不同值)")
                return False

        # 对于NR网络，需要同时检查模3和模30冲突
        if self.network_type == "NR" and self.dual_mod_requirement:
            # 检查模3冲突（NR网络的最高优先级）
            candidate_mod3 = candidate_pci % 3
            existing_mod3s = []
            mod3_conflict_cells = []

            for cell in same_site_cells:
                existing_pci = cell['pci']
                if pd.notna(existing_pci) and existing_pci != -1:
                    existing_mod3 = int(existing_pci) % 3
                    existing_mod3s.append(existing_mod3)
                    # NR网络绝对不允许同站同模3，即使没有达到3个小区也要避免
                    if existing_mod3 == candidate_mod3:
                        enodeb_id = cell.get('enodeb_id', 'N/A')
                        cell_id = cell.get('cell_id', 'N/A')
                        mod3_conflict_cells.append(f"基站{enodeb_id}-小区{cell_id}(PCI{existing_pci})")
                        print(f"      [严重警告] NR同站模3冲突: 候选PCI={candidate_pci} (mod3={candidate_mod3})")
                        print(f"      [严重警告] 冲突小区: 基站{enodeb_id}-小区{cell_id} (PCI={existing_pci}, mod3={existing_mod3})")
                        return False  # 直接返回，不允许模3冲突

            # 如果发现任何模3冲突，直接返回False
            if mod3_conflict_cells:
                print(f"      [严重警告] NR同站点已有模3值: {sorted(existing_mod3s)}，候选PCI模3={candidate_mod3}冲突")
                return False

        # 如果发现冲突，记录详细信息
        if conflict_cells:
            print(f"      [DEBUG] 发现同站点模{self.mod_value}冲突: 候选PCI={candidate_pci} (mod{self.mod_value}={candidate_mod})")
            print(f"      [DEBUG] 冲突小区: {', '.join(conflict_cells)}")
            print(f"      [DEBUG] 同站点已有模{self.mod_value}值: {sorted(existing_mods)}")
            return False  # 发现冲突

        return True  # 无冲突

    def get_reuse_compliant_pcis(self, target_lat: float, target_lon: float, target_earfcn: float,
                                exclude_enodeb: int = None, exclude_cell: int = None,
                                target_mod: Optional[int] = None) -> List[Tuple[int, float]]:
        """
        获取满足复用距离要求的PCI列表，按照多重优先级排序
        增强同站点模值冲突避免逻辑

        Returns:
            List of (pci, min_distance) tuples, sorted by preference
        """
        compliant_pcis = []

        # 如果需要继承mod值，严格按照mod值筛选PCI范围
        if self.inherit_mod and target_mod is not None:
            # 严格继承模式：只考虑匹配mod值的PCI
            candidate_pcis = [pci for pci in self.pci_range if pci % self.mod_value == target_mod]
            print(f"    严格mod{self.mod_value}继承模式: 只检查mod{self.mod_value}={target_mod}的PCI，候选PCI数量: {len(candidate_pcis)}")
        else:
            # 自由规划模式：考虑所有PCI，不考虑mod值
            candidate_pcis = self.pci_range
            print(f"    自由规划模式: 检查所有PCI，不考虑mod{self.mod_value}值，候选PCI数量: {len(candidate_pcis)}")

        # 获取同站点小区的模值信息，用于避免冲突
        same_site_mods = set()
        # 关键修复：直接基于经纬度获取同站点小区，不依赖eNodeBID
        same_site_cells = self.get_same_site_cells(target_lat, target_lon, exclude_enodeb, exclude_cell)
        for cell in same_site_cells:
            # 增强：考虑同站点的所有小区，包括不同基站但同位置的小区
            # 优先级：已分配PCI > 原PCI > 忽略无PCI的小区
            pci_value = cell.get('pci')
            original_pci = cell.get('原PCI')

            # 首先检查已分配的PCI
            if pd.notna(pci_value) and pci_value != -1:
                same_site_mods.add(int(pci_value) % self.mod_value)
            # 如果没有分配PCI，检查原PCI
            elif pd.notna(original_pci) and original_pci != -1:
                same_site_mods.add(int(original_pci) % self.mod_value)
            # 如果都没有PCI值，则跳过该小区

        print(f"    同站点已有模{self.mod_value}值: {sorted(list(same_site_mods))}")

        # 关键修复：对于LTE网络，即使自由规划模式也要考虑同站点模3冲突避免
        if self.network_type == "LTE" and not self.inherit_mod:
            print(f"    LTE自由规划模式：将严格避免同站点模3冲突")
            # 检查同站点是否已经有3个不同模值
            if len(same_site_mods) >= 3:
                print(f"    警告：同站点已有3种不同模{self.mod_value}值，将强制选择不同模值")

        # 对于NR网络，需要同时考虑模3和模30约束
        if self.network_type == "NR" and self.dual_mod_requirement:
            print(f"    NR双模约束模式：将同时避免同站点模3和模30冲突")
            # 统计同站点已有的模3值
            same_site_mod3s = set()
            for cell in same_site_cells:
                pci_value = cell.get('pci')
                original_pci = cell.get('原PCI')

                # 首先检查已分配的PCI
                if pd.notna(pci_value) and pci_value != -1:
                    same_site_mod3s.add(int(pci_value) % 3)
                # 如果没有分配PCI，检查原PCI
                elif pd.notna(original_pci) and original_pci != -1:
                    same_site_mod3s.add(int(original_pci) % 3)

            print(f"    同站点已有模3值: {sorted(list(same_site_mod3s))}")

            # 检查同站点是否已经有3个不同模3值
            if len(same_site_mod3s) >= 3:
                print(f"    警告：同站点已有3种不同模3值，将强制选择不同模3值")

        # 向量化优化：批量验证所有候选PCI的复用距离，避免循环调用
        print(f"    向量化优化：批量验证{len(candidate_pcis)}个候选PCI的复用距离...")
        batch_validation_results = self.validate_multiple_pci_reuse_distance(
            candidate_pcis, target_lat, target_lon, target_earfcn, exclude_enodeb, exclude_cell
        )

        # 检查候选PCI
        for pci, is_valid, min_distance in batch_validation_results:
            # 1. 首先检查复用距离（最高优先级）
            if not is_valid:
                continue  # 不满足复用距离要求的PCI直接跳过

            # 2. 检查同站点模值冲突（第二优先级）
            candidate_mod = pci % self.mod_value
            has_same_site_mod_conflict = candidate_mod in same_site_mods

            # 对于NR网络，额外检查模3冲突（最高优先级）
            has_nr_mod3_conflict = False
            if self.network_type == "NR" and self.dual_mod_requirement:
                # 调用专门的模3冲突检查函数
                no_conflict = self.check_same_site_mod_conflict(
                    pci, target_lat, target_lon, exclude_enodeb, exclude_cell
                )
                has_nr_mod3_conflict = not no_conflict

                # 调试信息
                if has_nr_mod3_conflict:
                    # print(f"      [关键调试] NR候选PCI={pci} (mod3={pci%3}) 检测到同站模3冲突，执行continue跳过")
                    continue  # NR网络模3冲突直接跳过，不进入候选列表
                else:
                    # print(f"      [关键调试] NR候选PCI={pci} (mod3={pci%3}) 无同站模3冲突，可以接受")
                    pass

            # 3. 计算PCI分布均衡性指标（第三优先级）
            # 距离越接近阈值越均衡
            if min_distance == float('inf'):
                balance_score = 0  # 无复用PCI，最均衡
            else:
                # 距离接近阈值得分更高（分数越小越好）
                balance_score = abs(min_distance - self.reuse_distance_km)

            # 关键修复：对于NR网络，需要考虑模3冲突状态
            # 统一冲突标志：mod30冲突或NR模3冲突任一为真都表示有冲突
            has_any_conflict = has_same_site_mod_conflict
            if self.network_type == "NR" and self.dual_mod_requirement:
                has_any_conflict = has_any_conflict or has_nr_mod3_conflict

            compliant_pcis.append((pci, min_distance, has_any_conflict, balance_score))

        if not compliant_pcis:
            return []

        # 关键修复：对于LTE网络，无论是否继承模值，都必须严格避免同站同模
        if self.network_type == "LTE":
            # LTE网络强制同站点模3错开原则
            print(f"    LTE网络：强制执行同站点模3错开原则")

            # 按模3值分组
            mod_groups = {}
            for pci_info in compliant_pcis:
                pci, distance, has_mod_conflict, balance_score = pci_info
                mod_value = pci % self.mod_value
                if mod_value not in mod_groups:
                    mod_groups[mod_value] = []
                mod_groups[mod_value].append(pci_info)

            # 获取同站点已分配的模3值
            same_site_assigned_mods = set()
            same_site_cells_for_check = self.get_same_site_cells(target_lat, target_lon, exclude_enodeb, exclude_cell)
            for cell in same_site_cells_for_check:
                if 'pci' in cell and pd.notna(cell['pci']) and cell['pci'] != -1:
                    same_site_assigned_mods.add(int(cell['pci']) % self.mod_value)

            print(f"    同站点已有模3值: {sorted(list(same_site_assigned_mods))}")
            print(f"    候选PCI模3值分布: {sorted(list(mod_groups.keys()))}")

            # 关键修复：LTE网络必须保证同站3个小区模3值不同
            # 如果同站点已有2个不同模3值，必须选择第3个不同的模3值
            if len(same_site_assigned_mods) >= 2:
                print(f"    同站点已有{len(same_site_assigned_mods)}个不同模3值，必须选择不同的模3值")
                # 强制选择未被使用的模3值
                available_mod_groups = {mod: pci_list for mod, pci_list in mod_groups.items()
                                      if mod not in same_site_assigned_mods}

                if available_mod_groups:
                    # 有可用模3值，优先选择
                    final_pcis = []
                    for mod_value, pci_list in available_mod_groups.items():
                        # 对每个模3组内的PCI进行排序（复用距离优先）
                        pci_list.sort(key=lambda x: (-x[1], x[3], x[0]))
                        final_pcis.extend(pci_list)
                    print(f"    成功找到{len(final_pcis)}个无同站点模3冲突的PCI")
                else:
                    # 所有模3值都被使用，这是严重问题 - 需要强制选择不同的模3值
                    print(f"    严重警告：所有模3值都已被同站点小区使用，违反LTE同站3小区模3错开原则！")

                    # 统计每个mod值的使用次数
                    mod_usage_count = {}
                    for mod in same_site_assigned_mods:
                        mod_usage_count[mod] = 0

                    # 计算每个mod值在同站点的使用次数
                    for cell in same_site_cells_for_check:
                        if 'pci' in cell and pd.notna(cell['pci']) and cell['pci'] != -1:
                            cell_mod = int(cell['pci']) % self.mod_value
                            if cell_mod in mod_usage_count:
                                mod_usage_count[cell_mod] += 1

                    # 找到使用次数最少的mod值
                    if mod_usage_count:
                        min_usage_mod = min(mod_usage_count.keys(), key=lambda x: mod_usage_count[x])
                        print(f"    使用次数最少的mod值: {min_usage_mod} (使用次数: {mod_usage_count[min_usage_mod]})")

                        # 优先选择使用次数最少的mod值的PCI
                        if min_usage_mod in mod_groups:
                            best_mod_pci_list = mod_groups[min_usage_mod]
                            best_mod_pci_list.sort(key=lambda x: (-x[1], x[3], x[0]))
                            final_pcis = best_mod_pci_list
                        else:
                            # 如果没有找到，从所有mod组中选择
                            final_pcis = []
                            for mod_value, pci_list in mod_groups.items():
                                pci_list.sort(key=lambda x: (-x[1], x[3], x[0]))
                                final_pcis.extend(pci_list)
                    else:
                        # 如果没有使用统计，从所有mod组中选择
                        final_pcis = []
                        for mod_value, pci_list in mod_groups.items():
                            pci_list.sort(key=lambda x: (-x[1], x[3], x[0]))
                            final_pcis.extend(pci_list)
            else:
                # 同站点模3值不足2个，可以自由选择但优先避免冲突
                available_mod_groups = {mod: pci_list for mod, pci_list in mod_groups.items()
                                      if mod not in same_site_assigned_mods}

                if available_mod_groups:
                    # 优先选择未被使用的模3值
                    final_pcis = []
                    for mod_value, pci_list in available_mod_groups.items():
                        pci_list.sort(key=lambda x: (-x[1], x[3], x[0]))
                        final_pcis.extend(pci_list)
                    print(f"    成功找到{len(final_pcis)}个无同站点模3冲突的PCI")
                else:
                    # 如果没有可用模3值，使用所有候选PCI
                    final_pcis = []
                    for mod_value, pci_list in mod_groups.items():
                        pci_list.sort(key=lambda x: (-x[1], x[3], x[0]))
                        final_pcis.extend(pci_list)
                    print(f"    使用所有候选PCI，共{len(final_pcis)}个")
        else:
            # NR网络或非LTE网络的原有逻辑
            # 新优先级策略：复用距离 > 同站点模值冲突避免 > PCI分布均衡性
            # 首先尝试无同站点模值冲突的PCI（最高优先级）
            no_mod_conflict_pcis = [(pci, distance, has_mod_conflict, balance_score)
                                   for pci, distance, has_mod_conflict, balance_score in compliant_pcis
                                   if not has_mod_conflict]

            # 对于NR网络的双模约束，需要额外检查模3冲突
            if self.network_type == "NR" and self.dual_mod_requirement:
                print(f"    NR双模约束：将同时检查模30和模3冲突")

                # 统计同站点已有的模3值
                same_site_mod3s = set()
                same_site_cells_for_check = self.get_same_site_cells(target_lat, target_lon, exclude_enodeb, exclude_cell)
                for cell in same_site_cells_for_check:
                    if 'pci' in cell and pd.notna(cell['pci']) and cell['pci'] != -1:
                        same_site_mod3s.add(int(cell['pci']) % 3)

                print(f"    同站点已有模3值: {sorted(list(same_site_mod3s))}")

                # 过滤掉有模3冲突的PCI
                no_mod3_conflict_pcis = []
                for pci_info in no_mod_conflict_pcis:
                    pci = pci_info[0]
                    mod3_value = pci % 3
                    if mod3_value not in same_site_mod3s:
                        no_mod3_conflict_pcis.append(pci_info)

                if no_mod3_conflict_pcis:
                    print(f"    成功找到{len(no_mod3_conflict_pcis)}个无模30和模3冲突的PCI")
                    final_pcis = no_mod3_conflict_pcis
                else:
                    print(f"    警告：所有无模30冲突的PCI都存在模3冲突")
                    # 如果没有无模3冲突的PCI，使用所有无模30冲突的PCI
                    final_pcis = no_mod_conflict_pcis
                    print(f"    使用{len(final_pcis)}个无模30冲突但可能存在模3冲突的PCI")
            else:
                 # 非NR双模约束或非NR网络，使用原有逻辑
                 if no_mod_conflict_pcis:
                     # 优先使用无模值冲突的PCI
                     final_pcis = no_mod_conflict_pcis
                     print(f"    成功找到{len(final_pcis)}个无同站点模{self.mod_value}冲突的PCI")
                 else:
                     # 关键修复：即使所有PCI都有模3冲突，也要优先选择模3值不同的PCI
                     # 避免同站所有小区模3值相同的情况
                     print(f"    警告：所有满足复用距离的PCI都与同站点存在模{self.mod_value}冲突")
                     print(f"    同站点已有模{self.mod_value}值: {sorted(list(same_site_mods))}")

                 # 按模3值分组，优先选择模3值不同的PCI
                 mod_groups = {}
                 for pci_info in compliant_pcis:
                     pci, distance, has_mod_conflict, balance_score = pci_info
                     mod_value = pci % self.mod_value
                     if mod_value not in mod_groups:
                         mod_groups[mod_value] = []
                     mod_groups[mod_value].append(pci_info)

                 # 关键修复：绝对不允许同站同模，必须选择不同的模3值
                 if len(mod_groups) > 1:
                     print(f"    发现{len(mod_groups)}个不同的模{self.mod_value}值组，将强制选择不同的模3值")

                     # 获取当前位置已分配PCI的模3值，避免重复选择
                     same_site_assigned_mods = set()
                     # 基于经纬度获取同站点已分配的小区PCI模3值
                     same_site_cells_for_check = self.get_same_site_cells(target_lat, target_lon, exclude_enodeb, exclude_cell)
                     for cell in same_site_cells_for_check:
                         if 'pci' in cell and pd.notna(cell['pci']):
                             same_site_assigned_mods.add(int(cell['pci']) % self.mod_value)

                     # 强制选择未被同基站其他小区使用的模3值
                     available_mod_groups = {mod: pci_list for mod, pci_list in mod_groups.items()
                                           if mod not in same_site_assigned_mods}

                     if available_mod_groups:
                         # 优先选择未被使用的模3值
                         final_pcis = []
                         for mod_value, pci_list in available_mod_groups.items():
                             # 对每个模3组内的PCI进行排序（复用距离优先）
                             pci_list.sort(key=lambda x: (-x[1], x[3], x[0]))
                             final_pcis.append(pci_list[0])
                     else:
                         # 如果所有模3值都被使用，必须选择不同的模3值来避免同站同模
                         # 选择与已分配模3值不同的模3值
                         final_pcis = []
                         for mod_value, pci_list in remaining_mod_groups.items():
                             if mod_value not in same_site_assigned_mods:
                                 pci_list.sort(key=lambda x: (-x[1], x[3], x[0]))
                                 final_pcis.append(pci_list[0])
                         if not final_pcis:
                             # 如果无法避免，选择最优的PCI，但记录严重警告
                             print(f"    严重警告：无法避免同站同模！将选择最优PCI")
                             final_pcis = []
                             for mod_value, pci_list in mod_groups.items():
                                 pci_list.sort(key=lambda x: (-x[1], x[3], x[0]))
                                 final_pcis.append(pci_list[0])
                 else:
                     # 如果所有PCI模3值都相同，无法避免同站同模
                     print(f"    严重警告：所有候选PCI模{self.mod_value}值相同，无法避免同站同模！")
                     # 强制选择不同的PCI，即使模3值相同也要避免同站同模
                     final_pcis = []
                     for mod_value, pci_list in mod_groups.items():
                         pci_list.sort(key=lambda x: (-x[1], x[3], x[0]))
                         final_pcis.append(pci_list[0])

        # 多重优先级排序策略 - 修正优先级顺序，强化NR网络同站模3错开
        def sort_key(pci_info):
            pci, distance, has_mod_conflict, balance_score = pci_info

            # 对于NR网络，需要检查同站模3冲突（最高优先级）
            mod3_conflict_priority = 0
            if self.network_type == "NR" and self.dual_mod_requirement:
                # 获取同站点已有的模3值
                same_site_mod3s = set()
                same_site_cells_for_check = self.get_same_site_cells(target_lat, target_lon, exclude_enodeb, exclude_cell)
                for cell in same_site_cells_for_check:
                    if 'pci' in cell and pd.notna(cell['pci']) and cell['pci'] != -1:
                        same_site_mod3s.add(int(cell['pci']) % 3)

                candidate_mod3 = pci % 3
                # 同站模3冲突具有最高优先级，必须避免
                mod3_conflict_priority = 0 if candidate_mod3 not in same_site_mod3s else 1

                # 调试信息
                if mod3_conflict_priority == 1:
                    print(f"      [NR模3冲突] PCI={pci} (mod3={candidate_mod3}) 与同站模3值冲突: {sorted(list(same_site_mod3s))}")

            # 优先级1（LTE）/ 优先级2（NR）：同站点模值冲突（mod30）- 避免同站同mod30
            mod30_conflict_priority = 0 if not has_mod_conflict else 1

            # 优先级2（LTE）/ 优先级3（NR）：复用距离合规性 - 必须满足用户设定的最小复用距离
            if distance >= self.reuse_distance_km:
                distance_compliance = 0  # 满足复用距离要求
            else:
                distance_compliance = 1  # 不满足复用距离要求

            # 优先级3（LTE）/ 优先级4（NR）：PCI分布均衡性 - 越接近阈值越好，避免选择过远距离的PCI
            if distance >= self.reuse_distance_km:
                # 对于满足复用距离要求的PCI，选择距离接近阈值的
                balance_priority = abs(distance - self.reuse_distance_km)
            else:
                # 对于不满足复用距离要求的PCI，给予惩罚
                balance_priority = 999999 + abs(distance - self.reuse_distance_km)

            # 优先级4（LTE）/ 优先级5（NR）：复用距离大小 - 在满足条件下适当考虑距离，但不作为主要因素
            if distance >= self.reuse_distance_km:
                # 在满足复用距离的条件下，略微偏向距离较大的，但权重很低
                distance_priority = -distance * 0.1  # 降低距离权重
            else:
                # 不满足复用距离的PCI，距离大的优先
                distance_priority = -distance

            # 优先级5（LTE）/ 优先级6（NR）：连续PCI分配优化 - 优先选择能与同站点已分配PCI形成连续序列的PCI
            # 获取同站点已分配的PCI值
            same_site_assigned_pcis = self.get_same_site_assigned_pcis(target_lat, target_lon, exclude_enodeb, exclude_cell)
            continuity_priority = self.calculate_pci_continuity_score(pci, same_site_assigned_pcis)

            # 优先级6（LTE）/ 优先级7（NR）：PCI值 - 小的优先，确保确定性
            pci_priority = pci

            # 根据网络类型返回不同的优先级元组
            if self.network_type == "NR" and self.dual_mod_requirement:
                # NR网络：模3冲突 > mod30冲突 > 复用距离 > 均衡性 > 距离 > 连续性 > PCI值
                return (mod3_conflict_priority, mod30_conflict_priority, distance_compliance,
                       balance_priority, distance_priority, continuity_priority, pci_priority)
            else:
                # LTE网络：mod30冲突 > 复用距离 > 均衡性 > 距离 > 连续性 > PCI值
                return (mod30_conflict_priority, distance_compliance, balance_priority,
                       distance_priority, continuity_priority, pci_priority)

        final_pcis.sort(key=sort_key)

        # 输出排序结果用于调试
        if len(final_pcis) > 0:
            top_candidates = final_pcis[:min(5, len(final_pcis))]

            # 根据网络类型显示不同的优先级信息
            if self.network_type == "NR" and self.dual_mod_requirement:
                print(f"    前{len(top_candidates)}个候选PCI排序结果 (NR优先级：模3冲突>mod30冲突>复用距离>均衡性>距离>连续性>PCI值):")
            else:
                print(f"    前{len(top_candidates)}个候选PCI排序结果 (LTE优先级：mod30冲突>复用距离>均衡性>距离>连续性>PCI值):")

            # 获取同站点已分配PCI用于连续性分析
            same_site_assigned_pcis = self.get_same_site_assigned_pcis(target_lat, target_lon, exclude_enodeb, exclude_cell)

            # 对于NR网络，显示模3信息
            if self.network_type == "NR" and self.dual_mod_requirement:
                same_site_mod3s = set()
                for assigned_pci in same_site_assigned_pcis:
                    same_site_mod3s.add(assigned_pci % 3)
                print(f"    同站点已分配PCI: {same_site_assigned_pcis}, 模3值: {sorted(list(same_site_mod3s))}")
            else:
                print(f"    同站点已分配PCI: {same_site_assigned_pcis}")

            for i, (pci, dist, conflict, _) in enumerate(top_candidates):
                status = "无冲突" if not conflict else "有冲突"
                if dist == float('inf'):
                    dist_str = "无复用PCI"
                else:
                    dist_str = f"{dist:.2f}km"
                    if dist >= self.reuse_distance_km:
                        dist_str += f"(满足≥{self.reuse_distance_km}km)"
                    else:
                        dist_str += f"(不满足<{self.reuse_distance_km}km)"

                # 计算连续性得分和描述
                continuity_score = self.calculate_pci_continuity_score(pci, same_site_assigned_pcis)
                if continuity_score == 0:
                    continuity_str = "连续"
                elif continuity_score == 1:
                    continuity_str = "接近连续"
                else:
                    continuity_str = "不连续"

                # 对于NR网络，添加模3冲突检查
                if self.network_type == "NR" and self.dual_mod_requirement:
                    mod3_conflict = pci % 3 in same_site_mod3s if same_site_mod3s else False
                    mod3_status = "模3冲突" if mod3_conflict else "模3正常"
                    print(f"      {i+1}. PCI={pci:3d} (mod3={pci%3}, mod30={pci%self.mod_value}), 距离={dist_str:>12}, {status:>6}, {mod3_status:>8}, 连续性={continuity_str}")
                else:
                    print(f"      {i+1}. PCI={pci:3d} (mod{self.mod_value}={pci%self.mod_value}), 距离={dist_str:>12}, {status:>6}, 连续性={continuity_str}")

        # 保持4元组格式以确保数据结构一致性
        return final_pcis

    def get_same_site_assigned_pcis(self, target_lat: float, target_lon: float,
                                   exclude_enodeb: int = None, exclude_cell: int = None) -> List[int]:
        """
        获取同站点已分配的PCI值列表

        Args:
            target_lat: 目标小区纬度
            target_lon: 目标小区经度
            exclude_enodeb: 要排除的基站ID
            exclude_cell: 要排除的小区ID

        Returns:
            List[int]: 同站点已分配的PCI值列表
        """
        same_site_cells = self.get_same_site_cells(target_lat, target_lon, exclude_enodeb, exclude_cell)
        assigned_pcis = []

        for cell in same_site_cells:
            pci_value = cell.get('pci')
            if pd.notna(pci_value) and pci_value != -1:
                assigned_pcis.append(int(pci_value))

        return sorted(assigned_pcis)

    def calculate_pci_continuity_score(self, candidate_pci: int, same_site_assigned_pcis: List[int]) -> int:
        """
        计算候选PCI的连续性得分

        Args:
            candidate_pci: 候选PCI值
            same_site_assigned_pcis: 同站点已分配的PCI值列表

        Returns:
            int: 连续性得分（越小越好，0表示完全不连续，1表示能形成连续序列）
        """
        if not same_site_assigned_pcis:
            # 同站点没有已分配的PCI，无法形成连续序列
            return 1

        # 检查是否能与已分配PCI形成连续序列
        # 连续序列是指相邻的PCI值（如 1,2,3 或 10,11,12）
        assigned_set = set(same_site_assigned_pcis)

        # 检查候选PCI是否能扩展现有连续序列
        # 例如：已分配 [10,11]，候选PCI=9或12可以形成连续序列
        for assigned_pci in same_site_assigned_pcis:
            if candidate_pci == assigned_pci - 1:  # 候选PCI在序列前面
                return 0  # 优先级最高
            elif candidate_pci == assigned_pci + 1:  # 候选PCI在序列后面
                return 0  # 优先级最高

        # 检查是否能与任意一个已分配PCI形成连续对（即使不是完整序列）
        for assigned_pci in same_site_assigned_pcis:
            if abs(candidate_pci - assigned_pci) == 1:
                return 0  # 优先级最高

        # 检查是否能形成接近连续的序列（相差2）
        for assigned_pci in same_site_assigned_pcis:
            if abs(candidate_pci - assigned_pci) == 2:
                return 1  # 优先级较高

        # 无法形成任何连续性
        return 2

    def calculate_fallback_pci(self, enodeb_id: int, cell_id: int, lat: float, lon: float) -> int:
        """
        计算保底PCI，避免同站同PCI冲突和模3冲突

        Args:
            enodeb_id: 基站ID
            cell_id: 小区ID
            lat: 纬度
            lon: 经度

        Returns:
            计算出的PCI值
        """
        # 获取同站点已分配的PCI
        same_site_assigned_pcis = self.get_same_site_assigned_pcis(lat, lon, enodeb_id, cell_id)
        used_pcis = set(same_site_assigned_pcis)

        # 基础PCI：使用基站ID和小区ID的组合
        base_pci = (enodeb_id * 100 + cell_id) % len(self.pci_range)

        # 对于NR网络，需要确保模3错开
        if self.network_type == "NR" and self.dual_mod_requirement:
            # 获取同站小区的模3值
            used_mod3s = set()
            for pci in used_pcis:
                if pd.notna(pci) and pci != -1:
                    used_mod3s.add(int(pci) % 3)

            # 尝试找到一个未被占用的模3值
            for mod3 in range(3):
                if mod3 not in used_mod3s:
                    # 计算符合该模3值的最小PCI
                    for offset in range(0, 100):  # 最多尝试100个偏移
                        candidate_pci = (base_pci + offset) % len(self.pci_range)
                        if candidate_pci % 3 == mod3 and candidate_pci not in used_pcis:
                            return candidate_pci

            # 如果所有模3值都被占用，选择任意未使用的PCI
            for offset in range(0, 100):
                candidate_pci = (base_pci + offset) % len(self.pci_range)
                if candidate_pci not in used_pcis:
                    return candidate_pci
        else:
            # LTE网络：简单避免同站同PCI
            for offset in range(0, 100):
                candidate_pci = (base_pci + offset) % len(self.pci_range)
                if candidate_pci not in used_pcis:
                    return candidate_pci

        # 如果仍然找不到，返回基础PCI（极少数情况）
        return base_pci

    def validate_fallback_pci_meets_constraints(self, pci: int, lat: float, lon: float,
                                              earfcn_dl: float, enodeb_id: int, cell_id: int) -> bool:
        """
        验证降级或保底PCI是否满足原始约束条件

        关键修复：验证的是待分配PCI是否满足原始复用距离要求，而不是检查冲突

        Args:
            pci: 要验证的PCI值
            lat: 纬度
            lon: 经度
            earfcn_dl: 下行频点
            enodeb_id: 基站ID
            cell_id: 小区ID

        Returns:
            bool: 是否满足约束条件
        """
        # 1. 检查复用距离约束
        # 暂时保存原始复用距离
        original_distance = self.reuse_distance_km

        try:
            print(f"    [验证开始] PCI={pci}, 原始复用距离={original_distance}km")

            # 向量化优化：批量检查同频复用距离冲突
            # 筛选出同频同PCI的小区（排除自身）
            same_freq_same_pci_cells = self.all_cells_combined[
                (self.all_cells_combined['pci'] == pci) &
                (self.all_cells_combined['pci'].notna()) &
                (self.all_cells_combined['pci'] != 0) &
                (self.all_cells_combined['earfcn_dl'] == earfcn_dl) &
                (self.all_cells_combined['earfcn_dl'].notna()) &
                ~((self.all_cells_combined['enodeb_id'] == enodeb_id) &
                  (self.all_cells_combined['cell_id'] == cell_id))
            ].copy()

            if not same_freq_same_pci_cells.empty:
                # 向量化计算距离
                distances = self.calculate_distance_vectorized(
                    lat, lon,
                    same_freq_same_pci_cells['lat'].values,
                    same_freq_same_pci_cells['lon'].values
                )

                # 找到最小距离
                min_distance = np.min(distances)
                print(f"    [检查] 同频PCI={pci}, 最小距离={min_distance:.2f}km")

                # 检查是否违反复用距离
                if 0 < min_distance < original_distance:
                    print(f"    [验证失败] 复用距离冲突: PCI={pci}, 距离={min_distance:.2f}km < {original_distance}km")
                    return False

            print(f"    [检查完成] 复用距离检查通过，无冲突")

            # 3. 检查NR网络模3和模30约束
            if self.network_type == "NR" and self.dual_mod_requirement:
                # 检查模3冲突（关键修复：临时分配PCI进行检查，然后撤销）
                temp_assigned = self.get_temp_assignment_for_validation(enodeb_id, cell_id, pci)

                try:
                    # 临时更新数据以进行验证
                    temp_cell_mask = (self.all_cells_combined['enodeb_id'] == enodeb_id) & \
                                    (self.all_cells_combined['cell_id'] == cell_id)
                    original_pci = self.all_cells_combined.loc[temp_cell_mask, 'pci'].values[0] if temp_cell_mask.any() else 0
                    self.all_cells_combined.loc[temp_cell_mask, 'pci'] = pci

                    # 检查模3冲突
                    if not self.check_same_site_mod_conflict(pci, lat, lon, enodeb_id, cell_id):
                        print(f"    [验证失败] NR同站模3冲突: PCI={pci} (mod3={pci%3})")
                        return False
                    print(f"    [检查通过] NR模3约束")

                    # 向量化优化：检查模30冲突（只检查同站其他小区）
                    # 筛选同站的其他已分配小区
                    same_site_other_cells = self.all_cells_combined[
                        (self.all_cells_combined['enodeb_id'] == enodeb_id) &
                        (self.all_cells_combined['cell_id'] != cell_id) &
                        (self.all_cells_combined['earfcn_dl'] == earfcn_dl) &
                        (self.all_cells_combined['pci'].notna()) &
                        (self.all_cells_combined['pci'] != 0)
                    ].copy()

                    if not same_site_other_cells.empty:
                        # 向量化计算距离
                        distances = self.calculate_distance_vectorized(
                            lat, lon,
                            same_site_other_cells['lat'].values,
                            same_site_other_cells['lon'].values
                        )

                        # 筛选出同站小区（距离<0.01km）
                        same_site_mask = distances < 0.01
                        same_site_cells = same_site_other_cells[same_site_mask]

                        # 检查模30冲突
                        for _, existing_cell in same_site_cells.iterrows():
                            if pci % 30 == existing_cell['pci'] % 30:
                                print(f"    [验证失败] NR同站模30冲突: PCI={pci} (mod30={pci%30}) vs 现有PCI={existing_cell['pci']} (mod30={existing_cell['pci']%30})")
                                return False

                    print(f"    [检查通过] NR模30约束")

                finally:
                    # 恢复原始PCI值
                    self.all_cells_combined.loc[temp_cell_mask, 'pci'] = original_pci

            # 4. 检查LTE网络模3约束
            elif self.network_type == "LTE":
                # LTE模3约束检查（同样使用临时分配机制）
                temp_cell_mask = (self.all_cells_combined['enodeb_id'] == enodeb_id) & \
                                (self.all_cells_combined['cell_id'] == cell_id)
                original_pci = self.all_cells_combined.loc[temp_cell_mask, 'pci'].values[0] if temp_cell_mask.any() else 0
                self.all_cells_combined.loc[temp_cell_mask, 'pci'] = pci

                try:
                    # 向量化优化：LTE模3约束检查
                    # 筛选同站的其他已分配小区
                    same_site_other_cells = self.all_cells_combined[
                        (self.all_cells_combined['enodeb_id'] == enodeb_id) &
                        (self.all_cells_combined['cell_id'] != cell_id) &
                        (self.all_cells_combined['earfcn_dl'] == earfcn_dl) &
                        (self.all_cells_combined['pci'].notna()) &
                        (self.all_cells_combined['pci'] != 0)
                    ].copy()

                    if not same_site_other_cells.empty:
                        # 向量化计算距离
                        distances = self.calculate_distance_vectorized(
                            lat, lon,
                            same_site_other_cells['lat'].values,
                            same_site_other_cells['lon'].values
                        )

                        # 筛选出同站小区（距离<0.01km）
                        same_site_mask = distances < 0.01
                        same_site_cells = same_site_other_cells[same_site_mask]

                        # 检查模3冲突
                        for _, existing_cell in same_site_cells.iterrows():
                            if pci % 3 == existing_cell['pci'] % 3:
                                print(f"    [验证失败] LTE同站模3冲突: PCI={pci} (mod3={pci%3}) vs 现有PCI={existing_cell['pci']} (mod3={existing_cell['pci']%3})")
                                return False

                    print(f"    [检查通过] LTE模3约束")

                finally:
                    # 恢复原始PCI值
                    self.all_cells_combined.loc[temp_cell_mask, 'pci'] = original_pci

            # 如果所有检查都通过，则满足约束
            print(f"    [验证成功] PCI={pci} 满足所有原始约束条件")
            return True

        except Exception as e:
            print(f"    [验证警告] 保底PCI验证时发生异常: {e}")
            import traceback
            traceback.print_exc()
            # 验证失败时保守处理，认为不满足约束
            return False

    def get_temp_assignment_for_validation(self, enodeb_id: int, cell_id: int, pci: int) -> bool:
        """获取临时分配信息（保留接口，可能后续扩展使用）"""
        return True

    def assign_pci_with_reuse_priority(self, enodeb_id: int, cell_id: int) -> Tuple[Optional[int], str, float, float]:
        """
        优先确保复用距离的PCI分配方法

        Returns:
            (assigned_pci, reason, earfcn_dl, actual_min_reuse_distance)
            assigned_pci: 分配的PCI值，如果无法分配则为None
        """
        # 获取小区信息
        cell_info, status = self.get_cell_info(enodeb_id, cell_id)

        if status != 'success':
            # 对于找不到的小区或缺少位置信息的小区，不进行PCI分配
            if status == 'cell_not_found':
                # 小区未找到，返回None表示不分配PCI
                return None, 'cell_not_found_no_pci', 0.0, 0.0
            elif status == 'no_location':
                # 缺少位置信息，不进行PCI分配，按优化要求返回特定格式
                earfcn_dl = cell_info['earfcn_dl'] if cell_info is not None else 0.0
                return None, 'no_location_fallback', earfcn_dl, 0.0

        target_lat = cell_info['lat']
        target_lon = cell_info['lon']
        earfcn_dl = cell_info['earfcn_dl']

        # 确定目标mod值（如果需要继承）
        target_mod = None
        if self.inherit_mod and not pd.isna(cell_info['pci']):
            target_mod = int(cell_info['pci']) % self.mod_value

        # 获取满足复用距离的PCI列表（同频检查）- 现在直接返回4元组格式
        compliant_pcis = self.get_reuse_compliant_pcis(
            target_lat, target_lon, earfcn_dl, enodeb_id, cell_id, target_mod
        )
        # 关键修复：对于LTE网络，不再进行额外的模3均衡化处理，避免破坏同站点模3检查逻辑
        # 因为get_reuse_compliant_pcis方法已经包含了完善的LTE同站点模3错开逻辑
        if self.network_type != "LTE" and len(compliant_pcis) > 0:
            # 仅对非LTE网络进行模3均衡化处理
            # 按模3值分组
            mod_groups = {}
            for pci_info in compliant_pcis:
                pci = pci_info[0]
                mod = pci % self.mod_value
                if mod not in mod_groups:
                    mod_groups[mod] = []
                mod_groups[mod].append(pci_info)

            # 确保每个模3组最多保留3个最佳候选
            balanced_pcis = []
            for mod in mod_groups:
                # 按距离和均衡性排序（4元组结构）
                mod_groups[mod].sort(key=lambda x: (-x[1], x[3]))
                balanced_pcis.extend(mod_groups[mod][:3])

            # 如果不同模3值不足3组，放宽距离要求
            if len(mod_groups) < 3 and len(compliant_pcis) > 20:
                print(f"  警告：候选PCI模3多样性不足({len(mod_groups)}种)，将放宽距离要求")
                # 临时放宽距离要求
                original_distance = self.reuse_distance_km
                self.reuse_distance_km = 2.0  # 临时放宽距离要求
                compliant_pcis = self.get_reuse_compliant_pcis(
                    target_lat, target_lon, earfcn_dl, enodeb_id, cell_id, target_mod
                )
                self.reuse_distance_km = original_distance  # 恢复原距离要求
                # 重新平衡
                mod_groups = {}
                for pci_info in compliant_pcis:
                    pci = pci_info[0]
                    mod = pci % self.mod_value
                    if mod not in mod_groups:
                        mod_groups[mod] = []
                    mod_groups[mod].append(pci_info)
                balanced_pcis = []
                for mod in mod_groups:
                    mod_groups[mod].sort(key=lambda x: (-x[1], x[3] if len(x) > 3 else 0))
                    balanced_pcis.extend(mod_groups[mod][:3])

            compliant_pcis = balanced_pcis
            print(f"  模3均衡化处理后候选PCI数量: {len(compliant_pcis)} (含{len(mod_groups)}种模3值)")
        else:
            # LTE网络直接使用get_reuse_compliant_pcis返回的结果，不进行额外处理
            print(f"  LTE网络：使用get_reuse_compliant_pcis的完整结果，共{len(compliant_pcis)}个候选PCI")

            # 对于LTE网络，如果候选PCI数量较少但大于0，直接使用所有候选
            if len(compliant_pcis) > 0 and len(compliant_pcis) <= 10:
                print(f"  候选PCI数量较少({len(compliant_pcis)})，直接使用所有候选确保同站模3错开")

        if not compliant_pcis:
            # 智能降级策略：当无法找到满足复用距离的PCI时，逐步放宽要求
            print(f"  警告：无法找到满足{self.reuse_distance_km}km复用距离的PCI，尝试智能降级")

            # 第一步：尝试放宽到3.0km（中等距离要求）
            original_distance = self.reuse_distance_km
            self.reuse_distance_km = 3.0
            compliant_pcis = self.get_reuse_compliant_pcis(
                target_lat, target_lon, earfcn_dl, enodeb_id, cell_id, target_mod
            )

            if not compliant_pcis:
                # 第二步：尝试放宽到2.0km（较低距离要求）
                self.reuse_distance_km = 2.0
                compliant_pcis = self.get_reuse_compliant_pcis(
                    target_lat, target_lon, earfcn_dl, enodeb_id, cell_id, target_mod
                )

            # 恢复原距离要求
            self.reuse_distance_km = original_distance

            if compliant_pcis:
                print(f"  智能降级成功：找到{len(compliant_pcis)}个候选PCI")
                # 使用降级后的候选PCI继续规划
                best_pci_info = compliant_pcis[0]
                best_pci = best_pci_info[0]
                min_distance = best_pci_info[1]

                # 验证智能降级找到的PCI是否满足原始约束条件
                if self.validate_fallback_pci_meets_constraints(best_pci, target_lat, target_lon, earfcn_dl, enodeb_id, cell_id):
                    print(f"    [验证通过] 智能降级PCI {best_pci} 实际满足原始约束条件")
                    reason = 'fallback_validated_meets_constraints'
                else:
                    print(f"    [验证失败] 智能降级PCI {best_pci} 不满足原始约束条件")
                    reason = f'no_compliant_pci_fallback_downgrade_{self.reuse_distance_km}km_to_{min(3.0, 2.0)}km'

                return best_pci, reason, earfcn_dl, min_distance
            else:
                # 如果降级后仍然找不到，使用改进的保底方案
                self.failure_reasons['no_compliant_pci'].append(f"{enodeb_id}-{cell_id}")

                # 关键修复：改进保底方案，避免同站同PCI冲突
                fallback_pci = self.calculate_fallback_pci(enodeb_id, cell_id, target_lat, target_lon)

                # 验证保底PCI是否满足原始约束条件
                if self.validate_fallback_pci_meets_constraints(fallback_pci, target_lat, target_lon, earfcn_dl, enodeb_id, cell_id):
                    print(f"    [验证通过] 保底PCI {fallback_pci} 实际满足原始约束条件")
                    fallback_reason = 'fallback_validated_meets_constraints'
                else:
                    print(f"    [验证失败] 保底PCI {fallback_pci} 不满足原始约束条件")
                    fallback_reason = 'no_compliant_pci_fallback'

                return fallback_pci, fallback_reason, earfcn_dl, 0.0

        # 选择最优的PCI（经过多重优先级排序后的第一个）
        best_pci_info = compliant_pcis[0]
        best_pci = best_pci_info[0]
        min_distance = best_pci_info[1]
        has_conflict = best_pci_info[2] if len(best_pci_info) > 2 else False
        balance_score = best_pci_info[3] if len(best_pci_info) > 3 else 0

        # 检查选择的PCI的特性
        best_pci_mod = best_pci % self.mod_value

        # 获取同站点小区的模值信息用于详细显示
        same_site_mods = set()
        same_site_cells = self.get_same_site_cells(target_lat, target_lon, enodeb_id, cell_id)
        for cell in same_site_cells:
            # 只考虑已经分配了PCI的小区（pci不为空且不为-1）
            if pd.notna(cell.get('pci')) and cell.get('pci') != -1:
                same_site_mods.add(int(cell['pci']) % self.mod_value)

        same_site_mod_ok = best_pci_mod not in same_site_mods

        # 显示PCI选择的详细信息
        print(f"    小区{enodeb_id}-{cell_id}: 候选PCI数量={len(compliant_pcis)}")
        print(f"    同站点已有模{self.mod_value}值: {sorted(list(same_site_mods))}")

        if len(compliant_pcis) > 1:
            top_candidates = compliant_pcis[:min(3, len(compliant_pcis))]
            print(f"    前{len(top_candidates)}个候选PCI: {[(pci, f'{dist:.2f}km') for pci, dist, _, _ in top_candidates]}")

        print(f"    选择PCI={best_pci} (mod{self.mod_value}={best_pci_mod})")
        print(f"      复用距离: {min_distance:.2f}km")
        print(f"      同站点模{self.mod_value}冲突: {'无' if same_site_mod_ok else '有'}")

        if not same_site_mod_ok:
            print(f"      警告: 与同站点小区模{self.mod_value}值冲突!")
            conflict_cells = []
            for cell in same_site_cells:
                # 只考虑已经分配了PCI的小区（pci不为空且不为-1）
                if pd.notna(cell.get('pci')) and cell.get('pci') != -1:
                    cell_mod = int(cell['pci']) % self.mod_value
                    if cell_mod == best_pci_mod:
                        conflict_cells.append(f"小区{cell['cell_id']}(PCI{cell['pci']})")
            if conflict_cells:
                print(f"      冲突小区: {', '.join(conflict_cells)}")

        if min_distance != float('inf'):
            if abs(min_distance - self.reuse_distance_km) <= 2:
                print(f"      均衡性: 优秀 (接近{self.reuse_distance_km}km阈值)")
            elif min_distance <= self.reuse_distance_km + 5:
                print(f"      均衡性: 良好 (适中距离)")
            else:
                print(f"      均衡性: 一般 (距离较远)")

        # 确定分配原因
        if self.inherit_mod and target_mod is not None:
            # 继承模式：由于严格筛选，分配的PCI必然匹配mod值
            if self.network_type == "LTE":
                reason = 'strict_mod3_inheritance'
            else:
                reason = 'strict_mod30_inheritance'
        else:
            # 自由规划模式：仅考虑复用距离
            reason = 'free_planning_reuse_compliant'

        return best_pci, reason, earfcn_dl, min_distance

    def calculate_pci_mod(self, pci_value) -> Optional[int]:
        """
        计算PCI的模值（LTE为mod3，NR为mod30）
        """
        if pci_value is None or pci_value == -1 or pd.isna(pci_value):
            return None
        try:
            return int(pci_value) % self.mod_value
        except (ValueError, TypeError):
            return None

    def update_cell_pci(self, enodeb_id: int, cell_id: int, assigned_pci: Optional[int]):
        """
        更新小区的PCI值

        Args:
            assigned_pci: 分配的PCI值，如果为None表示无法分配PCI
        """
        if assigned_pci is None:
            # 无法分配PCI，不进行更新
            print(f"      [DEBUG] 无法为基站{enodeb_id}-小区{cell_id}分配PCI，跳过更新")
            return

        mask = (self.target_cells['enodeb_id'] == enodeb_id) & (self.target_cells['cell_id'] == cell_id)
        self.target_cells.loc[mask, 'pci'] = assigned_pci

        # 同时更新合并数据
        if self.all_cells_combined is not None:
            combined_mask = ((self.all_cells_combined['enodeb_id'] == enodeb_id) &
                           (self.all_cells_combined['cell_id'] == cell_id))
            # 移除cell_type过滤条件，确保所有类型的小区都能更新
            self.all_cells_combined.loc[combined_mask, 'pci'] = assigned_pci

        # 彻底清除所有相关缓存（关键修复）
        # 1. 清除同站点缓存
        # 获取当前小区的经纬度用于清除相关缓存
        cell_info, _ = self.get_cell_info(enodeb_id, cell_id)
        if cell_info is not None and not pd.isna(cell_info['lat']) and not pd.isna(cell_info['lon']):
            # 清除基于经纬度的同站点缓存（与当前小区位置相关的缓存）
            target_lat = cell_info['lat']
            target_lon = cell_info['lon']
            same_site_keys = [key for key in self.same_site_cache.keys()
                             if len(key) >= 2 and
                             abs(key[0] - target_lat) < 0.0001 and
                             abs(key[1] - target_lon) < 0.0001]
            for key in same_site_keys:
                if key in self.same_site_cache:  # 检查键是否仍然存在
                    del self.same_site_cache[key]
        else:
            # 如果无法获取位置信息，清除所有缓存
            self.same_site_cache.clear()

        # 2. 清除距离缓存（避免使用旧的干扰计算结果）
        cell_info, _ = self.get_cell_info(enodeb_id, cell_id)
        distance_keys = []  # 初始化变量，提供安全默认值
        if cell_info is not None and not pd.isna(cell_info['lat']) and not pd.isna(cell_info['lon']):
            distance_keys = [k for k in self.distance_cache.keys()
                           if k[0] == round(cell_info['lat'], 6)
                           and k[1] == round(cell_info['lon'], 6)]
            for key in distance_keys:
                del self.distance_cache[key]

        # 3. 清除PCI验证缓存
        pci_keys = [k for k in self.pci_validity_cache.keys()
                   if k[3] == enodeb_id or k[4] == cell_id]
        for key in pci_keys:
            del self.pci_validity_cache[key]

        print(f"      [DEBUG] 更新基站{enodeb_id}-小区{cell_id}的PCI为{assigned_pci}，清除缓存:")
        print(f"        - 同站点缓存: {len(same_site_keys)}项")
        print(f"        - 距离缓存: {len(distance_keys)}项")
        print(f"        - PCI验证缓存: {len(pci_keys)}项")

    def calculate_final_min_reuse_distance(self, result_df: pd.DataFrame) -> List:
        """
        计算每个小区的最终最小PCI复用距离（验证分配结果）
        返回混合类型列表，数值距离保持为float，特殊情况使用字符串标记
        """
        print("\n正在验证最终复用距离...")
        min_reuse_distances = []

        for idx, cell in result_df.iterrows():
            # 根据网络类型获取正确的列名
            if self.network_type == "NR":
                enodeb_id = cell.get('gNodeBID')
            else:
                enodeb_id = cell.get('eNodeBID')
            cell_id = cell.get('CellID')
            assigned_pci = cell.get('分配的PCI')

            # 处理PCI分配为None的情况（无位置信息的小区）
            if assigned_pci is None:
                min_reuse_distances.append("位置信息缺失")
                continue

            if assigned_pci == -1:
                min_reuse_distances.append("分配失败")
                continue

            # 获取当前小区信息
            cell_info, status = self.get_cell_info(enodeb_id, cell_id)

            if status != 'success' or pd.isna(cell_info['lat']) or pd.isna(cell_info['lon']):
                min_reuse_distances.append("位置信息缺失")
                continue

            # 验证实际复用距离（同频检查）
            _, actual_min_distance = self.validate_pci_reuse_distance(
                assigned_pci, cell_info['lat'], cell_info['lon'], cell_info['earfcn_dl'], enodeb_id, cell_id
            )

            # 保持数值格式，特殊情况使用字符串
            if actual_min_distance == float('inf'):
                min_reuse_distances.append("无复用PCI")
            else:
                min_reuse_distances.append(round(actual_min_distance, 2))

            if (idx + 1) % 100 == 0 or (idx + 1) == len(result_df):
                print(f"已验证 {idx+1}/{len(result_df)} 个小区的复用距离")

        return min_reuse_distances

    def plan_pci_with_reuse_priority(self) -> pd.DataFrame:
        """
        执行优先确保复用距离的PCI规划
        """
        if self.cells_to_plan is None:
            raise ValueError("请先加载数据")

        start_time = time.time()

        print(f"\n开始{self.network_type}网络PCI规划（优先确保同频PCI复用距离）")
        print(f"同频PCI最小复用距离: {self.reuse_distance_km}km (优先级最高)")
        print(f"PCI范围: 0-{max(self.pci_range)}")

        if self.network_type == "LTE":
            if self.lte_inherit_mod3:
                print(f"规划模式: 严格mod3继承 - 只分配匹配原PCI mod3值的PCI")
            else:
                print(f"规划模式: 自由规划 - 不考虑mod3值，优先复用距离")
        else:
            if self.nr_inherit_mod30:
                print(f"规划模式: 严格mod30继承 - 只分配匹配原PCI mod30值的PCI")
            else:
                print(f"规划模式: 自由规划 - 不考虑mod30值，优先复用距离")

        print(f"核心规则: 只有同频同PCI的小区才需要遵循最小复用距离")

        # 重置统计
        self.failure_reasons = {
            'cell_not_found': [],
            'no_location': [],
            'reuse_distance_violation': [],
            'no_compliant_pci': [],
            'fallback_assignments': []
        }

        # 复制规划小区数据
        result_df = self.cells_to_plan.copy()

        # 初始化新列
        cell_names = []
        assigned_pcis = []
        original_pcis = []
        original_mods = []
        assigned_mods = []
        mod_same_flags = []
        earfcn_dls = []
        assignment_reasons = []
        predicted_min_distances = []

        total_cells = len(result_df)

        # 关键修改：按基站分组，确保同站点的小区一起处理
        # 这样可以避免同站同模问题
        print("按基站分组处理小区，确保同站点小区模值错开...")

        # 创建基站分组
        if self.network_type == "NR":
            enodeb_col = 'gNodeBID'
        else:
            enodeb_col = 'eNodeBID'

        # 关键修改：按物理位置分组处理，确保同站点小区模值错开...
        print("按物理位置分组处理小区，确保同站点小区模值错开...")

        # 创建位置到小区的映射
        location_groups = {}
        for idx, cell in result_df.iterrows():
            if self.network_type == "NR":
                enodeb_id = cell.get('gNodeBID')
            else:
                enodeb_id = cell.get('eNodeBID')
            cell_id = cell.get('CellID')

            # 获取小区位置信息
            cell_info, _ = self.get_cell_info(enodeb_id, cell_id)
            if cell_info is not None:
                lat = cell_info.get('lat')
                lon = cell_info.get('lon')
                if pd.notna(lat) and pd.notna(lon):
                    # 使用经纬度作为位置键（保留6位小数精度）
                    location_key = f"{lat:.6f},{lon:.6f}"
                    if location_key not in location_groups:
                        location_groups[location_key] = []
                    location_groups[location_key].append((idx, cell))

        print(f"发现 {len(location_groups)} 个不同的物理位置")

        # 创建处理顺序：同位置的小区连续处理
        processing_order = []
        for location_key, cells in location_groups.items():
            for idx, cell in cells:
                processing_order.append((idx, cell))

        # 处理没有位置信息的小区（放在最后）
        remaining_cells = []
        for idx, cell in result_df.iterrows():
            if not any(idx == order_idx for order_idx, _ in processing_order):
                remaining_cells.append((idx, cell))
        processing_order.extend(remaining_cells)

        print(f"总共 {len(processing_order)} 个小区需要处理")

        # 按新的顺序处理小区，但保持原始索引对应关系
        # 关键修复：使用索引对应的数据结构，避免顺序错位
        results_by_index = {}

        for order_idx, (idx, cell) in enumerate(processing_order):
            # 根据网络类型选择正确的基站ID列
            if self.network_type == "NR":
                enodeb_id = cell.get('gNodeBID')
            else:
                enodeb_id = cell.get('eNodeBID')
            cell_id = cell.get('CellID')

            # 显示进度
            if (order_idx + 1) % 50 == 0 or (order_idx + 1) == len(processing_order):
                print(f"正在规划小区 {order_idx+1}/{len(processing_order)}")

            # 获取原PCI信息和小区名称
            cell_info, _ = self.get_cell_info(enodeb_id, cell_id)
            original_pci = cell_info['pci'] if cell_info is not None else None
            # 获取小区名称并清理
            if cell_info is not None and 'cell_name' in cell_info and not pd.isna(cell_info['cell_name']):
                cell_name = str(cell_info['cell_name']).strip()
                # 过滤掉模板行的标识
                if cell_name in ['非必填', 'UserLabel', '必填']:
                    cell_name = f"小区_{enodeb_id}_{cell_id}"
            else:
                cell_name = f"小区_{enodeb_id}_{cell_id}"

            # 分配PCI（优先确保复用距离）
            assigned_pci, reason, earfcn_dl, predicted_distance = self.assign_pci_with_reuse_priority(enodeb_id, cell_id)

            # 仅当成功分配PCI时才更新小区PCI（优化：无位置信息不更新）
            if assigned_pci is not None and assigned_pci != -1:
                self.update_cell_pci(enodeb_id, cell_id, assigned_pci)

            # 记录特殊情况
            if 'not_found' in reason:
                self.failure_reasons['cell_not_found'].append(f"{enodeb_id}-{cell_id}")
            elif 'no_location' in reason:
                self.failure_reasons['no_location'].append(f"{enodeb_id}-{cell_id}")
            elif 'no_compliant' in reason:
                self.failure_reasons['no_compliant_pci'].append(f"{enodeb_id}-{cell_id}")
            elif 'fallback' in reason:
                self.failure_reasons['fallback_assignments'].append(f"{enodeb_id}-{cell_id}")

            # 计算模值
            original_mod = self.calculate_pci_mod(original_pci)
            assigned_mod = self.calculate_pci_mod(assigned_pci)

            # 比较模值
            if original_mod is not None and assigned_mod is not None:
                mod_same = '是' if original_mod == assigned_mod else '否'
            else:
                mod_same = '无法比较'

            # 关键修复：按原始索引存储结果，确保对应关系正确
            results_by_index[idx] = {
                'cell_name': cell_name,
                'original_pci': original_pci,
                'original_mod': original_mod,
                'assigned_pci': assigned_pci,
                'assigned_mod': assigned_mod,
                'mod_same': mod_same,
                'earfcn_dl': earfcn_dl,
                'assignment_reason': reason,
                'predicted_distance': predicted_distance
            }

            # 仅当成功分配PCI时才检查同站模冲突（优化：无位置信息不检查）
            if assigned_pci is not None and assigned_pci != -1:
                cell_info, _ = self.get_cell_info(enodeb_id, cell_id)
                if cell_info is not None and pd.notna(cell_info.get('lat')) and pd.notna(cell_info.get('lon')):
                    same_site_mods = set()
                    same_site_cells = self.get_same_site_cells(cell_info['lat'], cell_info['lon'], enodeb_id, cell_id)
                    for cell in same_site_cells:
                        if pd.notna(cell.get('pci')) and cell.get('pci') != -1:
                            same_site_mods.add(int(cell['pci']) % self.mod_value)

                    assigned_mod = assigned_pci % self.mod_value
                    if assigned_mod in same_site_mods:
                        print(f"      严重警告：小区{enodeb_id}-{cell_id}分配的PCI={assigned_pci}与同站点小区模{self.mod_value}冲突！")
                        print(f"      同站点已有模{self.mod_value}值: {sorted(list(same_site_mods))}")
                        print(f"      冲突详情：")
                        for cell in same_site_cells:
                            if pd.notna(cell.get('pci')) and cell.get('pci') != -1:
                                cell_mod = int(cell['pci']) % self.mod_value
                                if cell_mod == assigned_mod:
                                    print(f"        - 基站{cell.get('enodeb_id', 'N/A')}-小区{cell.get('cell_id', 'N/A')}: PCI={cell['pci']}, mod{self.mod_value}={cell_mod}")
                    else:
                        print(f"      检查通过：小区{enodeb_id}-{cell_id}的PCI={assigned_pci} (mod{self.mod_value}={assigned_mod}) 无同站点冲突")
                        print(f"      同站点已有模{self.mod_value}值: {sorted(list(same_site_mods))}")
                else:
                    print(f"      警告：无法检查同站点冲突 - 缺少位置信息")

        # 关键修复：按原始索引正确组装数据，避免顺序错位
        # 先保存原有的列
        original_columns = list(result_df.columns)

        # 根据网络类型决定是否包含eNodeBID列
        if self.network_type == "NR":
            # NR网络：删除eNodeBID列，只保留CellID和小区名称
            print("    NR网络：删除冗余的eNodeBID列")
            filtered_columns = [col for col in original_columns if col != 'eNodeBID']
            # 在CellID后插入小区名称
            if 'CellID' in filtered_columns:
                cellid_index = filtered_columns.index('CellID')
                new_columns = filtered_columns[:cellid_index+1] + ['小区名称'] + filtered_columns[cellid_index+1:]
            else:
                new_columns = ['小区名称'] + filtered_columns
        else:
            # LTE网络：保持原有结构，在第3列插入小区名称
            new_columns = original_columns[:2] + ['小区名称'] + original_columns[2:]

        # 关键修复：按原始索引顺序组装数据，确保对应关系正确
        result_df_new = pd.DataFrame()

        # 首先组装原始列（保持原始顺序）
        for col in new_columns:
            if col == '小区名称':
                # 按原始索引顺序组装小区名称
                ordered_cell_names = []
                for idx in result_df.index:
                    if idx in results_by_index:
                        ordered_cell_names.append(results_by_index[idx]['cell_name'])
                    else:
                        # 对于没有处理结果的小区，使用默认名称
                        if self.network_type == "NR":
                            enodeb_id = result_df.loc[idx, 'gNodeBID'] if 'gNodeBID' in result_df.columns else 'Unknown'
                        else:
                            enodeb_id = result_df.loc[idx, 'eNodeBID'] if 'eNodeBID' in result_df.columns else 'Unknown'
                        cell_id = result_df.loc[idx, 'CellID'] if 'CellID' in result_df.columns else 'Unknown'
                        ordered_cell_names.append(f"小区_{enodeb_id}_{cell_id}")
                result_df_new[col] = ordered_cell_names
            elif col in original_columns:
                # 保持原始列的数据顺序
                result_df_new[col] = result_df[col].values

        # 添加其他新列（按原始索引顺序）
        ordered_original_pcis = []
        ordered_assigned_pcis = []
        ordered_original_mods = []
        ordered_assigned_mods = []
        ordered_mod_same_flags = []
        ordered_earfcn_dls = []
        ordered_assignment_reasons = []
        ordered_predicted_distances = []

        for idx in result_df.index:
            if idx in results_by_index:
                ordered_original_pcis.append(results_by_index[idx]['original_pci'])
                ordered_assigned_pcis.append(results_by_index[idx]['assigned_pci'])
                ordered_original_mods.append(results_by_index[idx]['original_mod'])
                ordered_assigned_mods.append(results_by_index[idx]['assigned_mod'])
                ordered_mod_same_flags.append(results_by_index[idx]['mod_same'])
                ordered_earfcn_dls.append(results_by_index[idx]['earfcn_dl'])
                ordered_assignment_reasons.append(results_by_index[idx]['assignment_reason'])
                ordered_predicted_distances.append(results_by_index[idx]['predicted_distance'])
            else:
                # 对于没有处理结果的小区，使用默认值
                ordered_original_pcis.append(None)
                ordered_assigned_pcis.append(None)
                ordered_original_mods.append(None)
                ordered_assigned_mods.append(None)
                ordered_mod_same_flags.append('无法比较')
                ordered_earfcn_dls.append(0.0)
                ordered_assignment_reasons.append('未处理')
                ordered_predicted_distances.append(0.0)

        result_df_new['网络类型'] = self.network_type
        result_df_new['原PCI'] = ordered_original_pcis
        result_df_new['分配的PCI'] = ordered_assigned_pcis

        # 根据网络类型设置模值列名
        if self.network_type == "LTE":
            result_df_new['原PCI模3'] = ordered_original_mods
            result_df_new['新PCI模3'] = ordered_assigned_mods
            result_df_new['模3是否相同'] = ordered_mod_same_flags
        else:
            result_df_new['原PCI模30'] = ordered_original_mods
            result_df_new['新PCI模30'] = ordered_assigned_mods
            result_df_new['模30是否相同'] = ordered_mod_same_flags

        result_df_new['earfcnDl'] = ordered_earfcn_dls
        result_df_new['分配原因'] = ordered_assignment_reasons

        # 更新result_df
        result_df = result_df_new

        # 计算最终验证的最小复用距离
        final_min_distances = self.calculate_final_min_reuse_distance(result_df)
        result_df['最小复用距离(km)'] = final_min_distances

        # 计算执行时间
        end_time = time.time()
        execution_time = end_time - start_time

        # 统计结果
        self.print_reuse_focused_statistics(result_df, execution_time, self.params_file)

        return result_df

    def print_reuse_focused_statistics(self, result_df: pd.DataFrame, execution_time: float, params_file: str = None):
        """
        打印复用距离优先的统计信息
        """
        print(f"\n=== {self.network_type}网络PCI规划完成（同频PCI复用距离优先） ===")
        if params_file:
            print(f"参数文件: {os.path.basename(params_file)}")
        print(f"执行时间: {execution_time:.2f} 秒")
        print(f"平均每个小区处理时间: {execution_time/len(result_df):.4f} 秒")

        total_cells = len(result_df)
        success_count = len(result_df[result_df['分配的PCI'] != -1])

        print(f"规划统计:")
        print(f"  总小区数: {total_cells}")
        print(f"  成功分配: {success_count} ({success_count/total_cells*100:.1f}%)")

        # 同频PCI复用距离合规性统计
        distance_col = result_df['最小复用距离(km)']

        # 分类统计
        no_reuse_cells = result_df[distance_col == "无复用PCI"]
        failed_cells = result_df[distance_col == "分配失败"]
        location_missing_cells = result_df[distance_col == "位置信息缺失"]

        # 数值距离的小区
        numeric_distances = []
        numeric_indices = []
        for idx, dist in distance_col.items():
            if isinstance(dist, str) and dist not in ["无复用PCI", "分配失败", "位置信息缺失"]:
                try:
                    numeric_distances.append(float(dist))
                    numeric_indices.append(idx)
                except ValueError:
                    pass

        numeric_df = result_df.loc[numeric_indices] if numeric_indices else pd.DataFrame()

        if len(numeric_distances) > 0:
            compliant_numeric = [d for d in numeric_distances if d >= self.reuse_distance_km]
            violation_numeric = [d for d in numeric_distances if d < self.reuse_distance_km]
        else:
            compliant_numeric = []
            violation_numeric = []

        print(f"\n同频PCI复用距离详细统计:")
        print(f"  最小复用距离要求: {self.reuse_distance_km}km (仅适用于同频同PCI小区)")
        print(f"  无复用PCI小区: {len(no_reuse_cells)} ({len(no_reuse_cells)/total_cells*100:.1f}%)")
        print(f"  有复用且合规小区: {len(compliant_numeric)} ({len(compliant_numeric)/total_cells*100:.1f}%)")
        print(f"  有复用但违规小区: {len(violation_numeric)} ({len(violation_numeric)/total_cells*100:.1f}%)")

        # 复用距离分布统计（仅针对数值距离）
        if numeric_distances:
            print(f"\n复用距离分布统计 (有复用PCI的小区):")
            print(f"  平均复用距离: {np.mean(numeric_distances):.2f}km")
            print(f"  最小复用距离: {np.min(numeric_distances):.2f}km")
            print(f"  最大复用距离: {np.max(numeric_distances):.2f}km")
            print(f"  中位数复用距离: {np.median(numeric_distances):.2f}km")

            # 距离区间分布
            distance_ranges = [
                (0, self.reuse_distance_km, f"< {self.reuse_distance_km}km (违规)"),
                (self.reuse_distance_km, self.reuse_distance_km + 2, f"{self.reuse_distance_km}-{self.reuse_distance_km + 2}km (接近阈值)"),
                (self.reuse_distance_km + 2, self.reuse_distance_km + 5, f"{self.reuse_distance_km + 2}-{self.reuse_distance_km + 5}km (适中)"),
                (self.reuse_distance_km + 5, float('inf'), f"> {self.reuse_distance_km + 5}km (较远)")
            ]

            print(f"  距离区间分布:")
            for min_dist, max_dist, desc in distance_ranges:
                count = sum(1 for d in numeric_distances if min_dist <= d < max_dist)
                if count > 0:
                    print(f"    {desc}: {count} ({count/len(numeric_distances)*100:.1f}%)")

        if len(failed_cells) > 0:
            print(f"  分配失败小区: {len(failed_cells)} ({len(failed_cells)/total_cells*100:.1f}%)")
        if len(location_missing_cells) > 0:
            print(f"  位置信息缺失: {len(location_missing_cells)} ({len(location_missing_cells)/total_cells*100:.1f}%)")

        # 违规小区详情
        if len(violation_numeric) > 0:
            print(f"\n  违规小区详情 (前10个):")
            violation_indices = [numeric_indices[i] for i, d in enumerate(numeric_distances) if d < self.reuse_distance_km][:10]
            for idx in violation_indices:
                cell = result_df.loc[idx]
                print(f"    {cell['eNodeBID']}-{cell['CellID']}: 实际距离 {cell['最小复用距离(km)']}km")

        # 模值继承统计
        if self.inherit_mod:
            if self.network_type == "LTE":
                mod_col = '模3是否相同'
                mod_name = "模3"
            else:
                mod_col = '模30是否相同'
                mod_name = "模30"

            mod_inherited = result_df[result_df[mod_col] == '是']
            mod_rate = len(mod_inherited) / total_cells * 100
            print(f"\n{mod_name}继承统计:")
            print(f"  {mod_name}继承成功: {len(mod_inherited)} ({mod_rate:.1f}%)")

        # 同站点模值冲突分析
        self.analyze_same_site_mod_conflicts(result_df)

        # 分配原因统计
        print(f"\n分配原因统计:")
        reason_counts = result_df['分配原因'].value_counts()
        for reason, count in reason_counts.items():
            print(f"  {reason}: {count} ({count/total_cells*100:.1f}%)")

        # 特殊情况统计
        special_cases = 0
        for reason, cells in self.failure_reasons.items():
            if cells:
                special_cases += len(cells)
                print(f"  {reason}: {len(cells)} 个小区")

        if special_cases == 0:
            print("\n[成功] 所有小区均正常分配PCI")

        print(f"\n距离状态说明:")
        print(f"  '无复用PCI': 该小区PCI在同频点内无其他小区使用，满足复用距离要求")
        print(f"  数值(如3.45): 同频同PCI小区间的最小距离(km)")
        print(f"  '分配失败': PCI分配过程失败")
        print(f"  '位置信息缺失': 小区缺少经纬度信息，无法计算距离")

        print(f"\n规划完成！")

    def analyze_same_site_mod_conflicts(self, result_df: pd.DataFrame):
        """
        分析同站点模值冲突情况（基于经纬度判断）
        """
        print(f"\n同位置模{self.mod_value}冲突分析:")

        # 按经纬度分组统计
        conflicts = []
        total_locations = 0
        conflict_locations = 0

        # 检查是否有经纬度列
        lat_col = '小区纬度' if '小区纬度' in result_df.columns else 'lat'
        lon_col = '小区经度' if '小区经度' in result_df.columns else 'lon'


        if lat_col in result_df.columns and lon_col in result_df.columns:
            # 按经纬度分组（使用经纬度组合作为键）
            location_groups = result_df.groupby([lat_col, lon_col])

            for (lat, lon), group in location_groups:
                total_locations += 1

                # 获取该位置下所有小区的PCI和模值
                location_cells = []
                for _, cell in group.iterrows():
                    pci = cell['分配的PCI']
                    if pd.notna(pci) and pci != -1:
                        mod_value = int(pci) % self.mod_value
                        # 获取基站ID用于显示
                        if self.network_type == "LTE":
                            enodeb_col = 'eNodeBID'
                        else:
                            enodeb_col = 'gNodeBID'
                        enodeb_id = cell[enodeb_col] if enodeb_col in cell else 'N/A'

                        location_cells.append({
                            'cell_id': cell['CellID'],
                            'pci': int(pci),
                            'mod': mod_value,
                            'enodeb_id': enodeb_id
                        })

                if len(location_cells) <= 1:
                    continue  # 单小区位置无冲突可能

                # 检查模值冲突
                mod_values = [cell['mod'] for cell in location_cells]
                unique_mods = set(mod_values)

                if len(unique_mods) < len(mod_values):
                    # 发现冲突
                    conflict_locations += 1

                    # 找出冲突的模值
                    conflict_mods = []
                    for mod in unique_mods:
                        mod_count = mod_values.count(mod)
                        if mod_count > 1:
                            conflict_cells = [cell for cell in location_cells if cell['mod'] == mod]
                            conflict_mods.append((mod, conflict_cells))

                    conflicts.append({
                        'lat': lat,
                        'lon': lon,
                        'total_cells': len(location_cells),
                        'conflicts': conflict_mods
                    })

            # 输出统计结果
            print(f"  总位置数: {total_locations}")
            print(f"  有模{self.mod_value}冲突的位置: {conflict_locations} ({conflict_locations/total_locations*100:.1f}%)")
            print(f"  无模{self.mod_value}冲突的位置: {total_locations - conflict_locations} ({(total_locations - conflict_locations)/total_locations*100:.1f}%)")

            if conflicts:
                print(f"\n  冲突详情 (前5个位置):")
                for i, conflict in enumerate(conflicts[:5]):
                    lat = conflict['lat']
                    lon = conflict['lon']
                    print(f"    位置({lat}, {lon}): {conflict['total_cells']}个小区")
                    for mod, cells in conflict['conflicts']:
                        cell_info = ', '.join([f"小区{cell['cell_id']}(PCI{cell['pci']})" for cell in cells])
                        print(f"      模{self.mod_value}={mod}冲突: {cell_info}")

            if conflict_locations == 0:
                print(f"  优秀！所有位置都避免了同位置模{self.mod_value}冲突")
            elif conflict_locations / total_locations <= 0.1:
                print(f"  良好！大部分位置避免了同位置模{self.mod_value}冲突")
            else:
                print(f"  需要关注：较多位置存在同位置模{self.mod_value}冲突")
        else:
            print(f"  警告：缺少经纬度信息，无法进行同位置模{self.mod_value}冲突分析")


class PlaceholderEntry(ttk.Entry):
    """Custom Entry widget with placeholder text support"""
    def __init__(self, master=None, placeholder="", color="black", **kwargs):
        super().__init__(master, **kwargs)
        self.placeholder = placeholder
        self.placeholder_color = 'grey'
        self.default_fg_color = color  # User input color

        self.bind("<FocusIn>", self._focus_in)
        self.bind("<FocusOut>", self._focus_out)

        self._put_placeholder()
    
    def _put_placeholder(self):
        if not self.get():
            self.insert(0, self.placeholder)
            self['foreground'] = self.placeholder_color
    
    def _focus_in(self, event):
        if self['foreground'] == self.placeholder_color:
            self.delete('0', 'end')
            self['foreground'] = self.default_fg_color
    
    def _focus_out(self, event):
        if not self.get():
            self._put_placeholder()
    
    def get(self):
        content = super().get()
        if content == self.placeholder and self['foreground'] == self.placeholder_color:
            return ""
        return content



class Tooltip:
    """Simple tooltip class for showing bubble notifications"""
    def __init__(self, widget, text=''):
        self.widget = widget
        self.text = text
        self.tipwindow = None
        self.id = None
        self.x = self.y = 0
        self.x_offset = 10  # X offset from mouse pointer
        self.y_offset = 10  # Y offset from mouse pointer

        # Bind mouse events
        self.widget.bind("<Enter>", self.on_enter)
        self.widget.bind("<Leave>", self.on_leave)
        self.widget.bind("<Motion>", self.on_motion)

    def on_enter(self, event=None):
        """When mouse enters the widget"""
        self.schedule_show()

    def on_leave(self, event=None):
        """When mouse leaves the widget"""
        self.hide()

    def on_motion(self, event=None):
        """When mouse moves over the widget"""
        self.x = event.x_root + self.x_offset
        self.y = event.y_root + self.y_offset

    def schedule_show(self):
        """Schedule tooltip to show after delay"""
        self.widget.after_cancel(self.id) if self.id else None
        self.id = self.widget.after(500, self.show)  # Show after 500ms

    def show(self, text=None, x=None, y=None, duration=None):
        """Show tooltip at specified position for duration (ms)"""
        if text is not None:
            self.text = text
        if x is not None:
            self.x = x
        if y is not None:
            self.y = y

        if self.tipwindow or not self.text:
            return

        # Calculate position if not provided
        if x is None or y is None:
            try:
                x = self.x if hasattr(self, 'x') else 0
                y = self.y if hasattr(self, 'y') else 0
            except:
                # Fallback to mouse position if not set
                x = self.widget.winfo_pointerx() + self.x_offset
                y = self.widget.winfo_pointery() + self.y_offset

        # Create tooltip window
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(1)  # Remove window decorations
        tw.wm_geometry("+%d+%d" % (x, y))

        # Create label with styling
        label = tk.Label(tw, text=self.text, justify=tk.LEFT,
                        background="#ffffe0", foreground="black",
                        relief=tk.SOLID, borderwidth=1,
                        font=("TkDefaultFont", 9))
        label.pack(ipadx=1)

        # Auto-hide after duration if specified
        if duration is not None:
            self.widget.after(duration, self.hide)

    def hide(self):
        """Hide the tooltip"""
        self.widget.after_cancel(self.id) if self.id else None
        if self.tipwindow:
            self.tipwindow.destroy()
            self.tipwindow = None


class PCIGUIApp:
    def __init__(self, root):
        self.root = root
        self.root.title("NetworkPlanningTool_V1.0")
        self.root.geometry("1350x700")

        # Configure root window to be responsive
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        # Initialize variables
        import glob

        # Auto-detect latest files
        cell_files = glob.glob("待规划小区/*.xlsx")
        latest_cells_file = self._get_latest_file(cell_files)

        param_files = glob.glob("全量工参/ProjectParameter_mongoose*.xlsx")
        latest_params_file = self._get_latest_file(param_files)

        baseline_files = glob.glob("全量工参/BaselineLab_*.zip")
        latest_baseline_file = self._get_latest_file(baseline_files)

        # Initialize StringVars with default files
        self.cells_file = tk.StringVar(value=latest_cells_file or "")
        self.params_file = tk.StringVar(value=latest_params_file or "")
        self.param_update_file = tk.StringVar(value=latest_params_file or "")
        self.online_param_file = tk.StringVar(value=latest_baseline_file or "")
        self.neighbor_cells_file = tk.StringVar(value=latest_cells_file or "")
        self.neighbor_params_file = tk.StringVar(value=latest_params_file or "")
        self.coverage_distance_factor = tk.DoubleVar(value=5/9)  # k系数
        self.coverage_radius_factor = tk.DoubleVar(value=5/9)   # m系数
        self.reuse_distance = tk.DoubleVar(value=3.0)
        self.lte_inherit_mod3 = tk.BooleanVar()
        self.nr_inherit_mod30 = tk.BooleanVar()
        self.network_type = tk.StringVar(value="LTE")  # Default to LTE since Both option is removed
        self.pci_range_enabled = tk.BooleanVar(value=False)
        self.pci_range_var = tk.StringVar(value="0-503")
        
        # Store clicked column index for copy functionality
        self.nr_clicked_column = None
        self.lte_clicked_column = None

        # Initialize tooltip for copy notifications
        self.tooltip = Tooltip(self.root, '')

        # Setup custom styles
        self.setup_styles()

        # Create a queue for thread communication
        self.queue = queue.Queue()

        # Create main frames
        self.create_main_frame()
        self.create_status_bar()

        # Bind the queue processing to the main thread
        self.process_queue()

    def setup_styles(self):
        """Setup custom styles for the GUI"""
        style = ttk.Style()
        # Configure normal entry style
        style.configure("Normal.TEntry", fieldbackground="white")
        # Configure disabled entry style
        style.configure("Disabled.TEntry", fieldbackground="#d3d3d3")  # Light gray

    def toggle_pci_range_state(self):
        """Toggle the state of the PCI range entry field based on checkbox"""
        if self.pci_range_enabled.get():
            self.pci_range_entry.config(state="normal")
            # Change the background color to white when enabled
            self.pci_range_entry.config(style="Normal.TEntry")
        else:
            self.pci_range_entry.config(state="disabled")
            # Change the background color to gray when disabled
            self.pci_range_entry.config(style="Disabled.TEntry")

    def on_window_resize(self, event):
        """Handle window resize events to maintain proper spacing"""
        # Only process root window resize events
        if event.widget != self.root:
            return
        # This function is mainly for future expansion
        # The responsive layout is primarily handled by grid weights

    def setup_styles(self):
        """Setup custom styles for the GUI"""
        style = ttk.Style()
        # Configure normal entry style
        style.configure("Normal.TEntry", fieldbackground="white")
        # Configure disabled entry style
        style.configure("Disabled.TEntry", fieldbackground="#d3d3d3")  # Light gray

        """Create the main application frame with tabbed interface"""
        # Create notebook (tabbed interface) for different functions
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Configure the notebook to be responsive
        self.root.bind('<Configure>', self.on_window_resize)

        # Create tabs in requested order: 全量工参更新 -> 邻区规划 -> PCI规划
        self.create_param_update_tab()      # 全量工参更新 (was 现网工参更新)
        self.create_neighbor_planning_tab() # 邻区规划
        self.create_pci_planning_tab()      # PCI规划 (was PCI Planning)

    def create_pci_planning_tab(self):
        """Create the PCI Planning tab"""
        pci_frame = ttk.Frame(self.notebook)
        self.notebook.add(pci_frame, text="PCI规划")

        # Parameters frame
        param_frame = ttk.LabelFrame(pci_frame, text="规划参数")
        param_frame.pack(fill=tk.X, padx=10, pady=5)

        # Reuse distance
        ttk.Label(param_frame, text="最小复用距离 (km):").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(param_frame, textvariable=self.reuse_distance, width=10).grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)

        # Network type
        ttk.Label(param_frame, text="网络类型:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        network_combo = ttk.Combobox(param_frame, textvariable=self.network_type, values=["LTE", "NR"], state="readonly")
        network_combo.grid(row=1, column=1, sticky=tk.W, padx=5, pady=5)

        # Mod inheritance options
        ttk.Checkbutton(param_frame, text="LTE继承原PCI的模3值", variable=self.lte_inherit_mod3).grid(row=0, column=2, sticky=tk.W, padx=10, pady=5)
        ttk.Checkbutton(param_frame, text="NR继承原PCI的模30值", variable=self.nr_inherit_mod30).grid(row=1, column=2, sticky=tk.W, padx=10, pady=5)

        # PCI Range options
        pci_range_frame = ttk.LabelFrame(param_frame, text="PCI范围设置")
        pci_range_frame.grid(row=2, column=0, columnspan=3, sticky=tk.W, padx=5, pady=5)

        ttk.Label(pci_range_frame, text="PCI范围(如 0-100):").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.pci_range_var = tk.StringVar()
        self.pci_range_entry = ttk.Entry(pci_range_frame, textvariable=self.pci_range_var, width=20, state="disabled")
        self.pci_range_entry.grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)

        self.pci_range_entry = ttk.Entry(pci_range_frame, textvariable=self.pci_range_var, width=20, state="disabled")
        self.pci_range_entry.grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)

        self.pci_range_enabled_checkbutton = ttk.Checkbutton(pci_range_frame, text="启用自定义PCI范围", variable=self.pci_range_enabled, command=self.toggle_pci_range_state)
        self.pci_range_enabled_checkbutton.grid(row=0, column=2, sticky=tk.W, padx=10, pady=5)

        # Buttons frame
        btn_frame = ttk.Frame(pci_frame)
        btn_frame.pack(fill=tk.X, padx=10, pady=5)

        # Configure button frame to be responsive
        btn_frame.columnconfigure(1, weight=1)  # Progress frame column

        self.plan_btn = ttk.Button(btn_frame, text="开始规划", command=self.start_pci_planning)
        self.plan_btn.pack(side=tk.LEFT, padx=5)

        # Progress bar frame with percentage label
        progress_frame = ttk.Frame(btn_frame)
        progress_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

        # Configure progress frame to be responsive
        progress_frame.columnconfigure(0, weight=1)

        self.plan_progress = ttk.Progressbar(progress_frame, mode='determinate')
        self.plan_progress.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.plan_progress_label = ttk.Label(progress_frame, text="0%")
        self.plan_progress_label.pack(side=tk.RIGHT, padx=(5,0))

        # Results text area
        result_frame = ttk.LabelFrame(pci_frame, text="规划结果")
        result_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Configure result frame to be responsive
        pci_frame.rowconfigure(2, weight=1)  # Index of result_frame (after param_frame and btn_frame)
        result_frame.rowconfigure(0, weight=1)
        result_frame.columnconfigure(0, weight=1)

        self.result_text = scrolledtext.ScrolledText(result_frame, height=15)
        self.result_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        # Configure tags for text coloring
        self.result_text.tag_config("error", foreground="red")
        self.result_text.tag_config("warning", foreground="orange")
        self.result_text.tag_config("info", foreground="black")

    def create_param_update_tab(self):
        """Create the Parameter Update tab"""
        param_frame = ttk.Frame(self.notebook)
        self.notebook.add(param_frame, text="规划数据导入及工参更新")
        param_frame.columnconfigure(0, weight=1)
        param_frame.rowconfigure(1, weight=1)

        # Initialize search variables
        self.nr_search_value = tk.StringVar()
        self.lte_search_value = tk.StringVar()

        # File selection
        file_frame = ttk.Frame(param_frame)
        file_frame.pack(fill=tk.X, padx=10, pady=5)

        # Configure file_frame to be responsive
        file_frame.columnconfigure(0, weight=1)

        # Create left frame for file selection
        selection_frame = ttk.LabelFrame(file_frame, text="文件选择")
        selection_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        selection_frame.columnconfigure(1, weight=1)  # Entry column

        #待规划小区文件选择
        ttk.Label(selection_frame, text="待规划小区文件:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(selection_frame, textvariable=self.neighbor_cells_file, width=50).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(selection_frame, text="浏览", command=self.select_neighbor_cells_file).grid(row=0, column=2, padx=5, pady=5)
        ttk.Button(selection_frame, text="模板下载", command=self.download_template).grid(row=0, column=3, padx=5, pady=5)

        # Parameters file selection
        ttk.Label(selection_frame, text="全量工参文件:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(selection_frame, textvariable=self.param_update_file, width=50).grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(selection_frame, text="浏览", command=self.select_param_update_file).grid(row=1, column=2, padx=5, pady=5)
        # No template needed for full parameter file

        # Online parameters file selection
        ttk.Label(selection_frame, text="现网工参:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(selection_frame, textvariable=self.online_param_file, width=50).grid(row=2, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(selection_frame, text="浏览", command=self.select_online_param_file).grid(row=2, column=2, padx=5, pady=5)

        # Add the "更新工参" button next to the "现网工参" browse button
        self.update_btn = ttk.Button(selection_frame, text="更新工参", command=self.start_param_update)
        self.update_btn.grid(row=2, column=3, padx=5, pady=5)

        # Create right frame for instructions
        instruction_frame = ttk.LabelFrame(file_frame, text="文件导入说明")
        instruction_frame.pack(side=tk.RIGHT, fill=tk.Y)

        # Add instructions
        instruction_text = tk.Text(instruction_frame, width=40, height=6, wrap=tk.WORD)
        instruction_text.insert(tk.END, "1、请按模板填写待规划小区和全量工参。\n\n2、现网工参文件为中兴网管300基线导出的zip压缩包，无需解压直接导入。")
        instruction_text.config(state=tk.DISABLED)  # Make it read-only
        instruction_text.pack(padx=5, pady=5)

        # Cell search frame
        search_frame = ttk.LabelFrame(param_frame, text="小区搜索")
        search_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        search_frame.columnconfigure(0, weight=1)
        search_frame.rowconfigure(0, weight=1)

        # Create notebook for NR and LTE search
        self.search_notebook = ttk.Notebook(search_frame)
        self.search_notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # NR Cell Search tab
        nr_search_frame = ttk.Frame(self.search_notebook)
        self.search_notebook.add(nr_search_frame, text="NR小区查询")
        # Configure NR search frame to be responsive
        nr_search_frame.columnconfigure(0, weight=1)
        nr_search_frame.rowconfigure(1, weight=1)  # Row 1 is the result frame
        param_frame.pack(fill=tk.X, padx=10, pady=5)

        # Max neighbors
        ttk.Label(param_frame, text="最大邻区数:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.max_neighbors = tk.IntVar(value=32)
        ttk.Entry(param_frame, textvariable=self.max_neighbors, width=10).grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)

        # Coverage circle parameters with help tooltip
        self.coverage_distance_label = ttk.Label(param_frame, text="覆盖圆距离系数k:")
        self.coverage_distance_label.grid(row=0, column=2, sticky=tk.W, padx=10, pady=5)
        self.coverage_distance_factor_entry = ttk.Entry(param_frame, textvariable=self.coverage_distance_factor, width=10)
        self.coverage_distance_factor_entry.grid(row=0, column=3, sticky=tk.W, padx=5, pady=5)

        # 添加覆盖圆距离系数的说明帮助
        self.k_help_label = ttk.Label(param_frame, text="?", foreground="blue", cursor="hand2")
        self.k_help_label.grid(row=0, column=4, sticky=tk.W, padx=(0, 5))
        self.k_tooltip = Tooltip(self.k_help_label, "覆盖圆距离系数k：站点到覆盖圆心的距离系数\n- 默认值5/9(≈0.556)，覆盖圆心位于主瓣方向距离站点(5/9)*Co处\n- k值越大，覆盖圆心离基站越远，方向性越明显，邻区范围偏向前方\n- k值越小，覆盖圆心越接近基站，覆盖越接近全向，邻区分布更均匀")

        self.coverage_radius_label = ttk.Label(param_frame, text="覆盖圆半径系数m:")
        self.coverage_radius_label.grid(row=0, column=5, sticky=tk.W, padx=10, pady=5)
        self.coverage_radius_factor_entry = ttk.Entry(param_frame, textvariable=self.coverage_radius_factor, width=10)
        self.coverage_radius_factor_entry.grid(row=0, column=6, sticky=tk.W, padx=5, pady=5)

        # 添加覆盖圆半径系数的说明帮助
        self.m_help_label = ttk.Label(param_frame, text="?", foreground="blue", cursor="hand2")
        self.m_help_label.grid(row=0, column=7, sticky=tk.W, padx=(0, 5))
        self.m_tooltip = Tooltip(self.m_help_label, "覆盖圆半径系数m：覆盖半径系数\n- 默认值5/9(≈0.556)，覆盖半径为(5/9)*Co\n- m值越大，覆盖圆半径越大，可获得邻区数量增多\n- m值越小，覆盖圆半径越小，邻区数量减少")

        # Buttons frame
        btn_frame = ttk.Frame(neighbor_frame)
        btn_frame.pack(fill=tk.X, padx=10, pady=5)

        self.neighbor_plan_btn = ttk.Button(btn_frame, text="开始规划", command=self.start_neighbor_planning)
        self.neighbor_plan_btn.pack(side=tk.LEFT, padx=5)

        # Progress bar frame with percentage label
        progress_frame = ttk.Frame(btn_frame)
        progress_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

        self.neighbor_progress = ttk.Progressbar(progress_frame, mode='determinate')
        self.neighbor_progress.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.neighbor_progress_label = ttk.Label(progress_frame, text="0%")
        self.neighbor_progress_label.pack(side=tk.RIGHT, padx=(5,0))

        # Results text area
        result_frame = ttk.LabelFrame(neighbor_frame, text="规划结果")
        result_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        self.neighbor_result_text = scrolledtext.ScrolledText(result_frame, height=15)
        self.neighbor_result_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        # Configure tags for text coloring
        self.neighbor_result_text.tag_config("error", foreground="red")
        self.neighbor_result_text.tag_config("warning", foreground="orange")
        self.neighbor_result_text.tag_config("info", foreground="black")

    def create_status_bar(self):
        """Create the status bar"""

    def select_cells_file(self):
        """Select the cells file"""
        import glob
        # First try to find the latest cell file automatically
        cell_files = glob.glob("待规划小区/*.xlsx")
        latest_file = self._get_latest_file(cell_files)

        # Extract directory for initialdir
        initial_dir = "待规划小区" if cell_files else "."
        initial_file = os.path.basename(latest_file) if latest_file else ""

        filename = filedialog.askopenfilename(
            title="选择待规划小区文件",
            initialdir=initial_dir,
            initialfile=initial_file,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.cells_file.set(filename)

    def select_params_file(self):
        """Select the parameters file"""
        import glob
        # First try to find the latest parameter file automatically
        param_files = glob.glob("全量工参/ProjectParameter_mongoose*.xlsx")
        latest_file = self._get_latest_file(param_files)

        # Extract directory for initialdir
        initial_dir = "全量工参" if param_files else "."
        initial_file = os.path.basename(latest_file) if latest_file else ""

        filename = filedialog.askopenfilename(
            title="选择全量工参文件",
            initialdir=initial_dir,
            initialfile=initial_file,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.params_file.set(filename)

    def select_param_update_file(self):
        """Select the parameters file for update, default to latest timestamped file"""
        import glob
        # First try to find the latest parameter file automatically
        param_files = glob.glob("全量工参/ProjectParameter_mongoose*.xlsx")
        latest_file = self._get_latest_file(param_files)

        # Extract directory for initialdir
        initial_dir = "全量工参" if param_files else "."
        initial_file = os.path.basename(latest_file) if latest_file else ""

        filename = filedialog.askopenfilename(
            title="选择全量工参文件",
            initialdir=initial_dir,
            initialfile=initial_file,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.param_update_file.set(filename)

    def select_online_param_file(self):
        """Select the online parameters file, default to latest BaselineLab file"""
        import glob
        # First try to find the latest BaselineLab file automatically
        baseline_files = glob.glob("全量工参/BaselineLab_*.zip")
        latest_file = self._get_latest_file(baseline_files)

        # Extract directory for initialdir
        initial_dir = "全量工参" if baseline_files else "."
        initial_file = os.path.basename(latest_file) if latest_file else ""

        filename = filedialog.askopenfilename(
            title="选择现网工参文件",
            initialdir=initial_dir,
            initialfile=initial_file,
            filetypes=[("ZIP files", "*.zip"), ("All files", "*.*")]
        )
        if filename:
            self.online_param_file.set(filename)

    

    def select_neighbor_cells_file(self):
        """Select the cells file for neighbor planning"""
        import glob
        # First try to find the latest cell file automatically
        cell_files = glob.glob("待规划小区/*.xlsx")
        latest_file = self._get_latest_file(cell_files)

        # Extract directory for initialdir
        initial_dir = "待规划小区" if cell_files else "."
        initial_file = os.path.basename(latest_file) if latest_file else ""

        filename = filedialog.askopenfilename(
            title="选择待规划小区文件",
            initialdir=initial_dir,
            initialfile=initial_file,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.neighbor_cells_file.set(filename)

    def select_neighbor_params_file(self):
        """Select the parameters file for neighbor planning"""
        import glob
        # First try to find the latest parameter file automatically
        param_files = glob.glob("全量工参/ProjectParameter_mongoose*.xlsx")
        latest_file = self._get_latest_file(param_files)

        # Extract directory for initialdir
        initial_dir = "全量工参" if param_files else "."
        initial_file = os.path.basename(latest_file) if latest_file else ""

        filename = filedialog.askopenfilename(
            title="选择全量工参文件",
            initialdir=initial_dir,
            initialfile=initial_file,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.neighbor_params_file.set(filename)


    def _get_latest_file(self, file_list):
        """Get the latest file from a list based on timestamp in filename or modification time"""
        if not file_list:
            return None

        import re
        from datetime import datetime

        # Extract timestamps from filenames and sort by them
        file_timestamps = []
        for file_path in file_list:
            file_name = os.path.basename(file_path)
            # Look for 14-digit timestamp in filename (YYYYMMDDHHMMSS)
            timestamp_match = re.search(r'(\d{14})', file_name)
            if timestamp_match:
                timestamp_str = timestamp_match.group(1)
                try:
                    timestamp = datetime.strptime(timestamp_str, '%Y%m%d%H%M%S')
                    file_timestamps.append((timestamp, file_path))
                except ValueError:
                    # If not 14-digit timestamp, try 8-digit date (YYYYMMDD)
                    date_match = re.search(r'(\d{8})', file_name)
                    if date_match:
                        date_str = date_match.group(1)
                        try:
                            date = datetime.strptime(date_str, '%Y%m%d')
                            file_timestamps.append((date, file_path))
                        except ValueError:
                            # If no valid timestamp in filename, use modification time
                            file_timestamps.append((datetime.fromtimestamp(os.path.getmtime(file_path)), file_path))
            else:
                # If no timestamp in filename, use modification time
                file_timestamps.append((datetime.fromtimestamp(os.path.getmtime(file_path)), file_path))

        # Sort by timestamp in descending order and return the latest file
        if file_timestamps:
            file_timestamps.sort(key=lambda x: x[0], reverse=True)
            return file_timestamps[0][1]

        return file_list[0] if file_list else None

    def show_about(self):
        """Show about dialog"""
        messagebox.showinfo("About", "NetworkPlanningTool_V1\n邮箱：weicongpeng1@163.com\n微信：pengwc2010")

    def open_output_folder(self):
        """Open the output folder in file explorer"""
        import subprocess
        import sys
        output_dir = "输出文件"
        os.makedirs(output_dir, exist_ok=True)  # Ensure directory exists

        try:
            if sys.platform == "win32":
                subprocess.run(["explorer", output_dir], check=True)
            elif sys.platform == "darwin":  # macOS
                subprocess.run(["open", output_dir], check=True)
            else:  # Linux and other Unix-like
                subprocess.run(["xdg-open", output_dir], check=True)
        except (subprocess.CalledProcessError, FileNotFoundError):
            # Fallback: just show a message
            messagebox.showinfo("打开输出文件夹", f"输出文件夹路径: {os.path.abspath(output_dir)}")

    def start_pci_planning(self):
        """Start the PCI planning process in a separate thread"""
        # Validate inputs
        if not self.neighbor_cells_file.get() or not self.param_update_file.get():
            messagebox.showerror("错误", "请先在“规划数据导入及工参更新”标签页中选择待规划小区文件和全量工参文件")
            return

        # Disable the button during processing
        self.plan_btn.config(state=tk.DISABLED)
        self.plan_progress['value'] = 0
        self.plan_progress_label.config(text="0%")

        # Start the planning in a separate thread
        planning_thread = threading.Thread(target=self.run_pci_planning)
        planning_thread.daemon = True
        planning_thread.start()

    def run_pci_planning(self):
        """Run the actual PCI planning process"""
        planning_successful = False  # Flag to track if planning completed successfully

        try:
            # Log start time
            start_time = time.time()

            # Log initial message
            self.queue.put(("info", "开始PCI规划...\n"))

            # Create planner instance
            network_types = []
            if self.network_type.get() == "LTE":
                network_types.append("LTE")
            elif self.network_type.get() == "NR":
                network_types.append("NR")
            else:  # For "Both" option
                network_types = ["LTE", "NR"]

            # Calculate total progress steps
            total_networks = len(network_types)
            if total_networks > 0:
                progress_per_network = 100 // total_networks
            else:
                progress_per_network = 100

            for idx, network_type in enumerate(network_types):
                # Report progress for each network
                self.queue.put(("info", f"开始规划 {network_type} 网络\n"))
                self.queue.put(("progress", idx * progress_per_network))

                # Create planner
                pci_range = None
                if self.pci_range_enabled.get() and self.pci_range_var.get():
                    # Parse the PCI range
                    try:
                        range_str = self.pci_range_var.get().strip()
                        if '-' in range_str:
                            start_pci, end_pci = map(int, range_str.split('-'))
                            if 0 <= start_pci <= end_pci:
                                # Determine valid range based on network type
                                max_pci = 503 if network_type == "LTE" else 1007
                                if start_pci > max_pci:
                                    self.queue.put(("error", f"错误: 起始PCI值 ({start_pci}) 超出{network_type}网络范围 (最大: {max_pci})\n"))
                                    return  # Stop planning
                                if end_pci > max_pci:
                                    self.queue.put(("error", f"错误: 结束PCI值 ({end_pci}) 超出{network_type}网络范围 (最大: {max_pci})\n"))
                                    return  # Stop planning
                                pci_range = list(range(start_pci, end_pci + 1))
                            else:
                                self.queue.put(("error", f"错误: 起始PCI值应小于等于结束PCI值\n"))
                                return  # Stop planning
                        else:
                            self.queue.put(("error", f"错误: 请使用格式 '起始PCI-结束PCI' (例如: 0-100)\n"))
                            return  # Stop planning
                    except ValueError:
                        self.queue.put(("error", f"错误: PCI范围格式无效，请使用数字和'-' (例如: 0-100)\n"))
                        return  # Stop planning

                planner = LTENRPCIPlanner(
                    reuse_distance_km=self.reuse_distance.get(),
                    lte_inherit_mod3=self.lte_inherit_mod3.get() if network_type == "LTE" else False,
                    nr_inherit_mod30=self.nr_inherit_mod30.get() if network_type == "NR" else False,
                    network_type=network_type,
                    params_file=self.param_update_file.get(),
                    pci_range=pci_range
                )

                # Load data
                planner.load_data(self.neighbor_cells_file.get(), self.param_update_file.get())

                # Report progress after loading data
                progress_after_load = idx * progress_per_network + progress_per_network // 3
                self.queue.put(("progress", progress_after_load))

                # Perform planning
                result_df = planner.plan_pci_with_reuse_priority()

                # Report progress after planning
                progress_after_planning = idx * progress_per_network + 2 * progress_per_network // 3
                self.queue.put(("progress", progress_after_planning))

                # Generate output file name
                timestamp = planner.generate_timestamp_suffix()
                mod_suffix = "mod3" if (network_type == "LTE" and self.lte_inherit_mod3.get()) else \
                             "mod30" if (network_type == "NR" and self.nr_inherit_mod30.get()) else \
                             "nomod" + network_type.lower()

                output_dir = "输出文件"
                os.makedirs(output_dir, exist_ok=True)
                output_file = f"{output_dir}/pci_planning_{network_type.lower()}_{self.reuse_distance.get()}km_{mod_suffix}_{timestamp}.xlsx"

                # Save result
                result_df.to_excel(output_file, index=False)

                self.queue.put(("info", f"{network_type}规划完成，结果保存至: {output_file}\n"))

        except Exception as e:
            # Log error
            error_msg = f"PCI规划过程中发生错误: {str(e)}\n"
            self.queue.put(("error", error_msg))
            self.queue.put(("info", f"\n❌ PCI规划失败\n"))
            print(error_msg)
            import traceback
            traceback.print_exc()
        finally:
            # Re-enable the button
            self.queue.put(("progress", 100))
            self.queue.put(("enable_button", "plan_btn"))
            planning_successful = True

            # Calculate and log execution time
            end_time = time.time()
            execution_time = end_time - start_time
            completion_msg = f"PCI规划完成，总用时: {execution_time:.2f}秒\n"
            self.queue.put(("info", completion_msg))

            # Add success indicator
            self.queue.put(("info", f"\n✅ PCI规划完成\n"))

            # Prepare output directory info (but don't show popup)
            output_dir = os.path.abspath("输出文件")

            # Open output folder after planning completion
            if planning_successful:
                self.root.after(1000, self.open_output_folder)  # Delay to let GUI update


    def start_param_update(self):
        """Start the parameter update process in a separate thread"""
        if not self.param_update_file.get() or not self.online_param_file.get():
            messagebox.showerror("错误", "请先选择全量工参文件和现网工参文件")
            return

        # Disable the button during processing
        self.update_btn.config(state=tk.DISABLED)

        # Start the update in a separate thread
        update_thread = threading.Thread(target=self.run_param_update)
        update_thread.daemon = True
        update_thread.start()

    def run_param_update(self):
        """Run the actual parameter update process"""
        import sys
        import io
        from contextlib import redirect_stdout, redirect_stderr

        update_successful = False  # Flag to track if update completed successfully
        output_text = ""  # Store output for popup

        try:
            output_text += "开始现网工参更新...\n"

            # Create updater instance and set file paths
            updater = NetworkParameterUpdater()

            # Override the default file paths with user selections
            if self.param_update_file.get():
                updater.full_param_dir = os.path.dirname(self.param_update_file.get())

            if self.online_param_file.get():
                # Since we're using the NetworkParameterUpdater class as is,
                # we need to extract the zip path from the selected file
                updater.baseline_zip_pattern = os.path.basename(self.online_param_file.get())

            # Redirect print output to capture it in the GUI
            result_string = io.StringIO()
            with redirect_stdout(result_string), redirect_stderr(result_string):
                success = updater.update_network_parameters()

            # Get and display the captured output
            captured_output = result_string.getvalue()
            output_text += captured_output

            if success:
                output_text += "\n✅ 现网工参更新完成\n"
                update_successful = True  # Mark as successful
            else:
                output_text += "\n❌ 现网工参更新失败\n"

        except Exception as e:
            output_text += f"更新过程中发生错误: {str(e)}\n"
            import traceback
            output_text += traceback.format_exc()

        finally:
            # Show results in popup
            if update_successful:
                messagebox.showinfo("更新结果", output_text)
            else:
                messagebox.showerror("更新结果", output_text)

            # 现网工参更新后不再自动打开输出文件夹
            # if update_successful:
            #     self.root.after(1000, self.open_output_folder)  # Delay to let GUI update

    def start_neighbor_planning(self):
        """Start the neighbor planning process in a separate thread"""
        if not self.neighbor_cells_file.get() or not self.param_update_file.get():
            messagebox.showerror("错误", "请先在“规划数据导入及工参更新”标签页中选择待规划小区文件和全量工参文件")
            return

        # Disable the button during processing
        self.neighbor_plan_btn.config(state=tk.DISABLED)
        self.neighbor_progress['value'] = 0
        self.neighbor_progress_label.config(text="0%")

        # Start the planning in a separate thread
        neighbor_thread = threading.Thread(target=self.run_neighbor_planning)
        neighbor_thread.daemon = True
        neighbor_thread.start()

    def run_neighbor_planning(self):
        """Run the actual neighbor planning process"""
        import sys
        import io
        from contextlib import redirect_stdout, redirect_stderr

        neighbor_successful = False  # Flag to track if neighbor planning completed successfully

        try:
            self.queue.put(("neighbor_info", "开始邻区规划...\n"))
            self.queue.put(("neighbor_progress", 10))  # Start progress

            # Create neighbor planning tool instance (always uses coverage circle algorithm now)
            tool = NeighborPlanningTool(
                max_neighbors=self.max_neighbors.get(),
                coverage_distance_factor=self.coverage_distance_factor.get(),
                coverage_radius_factor=self.coverage_radius_factor.get()
            )

            # Determine planning type based on GUI selection
            planning_type = self.planning_type.get()

            # Redirect print output to capture it in the GUI
            result_string = io.StringIO()
            with redirect_stdout(result_string), redirect_stderr(result_string):
                success = tool.run_neighbor_planning(
                    self.neighbor_cells_file.get(),
                    self.param_update_file.get(),
                    planning_type
                )

            # Get and display the captured output
            output = result_string.getvalue()
            self.queue.put(("neighbor_info", output))

            if success:
                self.queue.put(("neighbor_info", f"\n✅ {planning_type}邻区规划完成\n"))
                self.queue.put(("neighbor_progress", 100))
                neighbor_successful = True  # Mark as successful
            else:
                self.queue.put(("neighbor_info", f"\n❌ {planning_type}邻区规划失败\n"))
                self.queue.put(("neighbor_progress", 100))

        except Exception as e:
            self.queue.put(("neighbor_error", f"规划过程中发生错误: {str(e)}\n"))
            import traceback
            self.queue.put(("neighbor_error", traceback.format_exc()))
        finally:
            self.queue.put(("neighbor_done", None))

            # Only open output folder if neighbor planning completed successfully
            if neighbor_successful:
                self.root.after(1000, self.open_output_folder)  # Delay to let GUI update

    def download_template(self):
        """Download the planning cell template file"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            import os

            # Create a new Excel workbook
            wb = Workbook()

            # Create LTE sheet with headers
            lte_ws = wb.active
            lte_ws.title = "LTE"
            lte_headers = ["MCC", "MNC", "eNodeBID", "CellID"]
            for col, header in enumerate(lte_headers, 1):
                lte_ws.cell(row=1, column=col, value=header)

            # Create NR sheet with headers
            nr_ws = wb.create_sheet(title="NR")
            nr_headers = ["MCC", "MNC", "gNodeBID", "CellID"]
            for col, header in enumerate(nr_headers, 1):
                nr_ws.cell(row=1, column=col, value=header)

            # Define the output file path
            output_dir = "输出文件"
            os.makedirs(output_dir, exist_ok=True)
            template_file = os.path.join(output_dir, "cell-tree-export-规划小区模版.xlsx")

            # Save the workbook
            wb.save(template_file)

            # Show success message
            messagebox.showinfo("模板下载", f"规划小区模板已成功下载至:\n{template_file}")

        except Exception as e:
            messagebox.showerror("模板下载错误", f"下载模板时发生错误: {str(e)}")

        """
        通过关键词模糊查找列名，支持部分匹配
        例如：列名为"子网ID SubNetwork ID long[0...128]"，只用"子网ID"即可匹配

        Args:
            df: 数据框
            keyword: 关键词

        Returns:
            匹配的列名，如果找不到返回None
        """
        # 首先尝试精确匹配
        if keyword in df.columns:
            return keyword

        # 然后尝试部分匹配 - 检查关键词是否包含在列名中
        for col in df.columns:
            if keyword in col:
                return col

        # 最后尝试前缀匹配作为备选
        return self._find_column_by_prefix(df, keyword)

    def _find_column_by_prefix(self, df: pd.DataFrame, prefix: str) -> str:
        """
        通过前缀模糊查找列名

        Args:
            df: 数据框
            prefix: 列名前缀

        Returns:
            匹配的列名，如果找不到返回None
        """
        for col in df.columns:
            if col.startswith(prefix):
                return col
        return None

    def process_queue(self):
        """Process messages from the thread queue"""
        while True:
            try:
                msg_type, content = self.queue.get_nowait()

                if msg_type == "info":
                    # For sector layer generation, show in the appropriate result area based on active tab
                    if hasattr(self, 'neighbor_result_text') and self.notebook.select() == self.notebook.tabs()[1]:
                        # If neighbor planning tab is active
                        self.neighbor_result_text.insert(tk.END, content)
                        self.neighbor_result_text.see(tk.END)
                    elif hasattr(self, 'result_text'):
                        # Otherwise use PCI planning result text
                        self.result_text.insert(tk.END, content)
                        self.result_text.see(tk.END)
                elif msg_type == "progress":
                    self.plan_progress['value'] = content
                    self.plan_progress_label.config(text=f"{content}%")
                elif msg_type == "error":
                    # Insert error text in red
                    if hasattr(self, 'neighbor_result_text') and self.notebook.select() == self.notebook.tabs()[1]:
                        # If neighbor planning tab is active
                        self.neighbor_result_text.insert(tk.END, content, "error")
                        self.neighbor_result_text.see(tk.END)
                    elif hasattr(self, 'result_text'):
                        # Otherwise use PCI planning result text
                        self.result_text.insert(tk.END, content, "error")
                        self.result_text.see(tk.END)
                elif msg_type == "done":
                    self.plan_btn.config(state=tk.NORMAL)

                # Update parameter update results - no longer used since we use popups
                elif msg_type == "update_info":
                    pass  # No longer insert to text area
                elif msg_type == "update_progress":
                    pass  # No longer used
                elif msg_type == "update_error":
                    # No longer insert to text area
                elif msg_type == "update_done":
                    self.update_btn.config(state=tk.NORMAL)

                # Handle button enabling for PCI planning
                elif msg_type == "enable_button":
                    if content == "plan_btn":
                        self.plan_btn.config(state=tk.NORMAL)
                    elif content == "update_btn":
                        self.update_btn.config(state=tk.NORMAL)


                # Update neighbor planning results
                elif msg_type == "neighbor_info":
                    self.neighbor_result_text.insert(tk.END, content)
                    self.neighbor_result_text.see(tk.END)
                elif msg_type == "neighbor_progress":
                    self.neighbor_progress['value'] = content
                    self.neighbor_progress_label.config(text=f"{content}%")
                elif msg_type == "neighbor_error":
                    # Insert error text in red
                    self.neighbor_result_text.insert(tk.END, content, "error")
                    self.neighbor_result_text.see(tk.END)
                elif msg_type == "neighbor_done":
                    self.neighbor_plan_btn.config(state=tk.NORMAL)

            except queue.Empty:
                break

        # Schedule the next check
        self.root.after(100, self.process_queue)


def main():
    """
    主函数 - LTE/NR分离式PCI规划工具 (GUI版)
    """
    # 启动GUI界面
    root = tk.Tk()
    app = PCIGUIApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
