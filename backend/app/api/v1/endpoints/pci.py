"""
PCI规划API端点
"""

import os
import uuid
import shutil
import re
from datetime import datetime
from fastapi import APIRouter, HTTPException, BackgroundTasks, Request
from pydantic import ValidationError
from typing import Dict, Any, Optional
from app.models.schemas import PCIConfig, PCIResult, NetworkType
from app.services.task_manager import task_manager
from app.services.data_service import data_service
from app.core.config import settings

router = APIRouter()


@router.post("/plan", response_model=Dict[str, Any])
async def start_pci_planning(
    request: Request, background_tasks: BackgroundTasks
) -> Dict[str, Any]:
    """启动PCI规划任务 - 增强参数验证"""
    try:
        # 手动解析和验证以提供更友好的错误消息
        body = await request.json()

        try:
            config = PCIConfig(**body)
        except ValidationError as e:
            # 提取验证错误并生成友好的错误消息
            errors = []
            for error in e.errors():
                loc = " -> ".join(str(l) for l in error["loc"])
                msg = error["msg"]
                errors.append(f"{loc}: {msg}")

            error_detail = "; ".join(errors)
            raise HTTPException(
                status_code=400,
                detail={
                    "message": "参数验证失败",
                    "errors": error_detail
                }
            )

        # 创建任务
        task_id = await task_manager.create_pci_task(config)
        return {
            "success": True,
            "data": {"taskId": task_id, "message": "PCI规划任务已启动"},
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))





@router.get("/progress/{task_id}", response_model=Dict[str, Any])
async def get_pci_progress(task_id: str) -> Dict[str, Any]:
    """获取PCI规划进度"""
    try:
        progress = task_manager.get_task_progress(task_id)
        if progress is None:
            raise HTTPException(status_code=404, detail="任务不存在")
        return {"success": True, "data": progress}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/result/{task_id}", response_model=Dict[str, Any])
async def get_pci_result(task_id: str) -> Dict[str, Any]:
    """获取PCI规划结果"""
    try:
        result = task_manager.get_task_result(task_id)
        if result is None:
            raise HTTPException(status_code=404, detail="任务不存在")
        return {"success": True, "data": result}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export/{task_id}")
async def export_pci_result(task_id: str, format: str = "xlsx"):
    """导出PCI规划结果"""
    try:
        from fastapi.responses import FileResponse

        file_path = task_manager.export_result(task_id, format)
        if file_path is None:
            raise HTTPException(status_code=404, detail="任务不存在或导出失败")

        filename = f"pci_result_{task_id}.{format}"
        return FileResponse(
            path=file_path, filename=filename, media_type="application/octet-stream"
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/task/{task_id}", response_model=Dict[str, Any])
async def cancel_pci_task(task_id: str) -> Dict[str, Any]:
    """取消PCI规划任务"""
    try:
        result = task_manager.cancel_task(task_id)
        return {"success": result, "message": "任务已取消" if result else "取消失败"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/apply-to-params/{task_id}", response_model=Dict[str, Any])
async def apply_pci_to_params(task_id: str) -> Dict[str, Any]:
    """
    将PCI规划结果应用到全量工参
    
    流程：
    1. 获取规划结果和当前全量工参信息
    2. 创建新的数据目录和ID
    3. 复制原始Excel文件到新目录
    4. 使用openpyxl更新新Excel文件中对应网络类型sheet的PCI列
    5. 解析新Excel生成新的data.json
    6. 生成带时间戳后缀的新文件名，保存到原始目录
    7. 更新数据索引，将新工参注册为当前全量工参
    8. 返回新文件ID和更新统计
    
    注意：规划结果中的originalPCI保持不变
    """
    import json
    import logging
    import pandas as pd
    import openpyxl
    
    logger = logging.getLogger(__name__)
    logger.info(f"[应用PCI到工参] 开始处理任务: {task_id}")

    # 获取任务结果
    result = task_manager.get_task_result(task_id)
    if result is None:
        raise HTTPException(status_code=404, detail="任务不存在")

    if result.get("status") != "completed":
        raise HTTPException(status_code=400, detail="任务尚未完成")

    # 获取规划结果
    results = result.get("results", [])
    if not results:
        raise HTTPException(status_code=400, detail="规划结果为空")
    
    # 获取网络类型
    network_type = result.get("networkType", "LTE")
    logger.info(f"[应用PCI到工参] 规划结果包含 {len(results)} 个站点, 网络类型: {network_type}")

    # 找到当前的全量工参
    full_params_id = None
    full_params_info = None
    for data_id, data_info in data_service.index.items():
        if data_info.get("fileType") == "full_params":
            full_params_id = data_id
            full_params_info = data_info
            break

    if not full_params_id:
        raise HTTPException(status_code=404, detail="未找到全量工参数据")

    logger.info(f"[应用PCI到工参] 找到全量工参: {full_params_id}")

    # 获取原始Excel文件路径
    original_excel_path = settings.DATA_DIR / full_params_id / "original.xlsx"
    if not original_excel_path.exists():
        raise HTTPException(status_code=404, detail="全量工参原始Excel文件不存在")

    # 在创建新数据之前，删除所有旧的备份工参（清理旧数据）
    deleted_backup_count = 0
    for old_id, old_info in list(data_service.index.items()):
        if old_info.get("fileType") == "full_params_backup":
            try:
                data_service.delete_data(old_id)
                deleted_backup_count += 1
                logger.info(f"[应用PCI到工参] 已删除旧备份工参: {old_id}")
            except Exception as e:
                logger.warning(f"[应用PCI到工参] 删除旧备份工参失败: {old_id}, 错误: {e}")
    
    if deleted_backup_count > 0:
        logger.info(f"[应用PCI到工参] 共删除 {deleted_backup_count} 个旧备份工参")

    # 创建新的数据ID和目录
    new_data_id = str(uuid.uuid4())
    new_data_dir = settings.DATA_DIR / new_data_id
    new_data_dir.mkdir(parents=True, exist_ok=True)
    
    # 复制原始Excel到新目录
    new_excel_path = new_data_dir / "original.xlsx"
    shutil.copy2(original_excel_path, new_excel_path)
    logger.info(f"[应用PCI到工参] 已复制原始Excel到新目录: {new_excel_path}")

    # 使用openpyxl打开新Excel文件并更新PCI
    try:
        wb = openpyxl.load_workbook(new_excel_path)
    except Exception as e:
        logger.error(f"[应用PCI到工参] 无法打开Excel文件: {e}")
        raise HTTPException(status_code=500, detail=f"无法打开Excel文件: {str(e)}")

    # 确定要更新的sheet名称
    target_sheet_name = f"{network_type} Project Parameters"
    if target_sheet_name not in wb.sheetnames:
        logger.warning(f"[应用PCI到工参] 未找到工作表 {target_sheet_name}，尝试查找匹配的sheet")
        for sheet_name in wb.sheetnames:
            if network_type in sheet_name and "Project" in sheet_name:
                target_sheet_name = sheet_name
                break
    
    if target_sheet_name not in wb.sheetnames:
        logger.error(f"[应用PCI到工参] 未找到工作表: {target_sheet_name}")
        raise HTTPException(status_code=404, detail=f"未找到工作表: {target_sheet_name}")

    ws = wb[target_sheet_name]
    logger.info(f"[应用PCI到工参] 开始处理工作表: {target_sheet_name}")

    # 获取表头行
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value) if cell.value else "")
    
    logger.info(f"[应用PCI到工参] 表头: {headers[:15]}...")

    # 查找关键列索引
    site_id_col_idx = None
    sector_id_col_idx = None
    pci_col_idx = None

    # 基站ID列关键词
    if network_type == "LTE":
        site_id_keywords = ["eNodeB标识", "eNodeB ID", "基站ID", "eNodeB"]
    else:
        site_id_keywords = ["gNodeB标识", "gNodeB ID", "基站ID", "gNodeB"]
    
    sector_id_keywords = ["小区标识", "cellLocalId", "SectorID", "小区ID", "LocalCellId"]
    pci_keywords = ["PCI", "物理小区识别码"]

    # 匹配基站ID列
    for keyword in site_id_keywords:
        for idx, header in enumerate(headers):
            header_clean = header.strip().replace("\n", " ")
            if keyword in header_clean and "Length" not in header_clean and "宽度" not in header_clean:
                site_id_col_idx = idx
                logger.info(f"[应用PCI到工参] 找到基站ID列: 列{idx+1} '{header}'")
                break
        if site_id_col_idx is not None:
            break

    # 匹配小区标识列
    for keyword in sector_id_keywords:
        for idx, header in enumerate(headers):
            header_clean = header.strip().replace("\n", " ")
            if keyword in header_clean and "名称" not in header_clean and "NAME" not in header_clean.upper():
                sector_id_col_idx = idx
                logger.info(f"[应用PCI到工参] 找到小区标识列: 列{idx+1} '{header}'")
                break
        if sector_id_col_idx is not None:
            break

    # 匹配PCI列
    for keyword in pci_keywords:
        for idx, header in enumerate(headers):
            header_clean = header.strip().replace("\n", " ")
            if keyword in header_clean:
                pci_col_idx = idx
                logger.info(f"[应用PCI到工参] 找到PCI列: 列{idx+1} '{header}'")
                break
        if pci_col_idx is not None:
            break

    if site_id_col_idx is None or sector_id_col_idx is None or pci_col_idx is None:
        logger.error(f"[应用PCI到工参] 缺少必要列: site_id={site_id_col_idx}, sector_id={sector_id_col_idx}, pci={pci_col_idx}")
        raise HTTPException(status_code=400, detail="Excel文件缺少必要的列（基站ID、小区标识或PCI列）")

    # 创建PCI更新映射 - 只使用精确的siteId_sectorId组合格式
    pci_updates = {}
    
    # 调试：记录规划结果中的所有ID
    logger.info(f"[应用PCI到工参] ===== 规划结果ID列表 =====")
    for site_result in results:
        result_site_id = str(site_result.get("siteId", ""))
        sectors = site_result.get("sectors", [])
        for sector in sectors:
            result_sector_id = str(sector.get("sectorId", ""))
            new_pci = sector.get("newPCI")
            original_pci = sector.get("originalPCI")
            if new_pci is not None:
                # 组合格式: siteId_sectorId（精确匹配）
                unique_key = f"{result_site_id}_{result_sector_id}"
                pci_updates[unique_key] = new_pci
                
                logger.info(f"[应用PCI到工参] 规划结果: siteId={result_site_id}, sectorId={result_sector_id}, unique_key={unique_key}, PCI {original_pci} -> {new_pci}")

    logger.info(f"[应用PCI到工参] 创建了 {len(pci_updates)} 个PCI更新映射（精确组合格式）")
    logger.info(f"[应用PCI到工参] 映射keys: {list(pci_updates.keys())}")

    # 遍历数据行更新PCI
    updated_count = 0
    updated_sectors = []
    unmatched_sectors = []  # 记录未匹配的小区
    
    # 调试：记录Excel数据中的前几行ID
    logger.info(f"[应用PCI到工参] ===== Excel数据前20行 =====")
    sample_rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        site_id_val = str(row[site_id_col_idx].value) if row[site_id_col_idx].value else ""
        sector_id_val = str(row[sector_id_col_idx].value) if row[sector_id_col_idx].value else ""
        pci_val = row[pci_col_idx].value
        
        # 处理Excel中可能的小数格式（如 0.0 -> "0"）
        original_sector_id = sector_id_val
        try:
            if "." in sector_id_val:
                sector_id_val = str(int(float(sector_id_val)))
        except:
            pass
        
        sample_rows.append({
            "row_idx": row_idx,
            "site_id": site_id_val,
            "sector_id": sector_id_val,
            "original_sector_id": original_sector_id,
            "pci": pci_val
        })
        
        if len(sample_rows) <= 20:
            logger.info(f"[应用PCI到工参] Excel行{row_idx}: siteId={site_id_val}, sectorId={sector_id_val}(原始={original_sector_id}), PCI={pci_val}")
    
    logger.info(f"[应用PCI到工参] Excel共 {len(sample_rows)} 行数据")
    
    # 重新遍历进行更新
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        site_id_val = str(row[site_id_col_idx].value) if row[site_id_col_idx].value else ""
        sector_id_val = str(row[sector_id_col_idx].value) if row[sector_id_col_idx].value else ""
        
        # 处理Excel中可能的小数格式（如 0.0 -> "0"）
        try:
            if "." in sector_id_val:
                sector_id_val = str(int(float(sector_id_val)))
        except:
            pass
        
        matched_pci = None
        match_method = ""
        
        # 方式1: 组合格式匹配 siteId_sectorId（精确匹配）
        unique_key = f"{site_id_val}_{sector_id_val}"
        if unique_key in pci_updates:
            matched_pci = pci_updates[unique_key]
            match_method = "组合格式精确匹配"
        
        # 方式2: 如果sector_id_val已经是组合格式（如"540951_0"）
        if matched_pci is None and "_" in sector_id_val:
            parts = sector_id_val.split("_")
            if len(parts) == 2:
                excel_site_id = parts[0]
                excel_sector_id = parts[1]
                # 尝试用解析后的ID匹配
                key1 = f"{excel_site_id}_{excel_sector_id}"
                if key1 in pci_updates:
                    matched_pci = pci_updates[key1]
                    match_method = "解析组合格式精确匹配"

        # 注意：移除了"纯sector_id唯一匹配"逻辑，因为会导致错误匹配
        # 只使用精确的site_id + sector_id组合匹配

        if matched_pci is not None:
            old_pci = row[pci_col_idx].value
            if old_pci != matched_pci:
                row[pci_col_idx].value = matched_pci
                updated_count += 1
                updated_sectors.append({
                    "siteId": site_id_val,
                    "sectorId": sector_id_val,
                    "oldPCI": old_pci,
                    "newPCI": matched_pci,
                    "matchMethod": match_method
                })
                logger.info(f"[应用PCI到工参] 更新行{row_idx}({match_method}): 基站={site_id_val}, 小区={sector_id_val}, PCI {old_pci} -> {matched_pci}")
            else:
                logger.debug(f"[应用PCI到工参] 跳过行{row_idx}: PCI未变化 ({old_pci})")
        else:
            # 记录未匹配的小区（仅记录有PCI值的小区）
            if row[pci_col_idx].value is not None:
                unmatched_sectors.append({
                    "siteId": site_id_val,
                    "sectorId": sector_id_val,
                    "currentPCI": row[pci_col_idx].value
                })
    
    # 输出未匹配的小区信息
    if unmatched_sectors:
        logger.warning(f"[应用PCI到工参] 有 {len(unmatched_sectors)} 个小区未找到匹配的规划结果")
        for us in unmatched_sectors[:10]:  # 只输出前10个
            logger.warning(f"[应用PCI到工参]   未匹配: 基站={us['siteId']}, 小区={us['sectorId']}, 当前PCI={us['currentPCI']}")

    # 保存更新后的Excel
    wb.save(new_excel_path)
    wb.close()
    logger.info(f"[应用PCI到工参] 已保存更新后的Excel文件: {new_excel_path}")
    logger.info(f"[应用PCI到工参] 共更新 {updated_count} 个小区的PCI")

    # 解析新Excel生成data.json
    try:
        parsed_data = {}
        xls = pd.ExcelFile(new_excel_path)
        
        for net in ["LTE", "NR"]:
            sheet_name = f"{net} Project Parameters"
            if sheet_name in xls.sheet_names:
                sites = data_service._parse_sheet_data(xls, sheet_name, net)
                parsed_data[net] = sites
                logger.info(f"[应用PCI到工参] 解析 {net} 工参: {len(sites)} 个站点")
        
        # 保存data.json
        with open(new_data_dir / "data.json", "w", encoding="utf-8") as f:
            json.dump(parsed_data, f, ensure_ascii=False, indent=2)
        logger.info(f"[应用PCI到工参] 已保存data.json")
        
    except Exception as e:
        logger.error(f"[应用PCI到工参] 解析新Excel失败: {e}")
        import traceback
        logger.error(traceback.format_exc())

    # 生成带时间戳和"已更新PCI"后缀的新文件名
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    original_name = full_params_info.get("name", "ProjectParameter_mongoose.xlsx")
    name_part, ext_part = os.path.splitext(original_name)
    # 移除可能已有的时间戳和"已更新PCI"后缀
    new_name = re.sub(r"_已更新PCI$", "", name_part)  # 移除已有的"已更新PCI"后缀
    new_name = re.sub(r"_?\d{14}$", "", new_name)  # 移除已有的时间戳
    new_filename = f"{new_name}_{timestamp}_已更新PCI{ext_part}"
    
    logger.info(f"[应用PCI到工参] 新文件名: {new_filename}")

    # 保存到原始目录（如果存在originalPath）
    saved_to_original = False
    original_path = full_params_info.get("originalPath")
    
    if original_path and original_path.strip():
        try:
            original_file_path = os.path.abspath(original_path.strip())
            original_dir = os.path.dirname(original_file_path)
            
            if os.path.exists(original_dir):
                target_path = os.path.join(original_dir, new_filename)
                shutil.copy2(new_excel_path, target_path)
                
                if os.path.exists(target_path):
                    saved_to_original = True
                    logger.info(f"[应用PCI到工参] 已保存副本到原始目录: {target_path}")
                else:
                    logger.warning(f"[应用PCI到工参] 保存副本失败")
            else:
                logger.warning(f"[应用PCI到工参] 原始目录不存在: {original_dir}")
        except Exception as e:
            logger.warning(f"[应用PCI到工参] 保存到原始目录失败: {e}")
    else:
        logger.info(f"[应用PCI到工参] 无原始目录路径，跳过保存副本")

    # 更新数据索引
    metadata = {}
    for net in ["LTE", "NR"]:
        if net in parsed_data:
            metadata[f"{net}SiteCount"] = len(parsed_data[net])
            metadata[f"{net}SectorCount"] = sum(len(s.get("sectors", [])) for s in parsed_data[net])

    data_service.index[new_data_id] = {
        "id": new_data_id,
        "name": new_filename,
        "type": "excel",
        "fileType": "full_params",
        "size": new_excel_path.stat().st_size,
        "uploadDate": datetime.now().isoformat(),
        "status": "ready",
        "originalPath": original_path,
        "metadata": metadata,
    }
    
    # 删除旧的全量工参（而不是标记为备份）
    if full_params_id in data_service.index:
        try:
            data_service.delete_data(full_params_id)
            logger.info(f"[应用PCI到工参] 已删除旧工参: {full_params_id}")
        except Exception as e:
            logger.warning(f"[应用PCI到工参] 删除旧工参失败: {full_params_id}, 错误: {e}")
    
    data_service._save_index()
    logger.info(f"[应用PCI到工参] 索引已更新，新工参ID: {new_data_id}")

    return {
        "success": True,
        "data": {
            "message": f"成功应用PCI规划结果到全量工参，更新了{updated_count}个小区的PCI",
            "updatedCount": updated_count,
            "newFileId": new_data_id,
            "newFileName": new_filename,
            "savedToOriginal": saved_to_original,
            "deletedBackupCount": deleted_backup_count
        }
    }
