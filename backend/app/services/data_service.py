"""
数据管理服务
"""

import os
import sys
import uuid
import shutil
import glob
import re
from datetime import datetime
from typing import List, Optional, Dict, Any, Set, Tuple
from pathlib import Path
import pandas as pd
import numpy as np
import openpyxl

from app.core.config import settings
from app.models.schemas import DataItem, DataType, DataStatus, SiteData

# 设置stdout编码为utf-8，避免Windows下GBK编码问题
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except:
        pass  # 如果无法配置，忽略错误


def safe_print(msg: str):
    """安全打印函数，处理可能的编码错误"""
    try:
        print(msg)
    except UnicodeEncodeError:
        # 如果打印失败，尝试使用GBK编码，替换无法编码的字符
        safe_msg = msg.encode("gbk", "replace").decode("gbk")
        print(safe_msg)
    except Exception:
        # 其他错误也尝试处理
        try:
            safe_msg = msg.encode("gbk", "replace").decode("gbk")
            print(safe_msg)
        except:
            pass  # 如果还是失败，放弃打印


class DataService:
    """数据管理服务类"""

    def __init__(self, lazy_load: bool = True):
        self.data_index_file = settings.DATA_DIR / "index.json"
        self._ensure_directories()
        self._load_index()
        # 数据缓存: {data_id: {"data": data_dict, "mtime": timestamp}}
        self._data_cache: Dict[str, Dict[str, Any]] = {}
        # 索引缓存时间戳
        self._index_mtime: Optional[float] = None
        # 扫描uploads目录 - 默认延迟加载以加快启动
        if not lazy_load:
            self._scan_uploads_directory()
        else:
            # 异步扫描，不阻塞启动
            import threading
            threading.Thread(target=self._scan_uploads_directory, daemon=True).start()

    def _ensure_directories(self):
        """确保必要的目录存在"""
        settings.UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        settings.DATA_DIR.mkdir(parents=True, exist_ok=True)

    def _load_index(self):
        """加载数据索引"""
        if self.data_index_file.exists():
            import json

            with open(self.data_index_file, "r", encoding="utf-8") as f:
                self.index = json.load(f)
        else:
            self.index = {}

    def reload_index(self):
        """重新加载数据索引（用于其他模块获取最新索引）"""
        self._load_index()
        safe_print(f"[DataService] 索引已重新加载，共 {len(self.index)} 个数据项")

    def _save_index(self):
        """保存数据索引"""
        import json

        self.data_index_file.parent.mkdir(parents=True, exist_ok=True)
        with open(self.data_index_file, "w", encoding="utf-8") as f:
            json.dump(self.index, f, ensure_ascii=False, indent=2)

    def _scan_uploads_directory(self):
        """扫描uploads目录中的文件并添加到index中"""
        safe_print(f"[DataService] 开始扫描uploads目录...")

        # 遍历uploads目录中的所有.xlsx文件
        for file_path in settings.UPLOAD_DIR.glob("*.xlsx"):
            # 从文件名中提取data_id
            filename = file_path.stem
            if "-" in filename and len(filename.split("-")) == 5:
                data_id = filename

                # 检查文件是否已经在index中
                if data_id not in self.index:
                    safe_print(f"[DataService] 找到新文件: {file_path.name}")

                    # 尝试确定文件类型
                    try:
                        xls = pd.ExcelFile(file_path)
                        sheet_names = xls.sheet_names

                        file_type = "default"
                        if (
                            "LTE Project Parameters" in sheet_names
                            and "NR Project Parameters" in sheet_names
                        ):
                            file_type = "full_params"
                        elif "LTE" in sheet_names and "NR" in sheet_names:
                            file_type = "target_cells"

                        # 添加到index中
                        self.index[data_id] = {
                            "id": data_id,
                            "name": file_path.name,
                            "type": "excel",
                            "fileType": file_type,
                            "size": file_path.stat().st_size,
                            "uploadDate": datetime.now().isoformat(),
                            "status": "ready",
                            "metadata": {},
                        }

                        safe_print(
                            f"[DataService] 添加文件到索引: {data_id} ({file_type})"
                        )

                    except Exception as e:
                        safe_print(f"[DataService] 处理文件 {file_path.name} 失败: {e}")
                        continue

        # 保存索引
        self._save_index()
        safe_print(f"[DataService] 扫描完成，共 {len(self.index)} 个文件在索引中")

    def parse_points(self, file_path: str) -> List[Dict[str, Any]]:
        """从文件解析点数据"""
        path = Path(file_path)
        if not path.exists():
            raise ValueError(f"文件不存在: {file_path}")

        ext = path.suffix.lower()
        df = None

        try:
            if ext in [".xlsx", ".xls"]:
                df = pd.read_excel(path)
            elif ext == ".csv":
                try:
                    df = pd.read_csv(path, encoding="gbk")
                except:
                    df = pd.read_csv(path, encoding="utf-8")
            elif ext == ".txt":
                # 尝试各种分隔符
                success = False
                for sep in [",", "\t", ";", " "]:
                    try:
                        df = pd.read_csv(path, sep=sep, encoding="gbk")
                        if len(df.columns) > 1:
                            success = True
                            break
                    except:
                        continue
                if not success:
                    try:
                        df = pd.read_csv(path, sep=",", encoding="utf-8")
                    except:
                        raise ValueError("无法解析TXT文件")
            else:
                raise ValueError(f"不支持的文件格式: {ext}")
        except Exception as e:
            raise ValueError(f"读取文件失败: {str(e)}")

        if df is None or df.empty:
            return []

        # 智能匹配列名
        columns = df.columns.tolist()
        lat_col = None
        lng_col = None
        name_col = None

        lat_patterns = [r"lat", r"纬度", r"緯度", r"latitude"]
        lng_patterns = [r"lon", r"lng", r"经度", r"經度", r"longitude"]
        name_patterns = [r"name", r"名称", r"名稱", r"小区名", r"站名"]

        for col in columns:
            col_lower = str(col).lower()
            if not lat_col and any(re.search(p, col_lower) for p in lat_patterns):
                lat_col = col
            if not lng_col and any(re.search(p, col_lower) for p in lng_patterns):
                lng_col = col
            if not name_col and any(re.search(p, col_lower) for p in name_patterns):
                name_col = col

        if not lat_col or not lng_col:
            raise ValueError(
                "未找到经纬度列，请确保包含'经度/纬度'或'longitude/latitude'列"
            )

        points = []
        for _, row in df.iterrows():
            lat = row[lat_col]
            lng = row[lng_col]
            name = row[name_col] if name_col else "未命名点"

            try:
                lat_val = float(lat)
                lng_val = float(lng)
                if pd.isna(lat_val) or pd.isna(lng_val):
                    continue

                points.append(
                    {"name": str(name), "latitude": lat_val, "longitude": lng_val}
                )
            except:
                continue

        return points

    def update_parameters(
        self, full_param_id: str, current_param_id: str
    ) -> Dict[str, Any]:
        """更新工参"""
        safe_print(
            f"[DataService] 开始更新工参: Full={full_param_id}, Current={current_param_id}"
        )

        # 1. 获取文件路径
        if full_param_id not in self.index or current_param_id not in self.index:
            raise ValueError("找不到指定的文件ID")

        full_info = self.index[full_param_id]
        current_info = self.index[current_param_id]

        full_dir = settings.DATA_DIR / full_param_id
        current_dir = settings.DATA_DIR / current_param_id

        original_excel = full_dir / "original.xlsx"
        if not original_excel.exists():
            raise ValueError("全量工参原始文件不存在")

        # 2. 准备输出文件
        # 生成新文件名：原文件名 + 时间戳
        original_name = full_info["name"]
        name_part, ext_part = os.path.splitext(original_name)
        # 移除可能已有的时间戳（简单处理：如果文件名以数字结尾，可能是时间戳）
        # 这里直接追加时间戳
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")

        # 尝试替换文件名中的最后一段数字串（如果存在）
        new_name = re.sub(r"\d{14}", timestamp, name_part)
        if new_name == name_part:
            new_name = f"{name_part}_{timestamp}"

        new_filename = f"{new_name}{ext_part}"

        # 创建新的数据ID
        new_data_id = str(uuid.uuid4())
        new_data_dir = settings.DATA_DIR / new_data_id
        new_data_dir.mkdir(parents=True, exist_ok=True)

        output_excel_path = new_data_dir / "original.xlsx"

        # 3. 加载Excel
        safe_print(f"[DataService] 加载Excel: {original_excel}")
        try:
            wb = openpyxl.load_workbook(original_excel)
        except Exception as e:
            raise ValueError(f"无法打开Excel文件: {e}")

        # 4. 处理LTE
        if "LTE Project Parameters" in wb.sheetnames:
            safe_print(f"[DataService] 处理LTE工参...")
            self._update_sheet(wb["LTE Project Parameters"], current_dir, "LTE")
        else:
            safe_print(f"[DataService] 警告: 未找到 'LTE Project Parameters' Sheet")

        # 5. 处理NR
        if "NR Project Parameters" in wb.sheetnames:
            safe_print(f"[DataService] 处理NR工参...")
            self._update_sheet(wb["NR Project Parameters"], current_dir, "NR")
        else:
            safe_print(f"[DataService] 警告: 未找到 'NR Project Parameters' Sheet")

        # 6. 保存结果
        safe_print(f"[DataService] 保存结果到: {output_excel_path}")
        wb.save(output_excel_path)
        wb.close()

        # 7. 解析新文件以生成 data.json (复用 upload_excel 的部分逻辑)
        safe_print(f"[DataService] 解析新文件生成 data.json...")
        metadata = {}
        try:
            import json

            xls = pd.ExcelFile(output_excel_path)
            parsed_data = {}

            with xls:
                sheet_names = xls.sheet_names
                for network in ["LTE", "NR"]:
                    sheet_name = f"{network} Project Parameters"
                    if sheet_name in sheet_names:
                        sites = self._parse_sheet_data(xls, sheet_name, network)
                        parsed_data[network] = sites
                        metadata[f"{network}SiteCount"] = len(sites)
                        metadata[f"{network}SectorCount"] = sum(
                            len(s.get("sectors", [])) for s in sites
                        )

            with open(new_data_dir / "data.json", "w", encoding="utf-8") as f:
                json.dump(parsed_data, f, ensure_ascii=False, indent=2)

        except Exception as e:
            safe_print(f"[DataService] 解析新文件失败: {e}")
            # 即使解析失败，文件已经生成，我们还是注册它
            pass

        # 8. 注册新文件 & 保存到原始路径（如果存在）
        # 复制一份到uploads目录以便下载（模拟上传行为）
        new_upload_path = settings.UPLOAD_DIR / f"{new_data_id}.xlsx"
        shutil.copy(output_excel_path, new_upload_path)

        # 尝试保存到原始目录
        saved_to_original = False
        original_path = full_info.get("originalPath")
        safe_print(
            f"[DataService] 原始路径信息: originalPath='{original_path}', 类型={type(original_path).__name__}"
        )

        if original_path:
            try:
                # 确保路径是绝对路径并处理编码
                original_file_path = os.path.abspath(original_path)
                original_dir = os.path.dirname(original_file_path)

                safe_print(
                    f"[DataService] 解析后路径: original_file_path='{original_file_path}', original_dir='{original_dir}'"
                )

                if os.path.exists(original_dir):
                    target_path = os.path.join(original_dir, new_filename)
                    safe_print(f"[DataService] 准备保存副本到: {target_path}")

                    shutil.copy2(output_excel_path, target_path)
                    if os.path.exists(target_path):
                        safe_print(f"[DataService] [OK] 副本保存成功到原路径!")
                        saved_to_original = True
                    else:
                        safe_print(f"[DataService] [X] 副本保存失败 - 文件不存在")
                else:
                    safe_print(f"[DataService] [X] 原始目录不存在: {original_dir}")
            except Exception as e:
                safe_print(
                    f"[DataService] [X] 无法保存到原始目录 ({original_path}): {e}"
                )
                import traceback

                traceback.print_exc()
        else:
            safe_print(
                f"[DataService] [X] 未能获取到原始路径 (originalPath is empty or None)"
            )
            safe_print(f"[DataService] 提示: 请检查前端是否正确传递了 file_path 参数")

        self.index[new_data_id] = {
            "id": new_data_id,
            "name": new_filename,
            "type": "excel",
            "fileType": "full_params",
            "size": output_excel_path.stat().st_size,
            "uploadDate": datetime.now().isoformat(),
            "status": "ready",
            "originalPath": full_info.get("originalPath"),  # 继承路径
            "metadata": metadata,
        }
        self._save_index()
        safe_print(f"[DataService] 索引已更新")

        return {
            "newFileId": new_data_id,
            "newFileName": new_filename,
            "savedToOriginal": saved_to_original,
        }

    def _update_sheet(self, ws, current_dir: Path, network_type: str):
        """更新单个Sheet"""
        # 1. 确定列索引映射
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        col_map = {}  # Chinese Name -> Column Index (1-based)

        for idx, col_val in enumerate(header_row):
            if col_val:
                # 提取中文名称（处理可能的前导换行符，取第一行非空内容）
                # 例如: "\ngNodeBLength\n..." -> "gNodeBLength"
                clean_name = str(col_val).strip().split("\n")[0].strip()
                col_map[clean_name] = idx + 1

        # 调试：检查是否找到小区覆盖类型列
        if "小区覆盖类型" in col_map:
            safe_print(
                f"[DataService] 找到'小区覆盖类型'列，列索引={col_map['小区覆盖类型']}"
            )
        else:
            safe_print(f"[DataService] 警告: 未找到'小区覆盖类型'列")

        # 2. 读取现网工参数据
        updates = self._load_current_params(current_dir, network_type)
        safe_print(f"[DataService] 加载到 {len(updates)} 条 {network_type} 现网数据")

        # 3. 遍历Excel行进行更新
        # LTE Key: eNodeB标识 + 小区标识
        # NR Key: 移动国家码 + 移动网络码 + gNodeB标识 + 小区标识

        # 构建Excel中的现有Key集合
        existing_keys = set()

        # 映射字段配置
        if network_type == "LTE":
            key_cols = ["eNodeB标识", "小区标识"]
            # 源CSV字段 -> 目标Excel字段
            # 注意：允许多个源字段映射到同一个目标字段（后遍历的会覆盖前面的，或者根据是否存在来决定）
            # 建议将优先使用的字段放在后面（字典顺序），或者在逻辑中处理
            field_map = {
                "eNBName": "基站名称",
                "UserLabel": "基站名称",  # 备选
                "SubNetwork": "子网ID",
                "ManagedElement": "管理网元ID",
                "cellName": "小区名称",
                "CellName": "小区名称",  # 备选 (常见大写)
                "tac": "跟踪区码",
                "pci": "物理小区识别码",
                "frequency": "下行链路的中心载频",
                "cellLocalId": "小区标识",
                "eNBId": "eNodeB标识",
                "mcc": "移动国家码",
                "mnc": "移动网络码",
            }
        else:  # NR
            key_cols = ["移动国家码", "移动网络码", "gNodeB标识", "小区标识"]
            field_map = {
                "eNBName": "基站名称",
                "gNBName": "基站名称",
                "UserLabel": "基站名称",  # 备选
                "dra_GNBName": "基站名称",  # 备选
                "SubNetwork": "子网ID",
                "ManagedElement": "管理网元ID",
                "cellName": "小区名称",
                "CellName": "小区名称",  # 备选
                "tac": "跟踪区码",
                "pci": "物理小区识别码",
                "ssbFrequency": "填写SSB频点",  # 修改：使用ssbFrequency而不是frequency
                "frequencyBandList": "gNodeBLength",  # 新增映射
                "cellLocalId": "小区标识",
                "gNBId": "gNodeB标识",
                "mcc": "移动国家码",
                "mnc": "移动网络码",
                # 注意：coverageType 不在 field_map 中，而是单独处理
            }

        # 验证必要的列是否存在于Excel中
        missing_cols = [k for k in key_cols if k not in col_map]
        if missing_cols:
            safe_print(
                f"[DataService] 警告: Sheet中缺少Key列: {missing_cols}，跳过更新"
            )
            return

        # 遍历数据行（从第4行开始）
        # 注意：openpyxl的row是1-based
        for row in ws.iter_rows(min_row=4):
            # 获取Key值
            try:
                if network_type == "LTE":
                    enb_id = self._get_cell_val(row, col_map, "eNodeB标识")
                    cell_id = self._get_cell_val(row, col_map, "小区标识")
                    if enb_id is not None and cell_id is not None:
                        key = f"{enb_id}_{cell_id}"
                        existing_keys.add(key)

                        # 检查是否有更新
                        if key in updates:
                            # 获取当前基站名称（更新前）
                            current_bs_name = None
                            if "基站名称" in col_map:
                                current_bs_name = self._get_cell_val(row, col_map, "基站名称")

                            self._apply_row_update(
                                row, col_map, updates[key], field_map, network_type
                            )
                            # 标记为已处理（以便后续追加未处理的）
                            updates[key]["_processed"] = True

                            # 检查基站名称是否被更新
                            new_bs_name = None
                            if "基站名称" in col_map:
                                new_bs_name = self._get_cell_val(row, col_map, "基站名称")

                            if current_bs_name != new_bs_name:
                                safe_print(f"[LTE UPDATE] 基站名称已更新: Key={key}, 原值={current_bs_name}, 新值={new_bs_name}")
                        else:
                            # LTE Key未找到的调试日志（只记录前几条，避免日志过多）
                            missing_count = getattr(self, '_lte_missing_key_count', 0) + 1
                            self._lte_missing_key_count = missing_count
                            if missing_count <= 10:
                                safe_print(f"[LTE DEBUG] Key未在现网工参中找到: {key}")
                            elif missing_count == 11:
                                safe_print(f"[LTE DEBUG] ...（后续未匹配的Key将不再打印）")

                else:  # NR
                    mcc = self._get_cell_val(row, col_map, "移动国家码")
                    mnc = self._get_cell_val(row, col_map, "移动网络码")
                    gnb_id = self._get_cell_val(row, col_map, "gNodeB标识")
                    cell_id = self._get_cell_val(row, col_map, "小区标识")

                    if all(v is not None for v in [mcc, mnc, gnb_id, cell_id]):
                        key = f"{mcc}_{mnc}_{gnb_id}_{cell_id}"
                        existing_keys.add(key)

                        if key in updates:
                            try:
                                # 获取当前基站名称（更新前）
                                current_bs_name = None
                                if "基站名称" in col_map:
                                    current_bs_name = self._get_cell_val(row, col_map, "基站名称")

                                self._apply_row_update(
                                    row, col_map, updates[key], field_map, network_type
                                )
                                updates[key]["_processed"] = True

                                # 检查基站名称是否被更新
                                new_bs_name = None
                                if "基站名称" in col_map:
                                    new_bs_name = self._get_cell_val(row, col_map, "基站名称")

                                if current_bs_name != new_bs_name:
                                    safe_print(f"[UPDATE] 基站名称已更新: Key={key}, 原值={current_bs_name}, 新值={new_bs_name}")
                            except Exception as e:
                                import traceback

                                safe_print(f"[DataService] 更新行失败 (Key={key}): {e}")
                                traceback.print_exc()
                        else:
                            # Key未找到的调试日志（采样打印，避免日志过多）
                            safe_print(f"[DEBUG] Key未在现网工参中找到: {key}")
            except Exception as e:
                import traceback

                safe_print(f"[DataService] 处理行时发生错误: {e}")
                traceback.print_exc()
                continue

        # 5. 追加新行
        new_rows_count = 0
        for key, data in updates.items():
            if not data.get("_processed"):
                # 准备新行数据
                self._append_new_row(ws, col_map, data, field_map, network_type)
                new_rows_count += 1

        # 4. 处理特殊列：CGI、跟踪区码、是否共享、第一分组、gNodeBLength
        # 收集所有行数据用于处理（包括新追加的行）
        all_rows = list(ws.iter_rows(min_row=4))

        if network_type == "NR":
            # NR网络处理
            # 4.1 生成CGI列
            cgi_list = []
            for row in all_rows:
                gnb_id = self._get_cell_val(row, col_map, "gNodeB标识")
                cell_id = self._get_cell_val(row, col_map, "小区标识")
                if gnb_id and cell_id:
                    cgi = f"{gnb_id}-{cell_id}"
                    cgi_list.append(cgi)
                    # 更新CGI列
                    if "CGI" in col_map:
                        row[col_map["CGI"] - 1].value = cgi
                else:
                    cgi_list.append(None)

            # 4.2 更新跟踪区码（从现网工参获取）
            for i, row in enumerate(all_rows):
                mcc = self._get_cell_val(row, col_map, "移动国家码")
                mnc = self._get_cell_val(row, col_map, "移动网络码")
                gnb_id = self._get_cell_val(row, col_map, "gNodeB标识")
                cell_id = self._get_cell_val(row, col_map, "小区标识")
                if all([mcc, mnc, gnb_id, cell_id]):
                    key = f"{mcc}_{mnc}_{gnb_id}_{cell_id}"
                    if key in updates and "tac" in updates[key]:
                        tac = updates[key]["tac"]
                        if (
                            "跟踪区码" in col_map
                            and tac is not None
                            and not pd.isna(tac)
                        ):
                            row[col_map["跟踪区码"] - 1].value = tac

            # 4.3 计算是否共享列
            if "是否共享" in col_map:
                # 统计CGI出现次数
                cgi_counts = {}
                for cgi in cgi_list:
                    if cgi:
                        cgi_counts[cgi] = cgi_counts.get(cgi, 0) + 1
                # 更新是否共享列
                for i, row in enumerate(all_rows):
                    cgi = cgi_list[i]
                    if cgi:
                        count = cgi_counts.get(cgi, 0)
                        row[col_map["是否共享"] - 1].value = (
                            "是" if count == 2 else "否"
                        )

            # 4.4 更新第一分组列：在现网工参能找到的小区都填为"电信中兴"
            for i, row in enumerate(all_rows):
                mcc = self._get_cell_val(row, col_map, "移动国家码")
                mnc = self._get_cell_val(row, col_map, "移动网络码")
                gnb_id = self._get_cell_val(row, col_map, "gNodeB标识")
                cell_id = self._get_cell_val(row, col_map, "小区标识")
                if all([mcc, mnc, gnb_id, cell_id]):
                    key = f"{mcc}_{mnc}_{gnb_id}_{cell_id}"
                    if key in updates:
                        if "第一分组" in col_map:
                            row[col_map["第一分组"] - 1].value = "电信中兴"

            # 4.5 处理gNodeBLength：未找到的小区根据填写SSB频点判断
            for i, row in enumerate(all_rows):
                mcc = self._get_cell_val(row, col_map, "移动国家码")
                mnc = self._get_cell_val(row, col_map, "移动网络码")
                gnb_id = self._get_cell_val(row, col_map, "gNodeB标识")
                cell_id = self._get_cell_val(row, col_map, "小区标识")
                if all([mcc, mnc, gnb_id, cell_id]):
                    key = f"{mcc}_{mnc}_{gnb_id}_{cell_id}"
                    if key not in updates:
                        # 未找到的小区，根据填写SSB频点判断
                        ssb_frequency = self._get_cell_val(row, col_map, "填写SSB频点")
                        if ssb_frequency:
                            try:
                                # 尝试将SSB频点转换为数值，处理范围判断
                                # 情况1: 直接是数值，如 "925"
                                ssb_num = float(ssb_frequency)
                                if 900 <= ssb_num <= 955:
                                    if "gNodeBLength" in col_map:
                                        row[col_map["gNodeBLength"] - 1].value = 8
                                # 情况2: 范围字符串，如 "900-955"
                            except ValueError:
                                # 如果转换失败，检查是否是范围字符串
                                if "-" in ssb_frequency:
                                    try:
                                        # 提取范围的上下限
                                        lower, upper = ssb_frequency.split("-")
                                        lower_num = float(lower.strip())
                                        upper_num = float(upper.strip())
                                        # 检查范围是否与900-955有重叠
                                        if lower_num <= 955 and upper_num >= 900:
                                            if "gNodeBLength" in col_map:
                                                row[
                                                    col_map["gNodeBLength"] - 1
                                                ].value = 8
                                    except ValueError:
                                        # 如果范围解析失败，跳过
                                        pass
                                # 检查是否完全匹配 "900-955"
                                elif ssb_frequency == "900-955":
                                    if "gNodeBLength" in col_map:
                                        row[col_map["gNodeBLength"] - 1].value = 8

        else:  # LTE网络
            # LTE网络处理
            # 4.1 生成CGI列
            for row in all_rows:
                enb_id = self._get_cell_val(row, col_map, "eNodeB标识")
                cell_id = self._get_cell_val(row, col_map, "小区标识")
                if enb_id and cell_id:
                    cgi = f"{enb_id}-{cell_id}"
                    # 更新CGI列
                    if "CGI" in col_map:
                        row[col_map["CGI"] - 1].value = cgi

            # 4.2 更新是否共享列（基于plmnIdList）
            for i, row in enumerate(all_rows):
                enb_id = self._get_cell_val(row, col_map, "eNodeB标识")
                cell_id = self._get_cell_val(row, col_map, "小区标识")
                if enb_id and cell_id:
                    key = f"{enb_id}_{cell_id}"
                    if key in updates and "plmnIdList" in updates[key]:
                        plmn_id_list = updates[key]["plmnIdList"]
                        if "是否共享" in col_map:
                            if (
                                plmn_id_list
                                and isinstance(plmn_id_list, str)
                                and ";" in plmn_id_list
                            ):
                                row[col_map["是否共享"] - 1].value = "是"
                            else:
                                row[col_map["是否共享"] - 1].value = "否"

            # 4.3 更新第一分组列：在现网工参能找到的小区都填为"电信中兴"
            for i, row in enumerate(all_rows):
                enb_id = self._get_cell_val(row, col_map, "eNodeB标识")
                cell_id = self._get_cell_val(row, col_map, "小区标识")
                if enb_id and cell_id:
                    key = f"{enb_id}_{cell_id}"
                    if key in updates:
                        if "第一分组" in col_map:
                            row[col_map["第一分组"] - 1].value = "电信中兴"

        safe_print(
            f"[DataService] 更新完成: 现有匹配 {len(existing_keys)} 行, 追加新行 {new_rows_count} 行"
        )

    def _get_cell_val(self, row, col_map, col_name):
        """获取行中指定列的值"""
        if col_name in col_map:
            idx = col_map[col_name] - 1  # row is tuple, 0-based index
            val = row[idx].value
            if val is not None:
                return str(val).strip()
        return None

    def _apply_row_update(
        self, row, col_map, update_data, field_map, network_type: str
    ):
        """应用更新到行"""
        # LTE 专用：检查 Excel 中是否有"基站名称"列
        if network_type == "LTE" and "基站名称" not in col_map:
            safe_print(f"[LTE WARN] Excel 中没有 '基站名称' 列!")
            safe_print(f"[LTE WARN] Excel 可用列: {list(col_map.keys())}")

        # 关键修复：处理多个源字段映射到同一目标列的情况
        # 问题：如果后遍历的字段为空，不应覆盖之前已设置的值
        # 解决：为每个目标列只取第一个非空的有效值

        # 按目标列分组源字段（字典的迭代顺序在 Python 3.7+ 中是有序的）
        # 对于每个目标列，按源字段顺序检查，第一个非空值生效
        target_cols_map = {}
        for src_field, target_col in field_map.items():
            if target_col not in target_cols_map:
                target_cols_map[target_col] = []
            target_cols_map[target_col].append(src_field)

        # 对每个目标列，按优先级顺序查找第一个非空值
        for target_col, src_fields in target_cols_map.items():
            if target_col not in col_map:
                continue

            col_idx = col_map[target_col] - 1

            # 按顺序查找第一个非空的源字段值
            final_val = None
            used_src_field = None

            # 基站名称列的特殊调试：记录所有尝试的字段值
            if target_col == "基站名称":
                all_vals = []
                for sf in src_fields:
                    if sf in update_data:
                        all_vals.append(f"{sf}={repr(update_data[sf])}")
                    else:
                        all_vals.append(f"{sf}=不存在")
                if all_vals:
                    safe_print(f"[LTE DEBUG] 基站名称可用字段: {', '.join(all_vals)}")
                    # 同时显示 update_data 中所有字段（前20个）用于调试
                    available_fields = list(update_data.keys())
                    safe_print(f"[LTE DEBUG] update_data 中的所有字段（部分）: {available_fields[:20]}")

            for src_field in src_fields:
                if src_field not in update_data:
                    continue

                val = update_data[src_field]

                # 检查是否为空值（处理numpy NaN、pandas NA、None、空字符串等）
                if val is None:
                    continue
                try:
                    if pd.isna(val):
                        continue
                except (TypeError, ValueError):
                    pass
                if isinstance(val, float) and np.isnan(val):
                    continue
                # 空字符串也视为无效值
                if isinstance(val, str) and not val.strip():
                    continue

                # 找到第一个非空值，使用它
                final_val = val
                used_src_field = src_field
                break  # 找到有效值后立即停止，不被后续字段覆盖

            if final_val is not None:
                # 转换为Python原生类型（处理numpy类型）
                if hasattr(final_val, "item"):  # numpy scalar
                    final_val = final_val.item()
                elif isinstance(final_val, str):
                    final_val = final_val.strip()

                row[col_idx].value = final_val
                # 调试日志（对于基站名称列）
                if target_col == "基站名称":
                    safe_print(f"[LTE UPDATE] 更新基站名称: {used_src_field}={final_val}")
            elif target_col == "基站名称":
                # 特殊日志：基站名称列未找到有效值
                safe_print(f"[LTE WARN] 基站名称未更新: 所有字段都为空")

        # 特殊处理：NR coverageType 映射到小区覆盖类型
        # 只有NR网络类型才处理，且只在现网工参中有coverageType值时才更新
        # LTE网络不处理小区覆盖类型
        if network_type == "NR" and "coverageType" in update_data:
            val = update_data["coverageType"]

            # 调试日志
            safe_print(f"[DataService] NR小区: 发现coverageType字段，原始值={val}")

            # 检查是否为空值
            if not (
                pd.isna(val)
                or val is None
                or (isinstance(val, float) and np.isnan(val))
            ):
                # 转换为字符串并去除空格
                val_str = str(val).strip()

                # coverageType: "Micro" -> 4 (室内), "Macro" -> 1 (室外)
                # 只有明确的Micro或Macro才更新，其他值不更新
                if val_str == "Micro":
                    coverage_val = 4
                    possible_msg = (
                        f"[DataService] NR coverageType映射: Micro -> 小区覆盖类型=4"
                    )
                    safe_print(possible_msg)
                    # 如果Excel中有"小区覆盖类型"列，则更新
                    if "小区覆盖类型" in col_map:
                        col_idx = col_map["小区覆盖类型"] - 1
                        old_val = row[col_idx].value
                        row[col_idx].value = coverage_val
                        safe_print(
                            f"[DataService] 已更新: 原值={old_val}, 新值={coverage_val}"
                        )
                    else:
                        safe_print(
                            f"[DataService] 警告: Excel中没有'小区覆盖类型'列，无法更新"
                        )
                elif val_str == "Macro":
                    coverage_val = 1
                    safe_print(
                        f"[DataService] NR coverageType映射: Macro -> 小区覆盖类型=1"
                    )
                    # 如果Excel中有"小区覆盖类型"列，则更新
                    if "小区覆盖类型" in col_map:
                        col_idx = col_map["小区覆盖类型"] - 1
                        old_val = row[col_idx].value
                        row[col_idx].value = coverage_val
                        safe_print(
                            f"[DataService] 已更新: 原值={old_val}, 新值={coverage_val}"
                        )
                    else:
                        safe_print(
                            f"[DataService] 警告: Excel中没有'小区覆盖类型'列，无法更新"
                        )
                else:
                    # 未知值不更新（保持原值）
                    safe_print(
                        f"[DataService] NR coverageType未知值: {val_str}，跳过更新小区覆盖类型"
                    )
            else:
                safe_print(f"[DataService] NR coverageType为空值，跳过更新")

        # 固定值更新
        # 系统制式固定值是1
        if "系统制式" in col_map:
            row[col_map["系统制式"] - 1].value = 1

    def _append_new_row(self, ws, col_map, data, field_map, network_type):
        """追加新行"""
        # 创建一个空行列表，长度为最大列索引
        max_col = max(col_map.values()) if col_map else 0
        new_row = [None] * max_col

        # 填充映射字段
        for src_field, target_col in field_map.items():
            if src_field in data and target_col in col_map:
                col_idx = col_map[target_col] - 1
                val = data[src_field]

                # 检查是否为空值
                if (
                    pd.isna(val)
                    or val is None
                    or (isinstance(val, float) and np.isnan(val))
                ):
                    continue  # 跳过空值

                # 转换为Python原生类型（处理numpy类型）
                if hasattr(val, "item"):  # numpy scalar
                    val = val.item()
                elif isinstance(val, str):
                    val = val.strip()

                new_row[col_idx] = val

        # 特殊处理：NR coverageType 映射到小区覆盖类型
        # 只有NR网络类型才处理，且只在现网工参中有coverageType值时才更新
        # LTE网络不处理小区覆盖类型
        # 注意：新行（不在现网工参中的小区）只有在现网工参中有明确coverageType值时才设置
        if network_type == "NR" and "coverageType" in data:
            val = data["coverageType"]

            # 调试日志
            safe_print(f"[DataService] NR新行: 发现coverageType字段，原始值={val}")

            # 检查是否为空值
            if not (
                pd.isna(val)
                or val is None
                or (isinstance(val, float) and np.isnan(val))
            ):
                # 转换为字符串并去除空格
                val_str = str(val).strip()

                # coverageType: "Micro" -> 4 (室内), "Macro" -> 1 (室外)
                # 只有明确的Micro或Macro才设置，其他值不设置（保持默认或空）
                if val_str == "Micro":
                    coverage_val = 4
                    safe_print(
                        f"[DataService] NR新行 coverageType映射: Micro -> 小区覆盖类型=4"
                    )
                    # 如果Excel中有"小区覆盖类型"列，则设置
                    if "小区覆盖类型" in col_map:
                        col_idx = col_map["小区覆盖类型"] - 1
                        new_row[col_idx] = coverage_val
                        safe_print(
                            f"[DataService] NR新行已设置: 小区覆盖类型={coverage_val}"
                        )
                    else:
                        safe_print(
                            f"[DataService] NR新行警告: Excel中没有'小区覆盖类型'列"
                        )
                elif val_str == "Macro":
                    coverage_val = 1
                    safe_print(
                        f"[DataService] NR新行 coverageType映射: Macro -> 小区覆盖类型=1"
                    )
                    # 如果Excel中有"小区覆盖类型"列，则设置
                    if "小区覆盖类型" in col_map:
                        col_idx = col_map["小区覆盖类型"] - 1
                        new_row[col_idx] = coverage_val
                        safe_print(
                            f"[DataService] NR新行已设置: 小区覆盖类型={coverage_val}"
                        )
                    else:
                        safe_print(
                            f"[DataService] NR新行警告: Excel中没有'小区覆盖类型'列"
                        )
                else:
                    # 未知值不设置（保持Excel默认值或空）
                    safe_print(
                        f"[DataService] NR新行 coverageType未知值: {val_str}，不设置小区覆盖类型"
                    )
            else:
                safe_print(
                    f"[DataService] NR新行 coverageType为空值，不设置小区覆盖类型"
                )

        # 填充固定字段
        if "系统制式" in col_map:
            new_row[col_map["系统制式"] - 1] = 1

        # 生成CGI列
        if "CGI" in col_map:
            if network_type == "LTE":
                # LTE CGI: eNodeB标识-小区标识
                enb_id = data.get("eNBId")
                cell_id = data.get("cellLocalId")
                if enb_id and cell_id:
                    cgi = f"{enb_id}-{cell_id}"
                    new_row[col_map["CGI"] - 1] = cgi
            else:  # NR
                # NR CGI: gNodeB标识-小区标识
                gnb_id = data.get("gNBId")
                cell_id = data.get("cellLocalId")
                if gnb_id and cell_id:
                    cgi = f"{gnb_id}-{cell_id}"
                    new_row[col_map["CGI"] - 1] = cgi

        # 更新跟踪区码
        if "跟踪区码" in col_map:
            tac = data.get("tac")
            if tac is not None and not pd.isna(tac):
                # 转换为Python原生类型
                if hasattr(tac, "item"):  # numpy scalar
                    tac = tac.item()
                elif isinstance(tac, str):
                    tac = tac.strip()
                new_row[col_map["跟踪区码"] - 1] = tac

        # 更新是否共享列
        if "是否共享" in col_map:
            if network_type == "LTE":
                # LTE: 基于plmnIdList判断
                plmn_id_list = data.get("plmnIdList")
                if (
                    plmn_id_list
                    and isinstance(plmn_id_list, str)
                    and ";" in plmn_id_list
                ):
                    new_row[col_map["是否共享"] - 1] = "是"
                else:
                    new_row[col_map["是否共享"] - 1] = "否"
            else:  # NR
                # NR: 先设置默认值，后续会根据CGI出现次数更新
                new_row[col_map["是否共享"] - 1] = "否"

        # 将新行添加到工作表
        ws.append(new_row)

    def _load_current_params(
        self, current_dir: Path, network_type: str
    ) -> Dict[str, Dict]:
        """读取现网工参CSV文件"""
        updates = {}

        if network_type == "LTE":
            patterns = ["LTE_SDR_CellInfo_*.csv", "LTE_ITBBU_CellInfo_*.csv"]
        else:
            patterns = ["NR_CellInfo_*.csv"]

        files = []
        for pat in patterns:
            files.extend(list(current_dir.glob(pat)))
            # Also try recursive if inside subfolders
            files.extend(list(current_dir.glob(f"**/{pat}")))

        try:
            safe_print(
                f"[DataService] 找到 {len(files)} 个CSV文件: {[f.name for f in files]}"
            )
        except UnicodeEncodeError:
            # 如果文件名包含特殊字符，使用ASCII安全打印
            safe_names = [
                f.name.encode("ascii", "replace").decode("ascii") for f in files
            ]
            safe_print(f"[DataService] 找到 {len(files)} 个CSV文件: {safe_names}")

        for file_path in files:
            try:
                # 尝试不同的编码
                try:
                    df = pd.read_csv(file_path, encoding="gbk")
                except:
                    df = pd.read_csv(file_path, encoding="utf-8")

                # 标准化列名（去除空格）
                df.columns = [c.strip() for c in df.columns]

                # 调试：打印CSV列名（使用安全打印避免GBK编码错误）
                try:
                    safe_print(
                        f"[DataService] CSV文件 {file_path.name} 的列名: {list(df.columns)}"
                    )
                except UnicodeEncodeError:
                    # 如果列名包含特殊字符，使用ASCII安全打印
                    safe_columns = [
                        str(c).encode("ascii", "replace").decode("ascii")
                        for c in df.columns
                    ]
                    safe_print(
                        f"[DataService] CSV文件 {file_path.name} 的列名: {safe_columns}"
                    )

                # LTE 专用：检查基站名称相关字段
                if network_type == "LTE":
                    bs_name_fields = ["eNBName", "UserLabel"]
                    found_bs_fields = [f for f in bs_name_fields if f in df.columns]
                    if found_bs_fields:
                        safe_print(f"[LTE DEBUG] CSV文件 {file_path.name} 中找到基站名称字段: {found_bs_fields}")
                        # 统计非空值数量
                        for field in found_bs_fields:
                            non_empty_count = df[field].notna().sum()
                            safe_print(f"  - {field}: {non_empty_count} 条非空记录")
                            # 显示前几个非空值作为样本
                            sample_values = df[field].dropna().head(3).tolist()
                            safe_print(f"  - {field} 样本值: {sample_values}")
                    else:
                        safe_print(f"[LTE WARN] CSV文件 {file_path.name} 中未找到基站名称字段 ({bs_name_fields})")
                        safe_print(f"[LTE WARN] CSV文件实际列名: {list(df.columns)}")
                        # 检查是否有类似的列名（大小写、空格等）
                        for field in bs_name_fields:
                            similar = [c for c in df.columns if field.lower() in c.lower()]
                            if similar:
                                safe_print(f"[LTE HINT] 可能匹配 '{field}' 的列: {similar}")

                # 特别检查 coverageType 列（仅 NR）
                if network_type == "NR" and "coverageType" in df.columns:
                    unique_values = df["coverageType"].dropna().unique()
                    try:
                        safe_print(
                            f"[DataService] CSV文件 {file_path.name} 中 coverageType 的唯一值: {list(unique_values)}"
                        )
                    except UnicodeEncodeError:
                        # 如果值包含特殊字符，使用ASCII安全打印
                        safe_values = [
                            str(v).encode("ascii", "replace").decode("ascii")
                            for v in unique_values
                        ]
                        safe_print(
                            f"[DataService] CSV文件 {file_path.name} 中 coverageType 的唯一值: {safe_values}"
                        )
                else:
                    safe_print(
                        f"[DataService] CSV文件 {file_path.name} 中没有 coverageType 列"
                    )

                for _, row in df.iterrows():
                    item = row.to_dict()

                    # 处理PLMN分解
                    if "plmn" in item and pd.notna(item["plmn"]):
                        plmn_parts = str(item["plmn"]).split("-")
                        if len(plmn_parts) >= 2:
                            item["mcc"] = plmn_parts[0]
                            item["mnc"] = plmn_parts[1]
                        else:
                            item["mcc"] = ""
                            item["mnc"] = ""
                    else:
                        item["mcc"] = ""
                        item["mnc"] = ""

                    # 生成Key
                    key = None
                    if network_type == "LTE":
                        if "eNBId" in item and "cellLocalId" in item:
                            key = f"{item['eNBId']}_{item['cellLocalId']}"
                    else:
                        if all(
                            k in item for k in ["mcc", "mnc", "gNBId", "cellLocalId"]
                        ):
                            key = f"{item['mcc']}_{item['mnc']}_{item['gNBId']}_{item['cellLocalId']}"

                    if key:
                        updates[key] = item
                        # LTE 专用：检查基站名称字段（仅打印调试日志，不写入updates）
                        if network_type == "LTE":
                            bs_name_fields = ["eNBName", "UserLabel"]
                        else:
                            bs_name_fields = ["eNBName", "gNBName", "UserLabel", "dra_GNBName"]

                        bs_values = {}
                        has_bs_name = False
                        for field in bs_name_fields:
                            if field in item and pd.notna(item[field]) and str(item[field]).strip():
                                bs_values[field] = item[field]
                                has_bs_name = True

                        if has_bs_name:
                            # 只记录前3条有基站名称的数据作为样本（纯日志，不污染updates）
                            if len(updates) <= 3:
                                safe_print(f"[DEBUG] Key={key}, 基站名称字段: {bs_values}")
                        elif network_type == "LTE":
                            # LTE 专用：记录没有基站名称的数据（仅前5条，纯日志）
                            if len(updates) <= 5:
                                safe_print(f"[LTE WARN] Key={key} 没有任何基站名称字段")

            except Exception as e:
                import traceback

                safe_print(f"[DataService] 读取CSV失败 {file_path.name}: {e}")
                traceback.print_exc()

        safe_print(f"[DataService] 总共加载 {len(updates)} 条 {network_type} 现网数据")
        return updates

    async def upload_excel(
        self, file: Optional[Any], original_path: Optional[str] = None
    ) -> Dict[str, Any]:
        """上传并解析Excel工参文件"""
        safe_print(f"[DataService] ===== 开始上传Excel文件 =====")

        # 处理路径引号
        if original_path:
            original_path = original_path.strip("\"'")

        safe_print(f"[DataService] 原始路径: {original_path}")

        # 生成唯一ID
        data_id = str(uuid.uuid4())
        safe_print(f"[DataService] 生成ID: {data_id}")

        # 获取文件名和扩展名
        # 优先使用 original_path 来获取文件名，因为它包含完整路径和真实文件名
        if original_path:
            filename = os.path.basename(original_path)
            ext = os.path.splitext(filename)[1]
        elif file:
            filename = file.filename
            if not filename:
                raise ValueError("无法从上传的文件中获取文件名")
            ext = os.path.splitext(filename)[1]
        else:
            raise ValueError("必须提供文件或文件路径")

        safe_print(f"[DataService] 文件名: {filename}")

        # 清理文件名：移除或替换可能导致Windows路径问题的字符
        import re

        safe_filename = re.sub(r'[<>:"/\\|?*]', "_", filename)
        # 限制文件名长度（避免Windows路径长度限制）
        if len(safe_filename) > 100:
            name, text_ext = os.path.splitext(safe_filename)
            safe_filename = name[:90] + text_ext
        safe_print(f"[DataService] 安全文件名: {safe_filename}")

        # 保存文件 - 使用UUID避免中文文件名导致的路径问题
        file_path = settings.UPLOAD_DIR / f"{data_id}{ext}"
        safe_print(f"[DataService] 目标路径: {file_path}")

        try:
            # 保存或复制文件
            if file:
                safe_print(f"[DataService] 读取上传文件内容...")
                content = await file.read()
                safe_print(f"[DataService] 已读取 {len(content):,} 字节")
                with open(file_path, "wb") as f:
                    f.write(content)
            elif original_path:
                if not os.path.exists(original_path):
                    raise ValueError(f"文件不存在: {original_path}")
                safe_print(f"[DataService] 复制本地文件...")
                shutil.copy2(original_path, file_path)

            # 验证文件
            if file_path.exists():
                file_size = file_path.stat().st_size
                safe_print(f"[DataService] 文件验证成功，大小: {file_size:,} 字节")
            else:
                raise FileNotFoundError(f"文件保存失败: {file_path}")

        except Exception as e:
            safe_print(f"[DataService] 文件准备失败: {type(e).__name__}: {e}")
            raise ValueError(f"文件准备失败: {str(e)}")

        # 处理数据
        try:
            import json

            safe_print(f"[DataService] 开始解析Excel文件...")

            # 使用pandas.ExcelFile上下文管理器统一管理文件句柄
            # 这可以避免多次打开/关闭文件导致的Windows文件锁问题 (Errno 22)
            try:
                safe_print(f"[DataService] 打开Excel文件...")
                xls = pd.ExcelFile(file_path)
            except PermissionError as e:
                safe_print(f"[DataService] 权限错误: {e}")
            safe_print(f"[DataService] 开始解析文件...")

            # 使用 file_path 作为当前文件路径

            # 确保 filename 不是 None
            if filename is None:
                raise ValueError("文件名为 None，无法确定文件类型")

            # 根据文件名后缀判断格式
            if filename.lower().endswith(".csv"):
                safe_print(f"[DataService] 检测到CSV文件，使用CSV解析器...")
                # 处理CSV
                try:
                    df = pd.read_csv(file_path, encoding="utf-8")
                except:
                    df = pd.read_csv(file_path, encoding="gbk")

                # CSV文件没有sheet的概念，直接解析为默认类型
                file_type = "default"
                parsed_data = {
                    "default": self._parse_dataframe(df, filename, "unknown")
                }  # "unknown" network type for generic CSV
                metadata = {"siteCount": len(parsed_data["default"])}
                safe_print(
                    f"[DataService] CSV文件解析完成: {len(parsed_data['default'])} 个基站"
                )

            else:  # 默认为Excel文件
                safe_print(f"[DataService] 检测到Excel文件，使用Excel解析器...")
                # 使用pandas.ExcelFile上下文管理器统一管理文件句柄
                # 这可以避免多次打开/关闭文件导致的Windows文件锁问题 (Errno 22)
                try:
                    safe_print(f"[DataService] 打开Excel文件...")
                    xls = pd.ExcelFile(file_path)
                except PermissionError as e:
                    safe_print(f"[DataService] 权限错误: {e}")
                    raise ValueError(
                        f"文件被其他程序占用，请确保文件未被Excel打开: {str(e)}"
                    )
                except OSError as e:
                    safe_print(f"[DataService] 系统错误 [Errno {e.errno}]: {e}")
                    if e.errno == 22:
                        raise ValueError(
                            f"无法打开文件 (Errno 22)。可能原因：1) 文件损坏 2) 文件被占用 3) 路径包含特殊字符"
                        )
                    else:
                        raise ValueError(f"无法打开文件: {str(e)}")
                except Exception as e:
                    safe_print(f"[DataService] 未知错误: {type(e).__name__}: {e}")
                    raise ValueError(f"无法打开Excel文件: {str(e)}")

                try:
                    # 获取sheet名称
                    sheet_names = xls.sheet_names
                    try:
                        safe_print(f"[DataService] Excel文件包含的sheet: {sheet_names}")
                    except UnicodeEncodeError:
                        # 如果sheet名称包含特殊字符，使用ASCII安全打印
                        safe_sheet_names = [
                            s.encode("ascii", "replace").decode("ascii")
                            for s in sheet_names
                        ]
                        safe_print(
                            f"[DataService] Excel文件包含的sheet: {safe_sheet_names}"
                        )

                    # 判断文件类型
                    file_type = self._classify_file(filename, sheet_names)
                    safe_print(f"[DataService] 文件类型: {file_type}")

                    # === 单一实例逻辑 ===
                    # 如果是全量工参或待规划小区，先删除已存在的同类型文件
                    if file_type in ["full_params", "target_cells"]:
                        files_to_delete = []
                        for existing_id, existing_data in self.index.items():
                            # 检查fileType，如果没有fileType字段则忽略（或者是旧数据）
                            if existing_data.get("fileType") == file_type:
                                files_to_delete.append(existing_id)

                        if files_to_delete:
                            safe_print(
                                f"[DataService] 检测到已存在 {file_type} 文件 ({len(files_to_delete)} 个)，准备删除以保持唯一性..."
                            )
                            for data_id_to_delete in files_to_delete:
                                try:
                                    self.delete_data(data_id_to_delete)
                                    safe_print(
                                        f"[DataService] 已删除旧文件: {data_id_to_delete}"
                                    )
                                except Exception as e:
                                    safe_print(f"[DataService] 删除旧文件失败: {e}")

                    # 根据文件类型解析数据
                    parsed_data = {}
                    metadata = {}

                    if file_type == "full_params":
                        # 全量工参文件 - 读取LTE和NR Project Parameters子表
                        for network in ["LTE", "NR"]:
                            sheet_name = f"{network} Project Parameters"
                            if sheet_name in sheet_names:
                                safe_print(f"[DataService] 解析 {sheet_name}...")
                                # 传入xls对象而不是路径
                                sites = self._parse_sheet_data(xls, sheet_name, network)
                                parsed_data[network] = sites
                                metadata[f"{network}SiteCount"] = len(sites)
                                metadata[f"{network}SectorCount"] = sum(
                                    len(s.get("sectors", [])) for s in sites
                                )
                                safe_print(
                                    f"[DataService] {sheet_name} 解析完成: {len(sites)} 个基站"
                                )

                    elif file_type == "target_cells":
                        # 待规划小区文件 - 读取LTE和NR子表
                        for network in ["LTE", "NR"]:
                            if network in sheet_names:
                                safe_print(f"[DataService] 解析 {network}...")
                                # 传入xls对象而不是路径
                                sites = self._parse_sheet_data(xls, network, network)
                                parsed_data[network] = sites
                                metadata[f"{network}SiteCount"] = len(sites)
                                metadata[f"{network}SectorCount"] = sum(
                                    len(s.get("sectors", [])) for s in sites
                                )
                                safe_print(
                                    f"[DataService] {network} 解析完成: {len(sites)} 个基站"
                                )
                    else:
                        # 普通工参文件 - 尝试默认解析
                        safe_print(f"[DataService] 使用默认解析...")
                        # 传入xls对象而不是路径
                        sites = self._parse_default_excel(xls)
                        parsed_data["default"] = sites
                        metadata["siteCount"] = len(sites)
                        metadata["sectorCount"] = sum(
                            len(s.get("sectors", [])) for s in sites
                        )
                        safe_print(f"[DataService] 默认解析完成: {len(sites)} 个基站")

                finally:
                    # 确保关闭文件句柄
                    xls.close()

        except Exception as e:
            # 清理失败的文件
            safe_print(f"[DataService] ===== 上传失败 =====")
            safe_print(f"[DataService] 错误类型: {type(e).__name__}")
            safe_print(f"[DataService] 错误信息: {str(e)}")
            # 只有在是在这个函数里创建的临时文件才清理，但这里文件已经被保存到uploads了
            # 如果是上传过程中的临时文件，FastAPI会处理
            import traceback

            traceback.print_exc()
            raise ValueError(f"文件解析失败: {str(e)}")

        # 保存处理后的数据
        # 注意：之前的代码结构有点问题，保存逻辑应该在解析成功之后，且需要在xls关闭之后（如果是在Windows上移动文件的话）
        # 但这里 we are reading content, so xls closing later is fine as long as we data are in parsed_data

        safe_print(f"[DataService] 保存解析结果...")
        data_dir = settings.DATA_DIR / data_id
        data_dir.mkdir(parents=True, exist_ok=True)

        # 保存原始文件 - 此时file_path是uploads下的文件，xls已经关闭，应该可以复制
        try:
            shutil.copy(file_path, data_dir / "original.xlsx")
            safe_print(f"[DataService] 原始文件已保存")
        except Exception as e:
            safe_print(f"[DataService] 保存原始文件失败: {e}")
            # 继续保存JSON

        # 保存解析后的数据
        with open(data_dir / "data.json", "w", encoding="utf-8") as f:
            json.dump(parsed_data, f, ensure_ascii=False, indent=2)
        safe_print(f"[DataService] 数据已保存到 data.json")

        # 更新索引
        self.index[data_id] = {
            "id": data_id,
            "name": filename,
            "type": "excel",
            "fileType": file_type,
            "size": file_path.stat().st_size,
            "originalPath": original_path,
            "uploadDate": datetime.now().isoformat(),
            "status": "ready",
            "metadata": metadata,
        }
        safe_print(
            f"[DataService] 保存到索引: originalPath='{original_path}', filename='{filename}'"
        )
        self._save_index()
        safe_print(f"[DataService] 索引已更新")

        safe_print(f"[DataService] ===== 上传成功 =====")
        return {
            "id": data_id,
            "name": filename,
            "status": "ready",
            "fileType": file_type,
        }

    def _classify_file(self, filename: str, sheet_names: list) -> str:
        """根据文件名和sheet名称分类文件类型"""
        filename_lower = filename.lower()

        # 待规划小区文件特征 - 必须以"cell-tree-export"开头
        if filename_lower.startswith("cell-tree-export"):
            return "target_cells"

        # 全量工参文件特征 - 包含LTE Project Parameters和NR Project Parameters
        if "lte project parameters" in [
            s.lower() for s in sheet_names
        ] and "nr project parameters" in [s.lower() for s in sheet_names]:
            return "full_params"

        # 默认类型
        return "default"

    def _parse_sheet_data(
        self, excel_file, sheet_name: str, network_type: str
    ) -> List[Dict]:
        """解析指定sheet的数据"""
        import pandas as pd

        safe_print(f"[DataService] 正在解析sheet: {sheet_name} ({network_type})")

        # 检测是否是全量工参文件（LTE/NR Project Parameters）
        is_full_params = "Project Parameters" in sheet_name
        # 检测是否是待规划小区文件
        # 注意：这里简单判断sheet名是否为LTE/NR且似乎在待规划文件上下文中
        # 更好的方式是传递file_type，但为了保持签名兼容，我们可以通过sheet内容或上下文判断
        # 这里我们假设如果是 'LTE' 或 'NR' 这种简短名字，且解析失败时尝试target_cells逻辑
        # 但既然我们现在在这个方法里，我们不妨先尝试特定的解析逻辑

        # 实际上，调用者知道file_type。但在目前的架构中，_parse_sheet_data是通用的。
        # 我们可以检查列名来决定。

        if is_full_params:
            # 全量工参文件：第一行列名包含\n分隔的多行信息，从第4行开始是数据
            safe_print(f"[DataService] 检测到全量工参文件，使用专用解析逻辑")
            return self._parse_full_params_sheet(excel_file, sheet_name, network_type)

        # 尝试读取前几行来判断是否是target_cells（没有基站经纬度，只有ID）
        try:
            # 简单读取
            df_preview = pd.read_excel(excel_file, sheet_name=sheet_name, nrows=1)
            columns = list(df_preview.columns)

            # 待规划小区特征：包含 eNodeBID/gNodeBID 和 CellID，但不包含 Longitude/Latitude
            has_id = "eNodeBID" in columns or "gNodeBID" in columns
            has_cell = "CellID" in columns
            has_coords = (
                "Longitude" in columns or "Latitude" in columns or "基站经度" in columns
            )

            if has_id and has_cell and not has_coords:
                safe_print(f"[DataService] 检测到待规划小区文件格式，使用专用解析逻辑")
                return self._parse_target_cells_sheet(
                    excel_file, sheet_name, network_type
                )

        except Exception as e:
            safe_print(f"[DataService] 预检查失败: {e}")

        # 尝试不同的解析方式
        df = None

        # 方式1: 跳过前3行
        try:
            df = pd.read_excel(excel_file, sheet_name=sheet_name, header=3)
            # 简单验证一下是否有效
            if len(df.columns) > 1:
                safe_print(f"[DataService] 使用header=3解析成功，共 {len(df)} 行")
            else:
                df = None
        except Exception as e:
            safe_print(f"[DataService] header=3解析失败: {e}")

        # 方式2: 多级列名
        if df is None:
            try:
                df = pd.read_excel(excel_file, sheet_name=sheet_name, header=[0, 1, 2])
                df.columns = [
                    "_".join(str(c).strip() for c in col).strip()
                    for col in df.columns.values
                ]
                print(f"[DataService] 使用多级列名解析成功，共 {len(df)} 行")
            except Exception as e:
                print(f"[DataService] 多级列名解析失败: {e}")

        # 方式3: 默认
        if df is None:
            try:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                print(f"[DataService] 使用默认解析成功，共 {len(df)} 行")
            except Exception as e:
                print(f"[DataService] 默认解析失败: {e}")

        if df is None:
            raise ValueError(f"无法解析sheet: {sheet_name}")

        return self._parse_dataframe(df, sheet_name, network_type)

    def _parse_target_cells_sheet(
        self, excel_file, sheet_name: str, network_type: str
    ) -> List[Dict]:
        """解析待规划小区Sheet (只有ID信息)"""
        import pandas as pd

        print(f"[DataService] 解析待规划小区Sheet: {sheet_name}")

        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        sites = {}

        # 确定ID列
        site_id_col = "eNodeBID" if network_type == "LTE" else "gNodeBID"

        if site_id_col not in df.columns:
            # 尝试查找
            for col in df.columns:
                if site_id_col.lower() in col.lower():
                    site_id_col = col
                    break

        if site_id_col not in df.columns:
            raise ValueError(f"在Sheet {sheet_name} 中找不到 {site_id_col} 列")

        print(f"[DataService] 使用 {site_id_col} 作为站点ID列")

        for _, row in df.iterrows():
            try:
                if pd.isna(row[site_id_col]):
                    continue

                site_id = str(int(row[site_id_col]))
                cell_id = (
                    str(int(row["CellID"]))
                    if "CellID" in row and pd.notna(row["CellID"])
                    else "0"
                )

                # 构建唯一扇区ID
                unique_sector_id = f"{site_id}_{cell_id}"

                if site_id not in sites:
                    sites[site_id] = {
                        "id": site_id,
                        "name": f"Site_{site_id}",  # 自动生成名称
                        "longitude": 0.0,  # 默认值
                        "latitude": 0.0,  # 默认值
                        "networkType": network_type,
                        "sectors": [],
                    }

                # 添加扇区
                sector = {
                    "id": unique_sector_id,
                    "siteId": site_id,
                    "name": f"Cell_{site_id}_{cell_id}",
                    "longitude": 0.0,
                    "latitude": 0.0,
                    "azimuth": 0,
                    "beamwidth": 65,
                    "height": 30,
                    "pci": 0,
                    "earfcn": 0,
                }

                sites[site_id]["sectors"].append(sector)

            except Exception as e:
                print(f"[DataService] 解析行失败: {e}")
                continue

        return list(sites.values())

    def _parse_full_params_sheet(
        self, excel_file, sheet_name: str, network_type: str
    ) -> List[Dict]:
        """解析全量工参文件（LTE/NR Project Parameters）"""
        import pandas as pd

        print(f"[DataService] 解析全量工参sheet: {sheet_name}")

        try:
            # 直接使用传入的ExcelFile对象
            xls = excel_file

            # 读取第一行获取列名（包含\n分隔的多行信息）
            header_row = pd.read_excel(
                xls, sheet_name=sheet_name, header=None, nrows=1
            ).iloc[0]

            # 提取中文名称（第一行中\n之前的部分）
            clean_columns = []
            # 提取中文名称（第一行中\n之前的部分）并检测是否为复杂表头
            clean_columns = []
            is_complex_header = False
            for col in header_row:
                col_str = str(col).strip() if pd.notna(col) else ""
                if "\n" in col_str:
                    is_complex_header = True
                    # 提取第一个\n之前的中文名称
                    chinese_name = col_str.split("\n")[0].strip()
                    clean_columns.append(chinese_name)
                else:
                    clean_columns.append(col_str)

            print(f"[DataService] 提取到 {len(clean_columns)} 个列名 (复杂表头: {is_complex_header})")

            # 确定跳过的行数
            # 复杂表头通常占3行（第4行开始是数据）
            # 简单表头通常占1行（第2行开始是数据）
            skip_rows = 3 if is_complex_header else 1
            print(f"[DataService] 将跳过 {skip_rows} 行读取数据")

            # 读取数据
            df = pd.read_excel(xls, sheet_name=sheet_name, header=None, skiprows=skip_rows)
            
            # 检查列数是否匹配
            if len(df.columns) == len(clean_columns):
                # 设置清理后的列名
                df.columns = clean_columns
            else:
                print(f"[DataService] 警告: 数据列数({len(df.columns)})与表头列数({len(clean_columns)})不匹配")
                # 尝试重新读取，假设只有1行表头 (如果之前尝试了3行)
                if skip_rows == 3:
                     print(f"[DataService] 尝试使用 skip_rows=1 重新读取...")
                     df = pd.read_excel(xls, sheet_name=sheet_name, header=None, skiprows=1)
                     if len(df.columns) == len(clean_columns):
                         df.columns = clean_columns
                         print(f"[DataService] 重新读取成功")
                     else:
                         # 如果还是不匹配，可能是空数据或者结构严重不符
                         if len(df) == 0:
                             # 空数据，强行赋值columns可能会失败 if len(df.columns)==0
                             # 创建一个空的DataFrame with correct columns
                             df = pd.DataFrame(columns=clean_columns)
                             print(f"[DataService] 数据为空，创建空DataFrame")
                         else:
                             raise ValueError(f"列数不匹配: 期望 {len(clean_columns)}, 实际 {len(df.columns)}")
                else:
                     if len(df) == 0:
                         df = pd.DataFrame(columns=clean_columns)
                         print(f"[DataService] 数据为空，创建空DataFrame")
                     else:
                         raise ValueError(f"列数不匹配: 期望 {len(clean_columns)}, 实际 {len(df.columns)}")

            print(f"[DataService] 全量工参数据加载完成，共 {len(df)} 行")
            column_preview = list(df.columns)[:10]
            try:
                print(f"[DataService] 列名: {column_preview}...")
            except:
                pass

            # 使用专门的解析方法处理全量工参数据
            return self._parse_full_params_dataframe(df, network_type)

        except Exception as e:
            print(f"[DataService] 解析全量工参sheet失败: {sheet_name}, 错误: {e}")
            raise ValueError(f"解析工参表 {sheet_name} 失败: {str(e)}")

    def _parse_full_params_dataframe(
        self, df: pd.DataFrame, network_type: str
    ) -> List[Dict]:
        """解析全量工参DataFrame为站点数据

        使用唯一键进行小区去重:
        - LTE: eNodeB标识 + 小区标识
        - NR: 移动国家码 + 移动网络码 + gNodeB标识 + 小区标识
        """
        sites = {}

        print(f"[DataService] 解析全量工参DataFrame，网络类型: {network_type}")

        # 定义全量工参的列名映射（智能匹配）
        if network_type == "LTE":
            # LTE工参需要的列名（中文）
            required_columns = {
                "site_id": [
                    "eNodeB标识",
                    "eNodeBID",
                    "eNodeB ID",
                    "基站ID",
                ],  # 移除了'管理网元ID'，避免混淆
                "site_name": ["基站名称"],
                "longitude": ["基站经度", "经度", "小区经度"],
                "latitude": ["基站纬度", "纬度", "小区纬度"],
                "sector_id": ["小区标识", "小区ID"],
                "sector_name": ["小区名称"],
                "azimuth": ["天线方向角", "方位角"],
                "height": ["天线挂高", "挂高"],
                "pci": ["物理小区识别码", "PCI"],
                "earfcn": ["下行链路的中心载频", "EARFCN"],
                # 可选字段
                "tac": ["跟踪区码", "TAC"],
                "cell_cover_type": ["小区覆盖类型"],
                "is_shared": ["是否共享"],
                "first_group": ["第一分组"],
            }
        else:  # NR
            # NR工参需要的列名（中文）
            required_columns = {
                "site_id": [
                    "gNodeB标识",
                    "gNodeBID",
                    "gNodeB ID",
                    "基站ID",
                    "管理网元ID",
                ],
                "site_name": ["基站名称"],
                "longitude": ["基站经度", "经度", "小区经度"],
                "latitude": ["基站纬度", "纬度", "小区纬度"],
                "sector_id": ["小区标识", "小区ID"],
                "sector_name": ["小区名称"],
                "azimuth": ["天线方向角", "方位角"],
                "height": ["天线挂高", "挂高"],
                "pci": ["物理小区识别码", "PCI"],
                "ssb_frequency": ["填写SSB频点", "SSB频点", "SSB Frequency"],
                # 可选字段
                "tac": ["跟踪区码", "TAC"],
                "cell_cover_type": ["小区覆盖类型"],
                "is_shared": ["是否共享"],
                "mcc": ["移动国家码", "MCC"],
                "mnc": ["移动网络码", "MNC"],
                "first_group": ["第一分组"],
            }

        # 智能列名匹配函数
        def find_column(clean_cols, possible_names, column_type=""):
            """在清理后的列名列表中查找匹配的列，返回列索引"""
            clean_cols_lower = [str(c).strip().lower() for c in clean_cols]
            clean_cols_clean = [str(c).strip() for c in clean_cols]

            print(
                f"[DataService] {network_type}] 查找列: {possible_names}, 可用列: {clean_cols_clean}"
            )

            for possible in possible_names:
                possible_lower = possible.lower()
                possible_clean = possible.strip()

                print(
                    f"[DataService] {network_type}] 尝试匹配: {possible_clean} -> {possible_lower}"
                )

                # 精确匹配
                if possible_lower in clean_cols_lower:
                    index = clean_cols_lower.index(possible_lower)
                    print(
                        f"[DataService] {network_type}] 精确匹配成功: {possible_clean} -> {clean_cols_clean[index]}, 索引: {index}"
                    )
                    return index

                # 包含匹配
                for i, (col, col_lower) in enumerate(
                    zip(clean_cols_clean, clean_cols_lower)
                ):
                    if possible_lower in col_lower or col_lower in possible_lower:
                        print(
                            f"[DataService] {network_type}] 包含匹配成功: {possible_clean} -> {col}, 索引: {i}"
                        )
                        return i

                # 对于管理网元ID，尝试更宽松的匹配
                if column_type == "managed_element":
                    for i, (col, col_lower) in enumerate(
                        zip(clean_cols_clean, clean_cols_lower)
                    ):
                        if (
                            "网元" in col
                            or "element" in col_lower
                            or "managed" in col_lower
                            or "管理" in col
                        ):
                            print(
                                f"[DataService] {network_type}] 宽松匹配成功: {possible_clean} -> {col}, 索引: {i}"
                            )
                            return i

            print(f"[DataService] {network_type}] 未找到匹配列: {possible_names}")
            return None

        # 获取当前DataFrame的列名
        current_columns = list(df.columns)

        # 执行列名匹配
        mapped_columns = {}
        print(f"[DataService] {network_type}] 开始列名匹配，可用列: {current_columns}")
        for key, possible_names in required_columns.items():
            found_col_index = find_column(current_columns, possible_names)
            if found_col_index is not None:
                # 注意：这里返回的是列索引，而不是列名
                found_col = current_columns[found_col_index]
                mapped_columns[key] = found_col
                # 特别关注小区覆盖类型和是否共享列的映射
                if key in ["cell_cover_type", "is_shared"]:
                    try:
                        print(
                            f"[DataService] {network_type}] 列名映射: {key} -> {found_col}"
                        )
                    except UnicodeEncodeError:
                        safe_col = (
                            str(found_col).encode("ascii", "replace").decode("ascii")
                        )
                        print(
                            f"[DataService] {network_type}] 列名映射: {key} -> {safe_col}"
                        )
            else:
                if key == "cell_cover_type":
                    print(
                        f"[DataService] {network_type}] 警告: 未找到列 {key}，尝试匹配: {possible_names}"
                    )
                # 只对必需字段输出警告
                if key in ["site_id", "longitude", "latitude", "sector_id"]:
                    print(
                        f"[DataService] 警告: 未找到列 {key}，尝试匹配: {possible_names}"
                    )

        # 额外匹配管理网元ID字段
        # 注意：优先匹配"管理网元ID"和"ManagedElement ID"，避免误匹配到其他列
        managed_element_columns = [
            "管理网元ID",
            "ManagedElement ID",
            "ManagedElementId",
            "ManagedElement",
            "Managed Element",
            "MEID",
            "网元ID",  # 放在最后，避免误匹配
        ]
        managed_element_col_index = find_column(
            current_columns, managed_element_columns, column_type="managed_element"
        )
        if managed_element_col_index is not None:
            managed_element_col = current_columns[managed_element_col_index]
            mapped_columns["managed_element_id"] = managed_element_col
            print(
                f"[DataService] {network_type}] 找到管理网元ID列: {managed_element_col}"
            )
        else:
            # 尝试在所有列中查找包含"网元"或"element"的列
            print(f"[DataService] {network_type}] 尝试手动匹配管理网元ID列...")
            for i, col in enumerate(current_columns):
                if (
                    "网元" in col
                    or "element" in col.lower()
                    or "managed" in col.lower()
                    or "管理" in col
                ):
                    mapped_columns["managed_element_id"] = col
                    print(f"[DataService] {network_type}] 手动匹配管理网元ID列: {col}")
                    break

        print(f"[DataService] {network_type}] 最终映射结果: {mapped_columns}")

        # 检查必需列是否存在
        required_keys = ["site_id", "longitude", "latitude", "sector_id"]
        missing_keys = [k for k in required_keys if k not in mapped_columns]

        if missing_keys:
            raise ValueError(
                f"全量工参文件缺少必需的列: {missing_keys}。已找到的列: {list(df.columns)}"
            )

        # 开始解析数据
        parsed_count = 0
        skipped_count = 0

        for idx, row in df.iterrows():
            try:
                # 提取基站ID
                site_id_raw = row.get(mapped_columns["site_id"])
                if pd.isna(site_id_raw):
                    skipped_count += 1
                    continue

                site_id = str(site_id_raw).strip()
                if not site_id or site_id.lower() in ["nan", "none", ""]:
                    skipped_count += 1
                    continue

                # 提取基站名称
                site_name = "Unknown"
                if "site_name" in mapped_columns:
                    name_raw = row.get(mapped_columns["site_name"])
                    if pd.notna(name_raw):
                        site_name = str(name_raw).strip()

                # 提取经纬度
                longitude = 0.0
                latitude = 0.0
                lon_col = mapped_columns["longitude"]
                lat_col = mapped_columns["latitude"]

                if pd.notna(row.get(lon_col)):
                    try:
                        longitude = float(row[lon_col])
                    except:
                        pass

                if pd.notna(row.get(lat_col)):
                    try:
                        latitude = float(row[lat_col])
                    except:
                        pass

                # 移除经纬度为0的过滤条件 - 只使用唯一键去重统计小区数

                # 创建或获取站点（使用字典存储sectors，以唯一键为key实现去重）
                if site_id not in sites:
                    site_data = {
                        "id": site_id,
                        "name": site_name,
                        "longitude": longitude,
                        "latitude": latitude,
                        "networkType": network_type,
                        "sectors": {},  # 改为字典，key为唯一键，value为sector对象
                    }

                    # 提取管理网元ID（所有网络类型）
                    if "managed_element_id" in mapped_columns:
                        me_col = mapped_columns["managed_element_id"]
                        try:
                            # 直接访问列值，因为row是pandas Series对象
                            managed_element_id = row[me_col]
                            if pd.notna(managed_element_id):
                                managed_element_id_str = str(managed_element_id).strip()
                                if (
                                    managed_element_id_str
                                    and managed_element_id_str.lower()
                                    not in ["nan", "none", ""]
                                ):
                                    # 处理数值类型的管理网元ID，转换为整数
                                    try:
                                        # 尝试转换为float，再转为int，最后转为字符串
                                        me_id_float = float(managed_element_id_str)
                                        if me_id_float.is_integer():
                                            managed_element_id_str = str(
                                                int(me_id_float)
                                            )
                                    except (ValueError, TypeError):
                                        # 如果转换失败，保持原字符串
                                        pass
                                    site_data["managedElementId"] = (
                                        managed_element_id_str
                                    )
                                    print(
                                        f"[DataService] {network_type}] 站点 {site_id} 的管理网元ID: {managed_element_id_str}"
                                    )
                                else:
                                    print(
                                        f"[DataService] {network_type}] 站点 {site_id} 的管理网元ID为空或无效: {managed_element_id}"
                                    )
                            else:
                                print(
                                    f"[DataService] {network_type}] 站点 {site_id} 的管理网元ID为NaN"
                                )
                        except KeyError:
                            print(
                                f"[DataService] {network_type}] 站点 {site_id} 找不到列: {me_col}"
                            )
                        except Exception as e:
                            print(
                                f"[DataService] {network_type}] 站点 {site_id} 提取管理网元ID失败: {e}"
                            )
                    else:
                        print(
                            f"[DataService] {network_type}] 站点 {site_id} 没有映射的管理网元ID列"
                        )

                    sites[site_id] = site_data

                # 提取小区信息
                sector_id_raw = row.get(mapped_columns["sector_id"])
                if pd.notna(sector_id_raw):
                    try:
                        sector_id = str(int(float(sector_id_raw)))
                    except:
                        sector_id = str(sector_id_raw).strip()
                else:
                    # 获取当前sectors字典的大小作为默认ID
                    sector_id = f"{site_id}_{len(sites[site_id]['sectors'])}"

                # 小区名称
                sector_name = f"{site_name}_{sector_id}"
                if "sector_name" in mapped_columns:
                    sname_raw = row.get(mapped_columns["sector_name"])
                    if pd.notna(sname_raw):
                        sname_str = str(sname_raw).strip()
                        if sname_str and sname_str not in ["非必填", "nan", "None"]:
                            sector_name = sname_str

                # 方位角
                azimuth = 0.0
                if "azimuth" in mapped_columns:
                    az_raw = row.get(mapped_columns["azimuth"])
                    if pd.notna(az_raw):
                        try:
                            azimuth = float(az_raw)
                        except:
                            pass

                # 天线高度
                height = 30.0
                if "height" in mapped_columns:
                    h_raw = row.get(mapped_columns["height"])
                    if pd.notna(h_raw):
                        try:
                            height = float(h_raw)
                        except:
                            pass

                # 构建小区数据
                sector = {
                    "id": sector_id,
                    "siteId": site_id,
                    "name": sector_name,
                    "longitude": longitude,
                    "latitude": latitude,
                    "azimuth": azimuth,
                    "beamwidth": 65.0,
                    "height": height,
                }

                # 可选字段：PCI
                if "pci" in mapped_columns:
                    pci_raw = row.get(mapped_columns["pci"])
                    if pd.notna(pci_raw):
                        try:
                            pci_val = int(float(pci_raw))
                            # PCI值范围是0-503，包括0
                            sector["pci"] = pci_val
                        except Exception as e:
                            print(f"[DataService] 解析PCI失败: {pci_raw}, 错误: {e}")
                            sector["pci"] = str(pci_raw).strip()
                # 确保pci字段始终存在，即使值为None
                if "pci" not in sector:
                    sector["pci"] = None

                # 可选字段：是否共享
                if "is_shared" in mapped_columns:
                    is_shared_raw = row.get(mapped_columns["is_shared"])
                    if pd.notna(is_shared_raw):
                        sector["is_shared"] = str(is_shared_raw).strip()
                # 确保is_shared字段始终存在，即使值为None
                if "is_shared" not in sector:
                    sector["is_shared"] = None

                # 可选字段：第一分组
                if "first_group" in mapped_columns:
                    first_group_raw = row.get(mapped_columns["first_group"])
                    if pd.notna(first_group_raw):
                        sector["first_group"] = str(first_group_raw).strip()
                # 确保first_group字段始终存在，即使值为None
                if "first_group" not in sector:
                    sector["first_group"] = None

                # 可选字段：TAC（跟踪区码）
                if "tac" in mapped_columns:
                    tac_raw = row.get(mapped_columns["tac"])
                    if pd.notna(tac_raw):
                        try:
                            # 尝试直接转换为整数
                            if isinstance(tac_raw, (int, float)):
                                tac_val = int(tac_raw)
                            else:
                                # 尝试从字符串转换
                                tac_val = int(float(str(tac_raw).strip()))
                            sector["tac"] = tac_val
                        except Exception as e:
                            # 记录错误并继续，使用原始值
                            print(f"[DataService] 解析TAC失败: {tac_raw}, 错误: {e}")
                            sector["tac"] = str(tac_raw).strip()
                # 确保tac字段始终存在，即使值为None
                if "tac" not in sector:
                    sector["tac"] = None

                # 可选字段：EARFCN（LTE）或SSB频点（NR），并添加统一的下行频点字段
                if network_type == "LTE" and "earfcn" in mapped_columns:
                    earfcn_raw = row.get(mapped_columns["earfcn"])
                    if pd.notna(earfcn_raw):
                        try:
                            earfcn_val = float(earfcn_raw)
                            sector["earfcn"] = earfcn_val
                            sector["frequency"] = earfcn_val  # 添加统一的下行频点字段
                        except Exception as e:
                            print(
                                f"[DataService] 解析LTE频点失败: {earfcn_raw}, 错误: {e}"
                            )
                            # 使用原始值
                            sector["earfcn"] = str(earfcn_raw).strip()
                            sector["frequency"] = str(earfcn_raw).strip()
                elif network_type == "NR" and "ssb_frequency" in mapped_columns:
                    ssb_raw = row.get(mapped_columns["ssb_frequency"])
                    if pd.notna(ssb_raw):
                        try:
                            ssb_val = float(ssb_raw)
                            sector["ssb_frequency"] = ssb_val
                            sector["frequency"] = ssb_val  # 添加统一的下行频点字段
                        except Exception as e:
                            # 处理字符串格式的SSB频点，如"900-955"
                            print(
                                f"[DataService] 解析NR SSB频点失败: {ssb_raw}, 错误: {e}"
                            )
                            ssb_str = str(ssb_raw).strip()
                            sector["ssb_frequency"] = ssb_str
                            sector["frequency"] = ssb_str  # 添加统一的下行频点字段
                # 确保frequency字段始终存在，即使值为None
                if "frequency" not in sector:
                    sector["frequency"] = None

                # 可选字段：小区覆盖类型
                # 1=室外小区（扇形，半径60米，夹角40度，按方向角绘制）
                # 4=室内小区（圆形，半径30米，不考虑方向角）
                if "cell_cover_type" in mapped_columns:
                    cover_type_raw = row.get(mapped_columns["cell_cover_type"])
                    if pd.notna(cover_type_raw):
                        try:
                            cover_type_val = int(float(cover_type_raw))
                            sector["cell_cover_type"] = cover_type_val
                            # 调试日志：记录室内小区
                            if cover_type_val == 4 and network_type == "NR":
                                print(
                                    f"[DataService] NR室内小区: {sector.get('sector_name')} cell_cover_type=4"
                                )
                        except:
                            # 默认为室外小区
                            sector["cell_cover_type"] = 1
                    else:
                        sector["cell_cover_type"] = 1  # 默认室外
                else:
                    sector["cell_cover_type"] = 1  # 默认室外
                    if network_type == "NR":
                        print(
                            f"[DataService] NR未找到小区覆盖类型列，使用默认值1: {sector.get('sector_name')}"
                        )

                # 生成小区唯一键用于去重
                # LTE: site_id + sector_id
                # NR: mcc + mnc + site_id + sector_id
                if network_type == "LTE":
                    sector_unique_key = f"{site_id}_{sector_id}"
                else:  # NR
                    # 尝试获取MCC和MNC
                    mcc = ""
                    mnc = ""
                    if "mcc" in mapped_columns:
                        mcc_raw = row.get(mapped_columns["mcc"])
                        if pd.notna(mcc_raw):
                            mcc = str(mcc_raw).strip()
                            sector["mcc"] = mcc  # 添加到sector对象
                    if "mnc" in mapped_columns:
                        mnc_raw = row.get(mapped_columns["mnc"])
                        if pd.notna(mnc_raw):
                            mnc = str(mnc_raw).strip()
                            sector["mnc"] = mnc  # 添加到sector对象
                    sector_unique_key = f"{mcc}_{mnc}_{site_id}_{sector_id}"

                # 使用唯一键存储sector，自动去重（后面的数据覆盖前面的）
                sites[site_id]["sectors"][sector_unique_key] = sector
                parsed_count += 1

            except Exception as e:
                print(f"[DataService] 解析第{idx}行数据失败: {e}")
                skipped_count += 1
                continue

        # 将sectors字典转换为列表，保持向后兼容
        result = []
        unique_sector_count = 0
        for site_data in sites.values():
            # 将sectors从字典转换为列表
            site_data["sectors"] = list(site_data["sectors"].values())
            result.append(site_data)
            unique_sector_count += len(site_data["sectors"])

        print(f"\n[DataService] ===== 全量工参解析完成 =====")
        print(f"  共解析: {len(result)} 个基站")
        print(f"  唯一小区数: {unique_sector_count} 个 (去重后)")
        print(f"  处理总行数: {parsed_count} 行")
        print(f"  跳过无效行: {skipped_count} 行")
        for site in result[:5]:
            print(f"  站点 {site['id']}: {len(site['sectors'])} 个小区")
        if len(result) > 5:
            print(f"  ... (还有 {len(result) - 5} 个站点)")
        print(f"=====================================\n")

        return result

    def _parse_default_excel(self, excel_file) -> List[Dict]:
        """解析默认格式的Excel文件"""
        import pandas as pd

        df = None
        try:
            df = pd.read_excel(excel_file, header=3)
        except:
            try:
                df = pd.read_excel(excel_file)
            except:
                pass

        if df is None:
            raise ValueError("无法解析Excel文件")

        return self._parse_dataframe(df, "default", "LTE")

    def _parse_dataframe(
        self, df: pd.DataFrame, sheet_name: str, network_type: str
    ) -> List[Dict]:
        """解析DataFrame为站点数据"""
        print(
            f"[DataService] _parse_dataframe: sheet={sheet_name}, network={network_type}"
        )
        # 复用现有的_parse_excel_data方法，传入网络类型作为文件名标识
        return self._parse_excel_data(df, f"{sheet_name} ({network_type})")

    def _parse_excel_data(self, df: pd.DataFrame, filename: str = "") -> List[Dict]:
        """解析Excel数据 - 支持多种列名格式，包括多行列名"""
        sites = {}

        print(f"[DataService] 开始解析Excel文件: {filename}")
        print(f"[DataService] 数据形状: {df.shape}")
        try:
            print(f"[DataService] 所有列名: {list(df.columns)}")
        except UnicodeEncodeError:
            # 如果列名包含特殊字符，使用ASCII安全打印
            safe_columns = [
                str(c).encode("ascii", "replace").decode("ascii") for c in df.columns
            ]
            print(f"[DataService] 所有列名: {safe_columns}")

        # 改进的列名映射表 - 使用更灵活的模糊匹配
        column_mappings = {
            # 基站ID可能的列名（支持多级列名组合）
            "site_id": [
                "基站ID",
                "eNodeBID",
                "eNodeB ID",
                "gNodeBID",
                "gNodeB ID",
                "站点ID",
                "Site ID",
                "ENODEBID",
                "基站标识",
                "网元ID",
                "NE ID",
                # 多级列名可能的组合
                "基站_ID",
                "eNodeB_ID",
                "gNodeB_ID",
                "站点_ID",
                "Unnamed_0_基站ID",
                "Unnamed_1_基站ID",
            ],
            # 基站名称可能的列名
            "site_name": [
                "基站名称",
                "站点名称",
                "Site Name",
                "eNodeB Name",
                "gNodeB Name",
                "网元名称",
                "NE Name",
                "基站_Name",
                "站点_Name",
            ],
            # 经纬度可能的列名
            "longitude": [
                "经度",
                " Longitude",
                "小区经度",
                "Cell Longitude",
                "eNodeB Longitude",
                "基站经度",
                "经度_经度",
                "Longitude_经度",
            ],
            "latitude": [
                "纬度",
                " Latitude",
                "小区纬度",
                "Cell Latitude",
                "eNodeB Latitude",
                "基站纬度",
                "纬度_纬度",
                "Latitude_纬度",
            ],
            # 网络类型可能的列名
            "network_type": [
                "网络类型",
                "Network Type",
                "RAT",
                "RAT Type",
                "RAT_Type",
                "网络类型_网络类型",
            ],
            # 小区ID可能的列名
            "sector_id": [
                "小区ID",
                "CellID",
                "Cell ID",
                "小区标识",
                "cellLocalId",
                "Cell LocalId",
                "扇区ID",
                "Sector ID",
                "小区_ID",
                "Cell_ID",
            ],
            # 小区名称可能的列名
            "sector_name": [
                "小区名称",
                "Cell Name",
                "CELL NAME",
                "userLabel",
                "UserLabel",
                "扇区名称",
                "Sector Name",
                "小区_小区名称",
                "Cell_Name",
            ],
            # 方位角可能的列名
            "azimuth": [
                "方位角",
                "Azimuth",
                "方位角\nAzimuth",
                "机械下倾角",
                "Azimuth_Azimuth",
                "方位角_方位角",
            ],
            # 波束宽度可能的列名
            "beamwidth": [
                "波束宽度",
                "Beamwidth",
                "水平波束宽度",
                "水平半功率角",
                "波束宽度_波束宽度",
            ],
            # 天线高度可能的列名
            "height": [
                "天线高度",
                "Antenna Height",
                "Height",
                "天线挂高",
                "总高度",
                "天线高度_天线高度",
            ],
            # PCI可能的列名
            "pci": [
                "PCI",
                "物理小区识别码",
                "Physical Cell ID",
                "PCI_PCI",
                "PCI_物理小区识别码",
            ],
            # EARFCN可能的列名
            "earfcn": [
                "EARFCN",
                "earfcnDl",
                "下行链路的中心载频",
                "ARFCN",
                "SSB Frequency",
                "EARFCN_EARFCN",
                "DL EARFCN",
            ],
        }

        # 改进的模糊匹配函数
        def find_column(target_name: str) -> Optional[str]:
            """查找目标列对应的实际列名 - 使用模糊匹配"""
            possible_names = column_mappings.get(target_name, [])

            for col in df.columns:
                col_str = str(col).strip().lower()
                col_str_clean = (
                    col_str.replace(" ", "").replace("_", "").replace("\n", "")
                )

                for possible in possible_names:
                    possible_clean = (
                        possible.lower()
                        .replace(" ", "")
                        .replace("_", "")
                        .replace("\n", "")
                    )

                    # 精确匹配
                    if col_str == possible.lower():
                        return col
                    # 去除特殊字符后匹配
                    if col_str_clean == possible_clean:
                        return col
                    # 包含匹配（列名包含目标词或目标词包含列名）
                    if (
                        possible_clean in col_str_clean
                        or col_str_clean in possible_clean
                    ):
                        if len(possible_clean) > 3:  # 避免过短词误匹配
                            return col

            return None

        # 映射列名
        site_id_col = find_column("site_id")
        site_name_col = find_column("site_name")
        longitude_col = find_column("longitude")
        latitude_col = find_column("latitude")
        network_type_col = find_column("network_type")
        sector_id_col = find_column("sector_id")
        sector_name_col = find_column("sector_name")
        azimuth_col = find_column("azimuth")
        beamwidth_col = find_column("beamwidth")
        height_col = find_column("height")
        pci_col = find_column("pci")
        earfcn_col = find_column("earfcn")

        print(f"\n[DataService] ===== 列名映射结果 =====")
        print(f"  基站ID: {site_id_col}")
        print(f"  基站名称: {site_name_col}")
        print(f"  经度: {longitude_col}")
        print(f"  纬度: {latitude_col}")
        print(f"  网络类型: {network_type_col}")
        print(f"  小区ID: {sector_id_col}")
        print(f"  小区名称: {sector_name_col}")
        print(f"  方位角: {azimuth_col}")
        print(f"  波束宽度: {beamwidth_col}")
        print(f"  天线高度: {height_col}")
        print(f"  PCI: {pci_col}")
        print(f"  EARFCN: {earfcn_col}")
        print(f"==============================\n")

        # 撒点文件模式：只要有经纬度即可，不需要基站ID
        is_point_file = longitude_col is not None and latitude_col is not None

        if not is_point_file:
            raise ValueError(
                "无法找到经度/纬度列，请检查Excel文件格式。撒点文件必须包含经度和纬度列。"
            )

        print(f"[DataService] 开始解析 {len(df)} 行数据...")
        print(
            f"[DataService] 解析模式: {'撒点文件模式（只需经纬度）' if is_point_file else '工参文件模式（需基站ID）'}"
        )

        parsed_count = 0
        skipped_count = 0

        for idx, row in df.iterrows():
            try:
                # 提取经纬度（撒点文件必须有经纬度）
                longitude = 0.0
                latitude = 0.0

                if longitude_col:
                    lon_raw = row.get(longitude_col)
                    if pd.notna(lon_raw):
                        try:
                            lon_str = str(lon_raw).strip()
                            longitude = float(lon_str)
                        except:
                            try:
                                longitude = float(lon_raw)
                            except:
                                pass

                if latitude_col:
                    lat_raw = row.get(latitude_col)
                    if pd.notna(lat_raw):
                        try:
                            lat_str = str(lat_raw).strip()
                            latitude = float(lat_str)
                        except:
                            try:
                                latitude = float(lat_raw)
                            except:
                                pass

                # 跳过无效的经纬度
                if longitude == 0 and latitude == 0:
                    skipped_count += 1
                    continue

                # 撒点文件模式：没有基站ID时使用行号
                if site_id_col:
                    site_id_raw = row.get(site_id_col)
                    if pd.isna(site_id_raw):
                        site_id = f"point_{idx}"
                    else:
                        site_id = str(site_id_raw).strip()
                        # 跳过空行或标题行
                        if (
                            not site_id
                            or site_id.lower() in ["nan", "none", ""]
                            or site_id.startswith("Un")
                        ):
                            site_id = f"point_{idx}"
                else:
                    site_id = f"point_{idx}"

                # 提取基站名称（撒点文件可能有名称列）
                site_name = site_id
                if site_name_col:
                    name_raw = row.get(site_name_col)
                    if pd.notna(name_raw):
                        name_str = str(name_raw).strip()
                        if name_str and name_str.lower() not in ["nan", "none", ""]:
                            site_name = name_str

                # 提取网络类型
                network_type = "LTE"
                if network_type_col:
                    nt_raw = row.get(network_type_col)
                    if pd.notna(nt_raw):
                        nt_str = str(nt_raw).strip().upper()
                        if "NR" in nt_str or "5G" in nt_str or "5GC" in nt_str:
                            network_type = "NR"
                        else:
                            network_type = "LTE"

                # 创建或获取站点
                if site_id not in sites:
                    sites[site_id] = {
                        "id": site_id,
                        "name": site_name,
                        "longitude": longitude,
                        "latitude": latitude,
                        "networkType": network_type,
                        "sectors": [],
                    }

                # 提取小区信息
                sector_id = f"{site_id}_0"  # 默认
                if sector_id_col:
                    sid_raw = row.get(sector_id_col)
                    if pd.notna(sid_raw):
                        try:
                            # 处理科学计数法等格式
                            sector_id = str(int(float(sid_raw)))
                        except:
                            sector_id = str(sid_raw).strip()

                sector_name = f"{site_name}_{sector_id}"
                if sector_name_col:
                    sname_raw = row.get(sector_name_col)
                    if pd.notna(sname_raw):
                        sname_str = str(sname_raw).strip()
                        if sname_str and sname_str not in [
                            "非必填",
                            "UserLabel",
                            "必填",
                            "nan",
                            "None",
                        ]:
                            sector_name = sname_str

                # 收集所有其他列作为属性（用于撒点文件的标签显示）
                point_attributes = {}
                for col in df.columns:
                    if col not in [
                        site_id_col,
                        site_name_col,
                        longitude_col,
                        latitude_col,
                        network_type_col,
                        sector_id_col,
                        sector_name_col,
                        azimuth_col,
                        beamwidth_col,
                        height_col,
                        pci_col,
                        earfcn_col,
                    ]:
                        val = row.get(col)
                        if pd.notna(val):
                            val_str = str(val).strip()
                            if val_str and val_str.lower() not in ["nan", "none", ""]:
                                # 使用列名作为属性名
                                attr_name = str(col).strip()
                                point_attributes[attr_name] = val_str

                # 方位角
                azimuth = 0.0
                if azimuth_col:
                    az_raw = row.get(azimuth_col)
                    if pd.notna(az_raw):
                        try:
                            azimuth = float(az_raw)
                        except:
                            azimuth = 0.0

                # 波束宽度
                beamwidth = 65.0
                if beamwidth_col:
                    bw_raw = row.get(beamwidth_col)
                    if pd.notna(bw_raw):
                        try:
                            beamwidth = float(bw_raw)
                        except:
                            beamwidth = 65.0

                # 天线高度
                height = 30.0
                if height_col:
                    h_raw = row.get(height_col)
                    if pd.notna(h_raw):
                        try:
                            height = float(h_raw)
                        except:
                            height = 30.0

                # 构建小区数据
                sector = {
                    "id": sector_id,
                    "siteId": site_id,
                    "name": sector_name,
                    "longitude": longitude,
                    "latitude": latitude,
                    "azimuth": azimuth,
                    "beamwidth": beamwidth,
                    "height": height,
                    "networkType": network_type,
                }

                # 可选字段：PCI
                if pci_col:
                    pci_raw = row.get(pci_col)
                    if pd.notna(pci_raw):
                        try:
                            pci_val = int(float(pci_raw))
                            if 0 <= pci_val <= 503:  # PCI有效范围检查
                                sector["pci"] = pci_val
                        except:
                            pass

                # 可选字段：EARFCN
                if earfcn_col:
                    earfcn_raw = row.get(earfcn_col)
                    if pd.notna(earfcn_raw):
                        try:
                            earfcn_val = float(earfcn_raw)
                            if earfcn_val > 0:
                                sector["earfcn"] = int(earfcn_val)
                        except:
                            pass

                # 添加撒点文件的额外属性（用于标签显示）
                if point_attributes:
                    sector["attributes"] = point_attributes

                sites[site_id]["sectors"].append(sector)
                parsed_count += 1

            except Exception as e:
                print(f"[DataService] 解析第{idx}行数据失败: {e}")
                skipped_count += 1
                continue

        result = list(sites.values())
        print(f"\n[DataService] ===== 解析完成 =====")
        print(f"  共解析: {len(result)} 个基站")
        print(f"  成功解析: {parsed_count} 个小区")
        print(f"  跳过无效行: {skipped_count} 行")
        for site in result[:5]:  # 只显示前5个站点
            print(f"  站点 {site['id']}: {len(site['sectors'])} 个小区")
        if len(result) > 5:
            print(f"  ... (还有 {len(result) - 5} 个站点)")
        print(f"=====================\n")

        return result

    async def upload_map(
        self, file: Optional[Any], original_path: Optional[str] = None
    ) -> Dict[str, str]:
        """上传地图文件（支持 MapInfo 图层文件）"""
        data_id = str(uuid.uuid4())

        # 处理路径引号
        if original_path:
            original_path = original_path.strip("\"'")

        # 获取文件名和扩展名
        if file:
            filename = file.filename
            if not filename:
                raise ValueError("无法从上传的文件中获取文件名")
            ext = os.path.splitext(filename)[1]
        elif original_path:
            filename = os.path.basename(original_path)
            ext = os.path.splitext(filename)[1]
        else:
            raise ValueError("必须提供文件或文件路径")

        file_path = settings.UPLOAD_DIR / f"{data_id}{ext}"

        # 保存或复制文件
        if file:
            with open(file_path, "wb") as f:
                content = await file.read()
                f.write(content)
        elif original_path:
            if not os.path.exists(original_path):
                raise ValueError(f"文件不存在: {original_path}")
            shutil.copy2(original_path, file_path)

        # 创建数据目录
        data_dir = settings.DATA_DIR / data_id
        data_dir.mkdir(parents=True, exist_ok=True)

        # 检测文件类型
        is_mapinfo = filename.lower().endswith((".mif", ".tab", ".dat"))
        is_zip = filename.lower().endswith(".zip")

        # 处理 ZIP 文件（离线地图或 MapInfo 包）
        if is_zip:
            import zipfile

            with zipfile.ZipFile(file_path, "r") as zip_ref:
                zip_ref.extractall(data_dir)

            # 检查解压后的文件是否为 MapInfo
            extracted_files = list(data_dir.glob("*"))
            has_mapinfo = any(
                f.suffix.lower() in [".mif", ".tab"]
                for f in extracted_files
                if f.is_file()
            )

            if has_mapinfo:
                # 解析 MapInfo 图层
                from app.services.mapinfo_service import parse_mapinfo_files

                layers = parse_mapinfo_files(data_dir)

                # 保存图层元数据
                import json

                with open(data_dir / "layers.json", "w", encoding="utf-8") as f:
                    json.dump(layers, f, ensure_ascii=False, indent=2)

                print(f"[DataService] 解析到 {len(layers)} 个 MapInfo 图层")
                for layer in layers:
                    print(
                        f"  - {layer['name']} ({layer['type']}): {layer['feature_count']} 个要素"
                    )

                # 更新索引
                self.index[data_id] = {
                    "id": data_id,
                    "name": filename,
                    "type": "map",
                    "subType": "mapinfo",  # 标记为 MapInfo 图层文件
                    "size": file_path.stat().st_size,
                    "originalPath": original_path,
                    "uploadDate": datetime.now().isoformat(),
                    "status": "ready",
                    "metadata": {
                        "fileCount": len(extracted_files),
                        "layerCount": len(layers),
                        "layers": layers,
                    },
                }
            else:
                # 普通离线地图文件
                self.index[data_id] = {
                    "id": data_id,
                    "name": filename,
                    "type": "map",
                    "subType": "offline",  # 标记为离线地图
                    "size": file_path.stat().st_size,
                    "originalPath": original_path,
                    "uploadDate": datetime.now().isoformat(),
                    "status": "ready",
                    "metadata": {"fileCount": len(extracted_files)},
                }

        # 处理单个 MapInfo 文件
        elif is_mapinfo:
            # 复制文件到数据目录
            target_file = data_dir / filename
            shutil.copy(file_path, target_file)

            # 关键修复: 如果提供了原始路径(Electron/Local模式)，尝试复制关联文件
            # MapInfo .tab 文件通常依赖 .dat, .map, .id, .ind 等同名文件
            if original_path:
                try:
                    original_dir = os.path.dirname(original_path)
                    # 获取文件名(不带后缀)
                    stem = os.path.splitext(filename)[0]

                    # 常见的MapInfo关联文件后缀
                    extensions = [".dat", ".map", ".id", ".ind", ".mid"]

                    print(
                        f"[DataService] 检查关联文件: 目录={original_dir}, 基名={stem}"
                    )

                    for ext in extensions:
                        # 尝试查找同名不同后缀的文件
                        sibling_name = f"{stem}{ext}"
                        sibling_path = os.path.join(original_dir, sibling_name)

                        if os.path.exists(sibling_path):
                            target_sibling = data_dir / sibling_name
                            shutil.copy2(sibling_path, target_sibling)
                            print(f"[DataService] 已复制关联文件: {sibling_name}")
                except Exception as e:
                    print(f"[DataService] 复制关联文件失败: {e}")

            # 解析 MapInfo 图层
            from app.services.mapinfo_service import parse_mapinfo_files

            layers = parse_mapinfo_files(target_file)

            # 检查解析结果
            if not layers:
                print(f"[DataService] 警告：未解析到任何图层，文件可能格式不正确")
                print(f"[DataService] 文件路径: {target_file}")
                print(f"[DataService] 文件是否存在: {target_file.exists()}")
                if target_file.exists():
                    file_size = target_file.stat().st_size
                    print(f"[DataService] 文件大小: {file_size} 字节")
                    # 读取前几行内容检查格式
                    try:
                        with open(
                            target_file, "r", encoding="gbk", errors="ignore"
                        ) as f:
                            preview = "".join([f.readline() for _ in range(10)])
                            print(f"[DataService] 文件内容预览:\n{preview}")
                    except:
                        print(f"[DataService] 无法读取文件内容预览")

            # 保存图层元数据（即使为空也保存，便于调试）
            import json

            with open(data_dir / "layers.json", "w", encoding="utf-8") as f:
                json.dump(layers, f, ensure_ascii=False, indent=2)

            print(f"[DataService] 解析到 {len(layers)} 个 MapInfo 图层")
            for layer in layers:
                print(
                    f"  - 图层ID: {layer['id']}, 名称: {layer['name']}, 类型: {layer['type']}, 要素数: {layer['feature_count']}"
                )

            # 更新索引
            self.index[data_id] = {
                "id": data_id,
                "name": filename,
                "type": "map",
                "subType": "mapinfo",
                "size": file_path.stat().st_size,
                "originalPath": original_path,
                "uploadDate": datetime.now().isoformat(),
                "status": "ready",
                "metadata": {
                    "fileCount": 1,
                    "layerCount": len(layers),
                    "layers": layers,
                },
            }
        else:
            # 其他地图文件
            self.index[data_id] = {
                "id": data_id,
                "name": filename,
                "type": "map",
                "subType": "other",
                "size": file_path.stat().st_size,
                "originalPath": original_path,
                "uploadDate": datetime.now().isoformat(),
                "status": "ready",
                "metadata": {"fileCount": 1},
            }

        self._save_index()

        return {"id": data_id, "name": filename, "status": "ready"}

    def list_data(self, page: int = 1, page_size: int = 50) -> List[DataItem]:
        """获取数据列表（支持分页）"""
        items = []
        for data_id, data in self.index.items():
            try:
                # 创建数据副本并转换枚举类型
                data_copy = data.copy()
                # 确保type字段是枚举类型
                if isinstance(data_copy.get("type"), str):
                    if data_copy["type"] == "excel":
                        data_copy["type"] = DataType.EXCEL
                    elif data_copy["type"] == "map":
                        data_copy["type"] = DataType.MAP
                # 确保status字段是枚举类型
                if isinstance(data_copy.get("status"), str):
                    if data_copy["status"] == "ready":
                        data_copy["status"] = DataStatus.READY
                    elif data_copy["status"] == "processing":
                        data_copy["status"] = DataStatus.PROCESSING
                    elif data_copy["status"] == "error":
                        data_copy["status"] = DataStatus.ERROR
                # 确保必填字段存在
                if "size" not in data_copy:
                    data_copy["size"] = 0
                if "uploadDate" not in data_copy:
                    data_copy["uploadDate"] = datetime.now().isoformat()
                # 创建DataItem实例
                item = DataItem(**data_copy)
                items.append(item)
            except Exception as e:
                safe_print(f"[DataService] 创建DataItem失败 (id={data_id}): {e}")
                import traceback

                traceback.print_exc()
                # 跳过无效数据项，继续处理其他项
                continue

        # 按上传日期排序（最新的在前）
        items.sort(key=lambda x: x.uploadDate, reverse=True)

        # 分页逻辑
        total = len(items)
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        paginated_items = items[start_idx:end_idx]

        safe_print(f"[DataService] list_data返回 {len(paginated_items)}/{total} 个数据项 (page={page}, page_size={page_size})")
        return paginated_items

    def get_total_count(self) -> int:
        """获取数据总数"""
        return len(self.index)

    def _get_cache_key(self, data_id: str, filename: str) -> str:
        """生成缓存键"""
        return f"{data_id}:{filename}"

    def _get_file_mtime(self, file_path: Path) -> float:
        """获取文件修改时间"""
        try:
            return file_path.stat().st_mtime
        except:
            return 0

    def _get_cached_data(self, data_id: str, filename: str, file_path: Path) -> Optional[Dict]:
        """从缓存获取数据，如果缓存失效返回None"""
        cache_key = self._get_cache_key(data_id, filename)
        if cache_key in self._data_cache:
            cached = self._data_cache[cache_key]
            current_mtime = self._get_file_mtime(file_path)
            # 检查缓存是否有效（文件未修改）
            if cached.get("mtime") == current_mtime:
                safe_print(f"[DataService] 从缓存读取数据: {cache_key}")
                return cached["data"]
            else:
                safe_print(f"[DataService] 缓存失效: {cache_key}")
        return None

    def _set_cached_data(self, data_id: str, filename: str, file_path: Path, data: Dict):
        """设置缓存数据"""
        cache_key = self._get_cache_key(data_id, filename)
        self._data_cache[cache_key] = {
            "data": data,
            "mtime": self._get_file_mtime(file_path)
        }
        safe_print(f"[DataService] 数据已缓存: {cache_key}")

    def get_data(self, data_id: str) -> Optional[Dict]:
        """获取数据详情（带缓存）"""
        if data_id not in self.index:
            return None

        data_info = self.index[data_id]
        data_dir = settings.DATA_DIR / data_id

        if data_info["type"] == "excel":
            import json

            # 尝试多个可能的文件名
            possible_files = [
                "data.json",      # 标准格式（合并的 LTE+NR）
                "default.json",  # 地理化数据格式
                "LTE.json",      # LTE 分表
                "NR.json"        # NR 分表
            ]

            for filename in possible_files:
                data_file = data_dir / filename

                if data_file.exists():
                    # 尝试从缓存读取
                    cached_data = self._get_cached_data(data_id, filename, data_file)
                    if cached_data is not None:
                        return cached_data

                    # 缓存未命中，读取文件
                    with open(data_file, "r", encoding="utf-8") as f:
                        data = json.load(f)

                    # 如果是分表格式（LTE.json 或 NR.json），包装成标准格式
                    if filename in ["LTE.json", "NR.json"]:
                        result = {filename.replace(".json", "").lower(): data}
                        self._set_cached_data(data_id, filename, data_file, result)
                        return result

                    # 缓存并返回数据
                    self._set_cached_data(data_id, filename, data_file, data)
                    return data

            return None
        else:
            # 地图文件返回文件列表
            files = list(data_dir.glob("*")) if data_dir.exists() else []
            return {"files": [f.name for f in files]}

        return None

    def clear_cache(self, data_id: Optional[str] = None):
        """清除缓存"""
        if data_id is None:
            self._data_cache.clear()
            safe_print("[DataService] 所有缓存已清除")
        else:
            # 清除指定数据ID的缓存
            keys_to_remove = [k for k in self._data_cache.keys() if k.startswith(f"{data_id}:")]
            for key in keys_to_remove:
                del self._data_cache[key]
            safe_print(f"[DataService] 数据 {data_id} 的缓存已清除")

    def delete_data(self, data_id: str) -> bool:
        """删除数据"""
        if data_id not in self.index:
            return False

        # 删除文件
        data_dir = settings.DATA_DIR / data_id
        if data_dir.exists():
            shutil.rmtree(data_dir)

        # 删除上传的原始文件 - 尝试新旧两种格式
        # 新格式：{data_id}.xlsx
        upload_file = settings.UPLOAD_DIR / f"{data_id}.xlsx"
        if upload_file.exists():
            upload_file.unlink()
        else:
            # 旧格式：{data_id}_{filename}.xlsx (兼容旧数据)
            old_filename = self.index[data_id].get("name", "")
            if old_filename:
                old_upload_file = settings.UPLOAD_DIR / f"{data_id}_{old_filename}"
                if old_upload_file.exists():
                    try:
                        old_upload_file.unlink()
                    except:
                        pass  # 忽略删除失败

        # 更新索引
        del self.index[data_id]
        self._save_index()

        return True

    def preview_data(self, data_id: str, rows: int = 100) -> Optional[Dict]:
        """预览数据"""
        if data_id not in self.index:
            return None

        data_info = self.index[data_id]
        if data_info["type"] != "excel":
            return None

        data_dir = settings.DATA_DIR / data_id
        original_file = data_dir / "original.xlsx"

        if original_file.exists():
            df = pd.read_excel(original_file, nrows=rows)
            return {
                "columns": df.columns.tolist(),
                "rows": df.to_dict("records"),
                "totalRows": len(df),
            }

        return None


# 创建全局实例
data_service = DataService()
